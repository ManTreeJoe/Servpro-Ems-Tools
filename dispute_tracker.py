"""Dispute Tracker — wraps the shared `Dispute Tracker.xlsx` workbook.

Schema (header row at R4 on the "Dispute Tracker" sheet, data from R5):

    A  Received Date
    B  Claim #                  ← primary dedup key
    C  Insured / Customer       ← secondary dedup key
    D  Carrier
    E  Dispute Type             (dropdown from Options!C)
    F  Priority                 (dropdown from Options!E)
    G  Intake Source            (dropdown from Options!D — Email / XA / Other)
    H  Dispute Summary
    I  Ack Email Sent?          (dropdown from Options!F — No / Yes)
    J  Assigned Estimator       (dropdown from Options!A)
    K  Assigned Date
    L  Target Response Date     FORMULA: =IF(K?="","",WORKDAY(K?,3))
    M  Status                   (dropdown from Options!B — New / Under Review / Response Sent / Closed)
    N  Next Follow-Up Date
    O  Resolution Date
    P  Outcome                  (dropdown from Options!G)
    Q  Amount in Dispute
    R  Notes

Public API:
    path()                                       — configured workbook path
    read_rows()                                  — list of row dicts (with row_number)
    read_options()                               — {column_key: [allowed_values]}
    upsert(payload, *, source=None)              — (was_new, row_number)
    update_row(row_number, updates)              — bool
    delete_row(row_number)                       — bool
    upsert_from_apa(insured, claim, carrier, ...)
    upsert_from_email(insured, claim, carrier, ..., subject)

File-lock handling: when Excel has the workbook open, writes go to a
`.pending_disputes/` sidecar JSON queue next to the workbook. The next
write attempt drains the queue first, so an open Excel just means the
row gets applied seconds later instead of being lost.

Single source of truth for canonicalization is `_canon_key`, used by
every upsert path so a dispute first surfaced from APA Monitor and
then re-surfaced from an inbox keyword match lands on the same row.
"""
from __future__ import annotations

import json
import os
import re
import time
import uuid
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

import config


# ── Config / path resolution ───────────────────────────────────────────────

_DEFAULT_PATH = (
    r"X:\IE_Public\Front Operation\EMS Admin\Dispute Tracker.xlsx"
)

# Sheet names from the template. Keep stable — user may add custom
# sheets, but writes go here.
SHEET_DATA    = "Dispute Tracker"
SHEET_OPTIONS = "Options"

# Row coordinates within the data sheet.
HEADER_ROW = 4    # column headers
DATA_START = 5    # first data row

# Column layout — keys MUST match canonicalized header text (lower, no
# punctuation) so dynamic header parsing routes correctly even if the
# user re-orders columns in the workbook later.
COL_RECEIVED      = "received_date"
COL_CLAIM         = "claim"
COL_INSURED       = "insured"
COL_CARRIER       = "carrier"
COL_TYPE          = "dispute_type"
COL_PRIORITY      = "priority"
COL_INTAKE        = "intake_source"
COL_SUMMARY       = "dispute_summary"
COL_ACK           = "ack_email_sent"
COL_ESTIMATOR     = "assigned_estimator"
COL_ASSIGNED_DATE = "assigned_date"
COL_TARGET_DATE   = "target_response_date"
COL_STATUS        = "status"
COL_NEXT_FOLLOWUP = "next_follow_up_date"
COL_RESOLUTION    = "resolution_date"
COL_OUTCOME       = "outcome"
COL_AMOUNT        = "amount_in_dispute"
COL_NOTES         = "notes"

# Stable header → key map. Used to translate workbook headers (which the
# user might tweak punctuation/casing on) back to a canonical key.
# Header strings keyed by canon(text) for resilience to "Claim #" vs
# "Claim#" vs "claim_number" punctuation drift.
_HEADER_ALIASES: dict[str, str] = {}


def _canon_header(s: str) -> str:
    """Lowercase + strip non-alnum so 'Claim #' / 'Claim#' / 'CLAIM' all
    collapse to the same key. Lets header tolerance not depend on the
    user's exact punctuation."""
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


def _build_header_aliases():
    """Map every variant the workbook might use → canonical key."""
    pairs = [
        (COL_RECEIVED,      ("Received Date",)),
        (COL_CLAIM,         ("Claim #", "Claim#", "Claim Number")),
        (COL_INSURED,       ("Insured / Customer", "Insured", "Customer",
                              "Insured/Customer")),
        (COL_CARRIER,       ("Carrier",)),
        (COL_TYPE,          ("Dispute Type",)),
        (COL_PRIORITY,      ("Priority",)),
        (COL_INTAKE,        ("Intake Source",)),
        (COL_SUMMARY,       ("Dispute Summary", "Summary")),
        (COL_ACK,           ("Ack Email Sent?", "Ack Email Sent",
                              "Acknowledged?")),
        (COL_ESTIMATOR,     ("Assigned Estimator", "Estimator")),
        (COL_ASSIGNED_DATE, ("Assigned Date",)),
        (COL_TARGET_DATE,   ("Target Response Date",
                              "Target Response",)),
        (COL_STATUS,        ("Status",)),
        (COL_NEXT_FOLLOWUP, ("Next Follow-Up Date", "Next Follow Up Date",
                              "Next Followup Date")),
        (COL_RESOLUTION,    ("Resolution Date",)),
        (COL_OUTCOME,       ("Outcome",)),
        (COL_AMOUNT,        ("Amount in Dispute", "Amount")),
        (COL_NOTES,         ("Notes",)),
    ]
    for key, variants in pairs:
        for v in variants:
            _HEADER_ALIASES[_canon_header(v)] = key


_build_header_aliases()


# Column order used when CREATING a fresh workbook from scratch (only
# fires if the user nukes the file). When the file already exists, the
# workbook's actual header row drives column lookups via _read_headers.
DEFAULT_COLUMN_ORDER = (
    COL_RECEIVED, COL_CLAIM, COL_INSURED, COL_CARRIER, COL_TYPE,
    COL_PRIORITY, COL_INTAKE, COL_SUMMARY, COL_ACK, COL_ESTIMATOR,
    COL_ASSIGNED_DATE, COL_TARGET_DATE, COL_STATUS, COL_NEXT_FOLLOWUP,
    COL_RESOLUTION, COL_OUTCOME, COL_AMOUNT, COL_NOTES,
)


def _franchise_default_path() -> str:
    """Where a franchise with no tracker of its own keeps one.

    The base franchise gets the shared X: workbook. Everyone else gets a
    workbook beside their OWN job folders - blank must not mean "open
    IE's", because that is one file two offices both write to."""
    try:
        cfg = config.load()
        if config.base_department() in ("", config.active_department()):
            return _DEFAULT_PATH
        root = (cfg.get("audit_base") or "").strip()
    except Exception:
        return _DEFAULT_PATH
    if not root:
        return _DEFAULT_PATH
    return os.path.join(root, "Dispute Tracker.xlsx")


def path() -> str:
    """Resolve the workbook path: config first, else this franchise's own
    default. Override via `dispute_tracker_path` in config.json."""
    try:
        cfg = config.load()
    except Exception:
        cfg = {}
    p = (cfg.get("dispute_tracker_path") or "").strip()
    return p or _franchise_default_path()


def configured_board_link() -> str:
    """The disputes board to sync FROM, or "" when this franchise has
    none yet. Blank turns sync off rather than pulling another
    franchise's cards into this tracker."""
    try:
        cfg = config.load()
    except Exception:
        return DISPUTES_BOARD_SHORT_LINK
    if "disputes_board_short_link" in cfg:
        return (cfg.get("disputes_board_short_link") or "").strip()
    # Never configured at all: a single-office install keeps the board it
    # has always used.
    try:
        if config.base_department() in ("", config.active_department()):
            return DISPUTES_BOARD_SHORT_LINK
    except Exception:
        pass
    return DISPUTES_BOARD_SHORT_LINK


def set_path(p: str) -> None:
    """Persist a new workbook path. Lets the user move the file (e.g.
    once Zac's copy is mirrored to X:) without a code edit.

    Writes to the ACTIVE FRANCHISE's profile, not the flat base. It used
    to save the overlaid config wholesale, which copied every one of that
    franchise's overrides into the base - so the next office to open the
    panel inherited them."""
    p = (p or "").strip()
    try:
        base = config.load_base()
    except Exception:
        return
    dept = ""
    try:
        dept = config.active_department()
    except Exception:
        pass
    if dept and isinstance(base.get("departments"), dict)             and dept in base["departments"]:
        base["departments"][dept]["dispute_tracker_path"] = p
    else:
        base["dispute_tracker_path"] = p or _DEFAULT_PATH
    try:
        config.save(base)
    except Exception:
        pass


def _pending_dir() -> str:
    return os.path.join(os.path.dirname(path()) or ".",
                         ".pending_disputes")


# ── Canonicalization ───────────────────────────────────────────────────────

def _canon_text(s) -> str:
    """Whitespace-collapse + casefold. Used as a building block — does
    NOT strip punctuation, because some canon paths (insured names) need
    to keep ampersands / hyphens distinguishable."""
    if s is None:
        return ""
    return " ".join(str(s).split()).strip().lower()


def _canon_claim(s) -> str:
    """Digits-only claim # so '17722059' / '017722059' / '0177-22059'
    all match. Insurance carriers wildly inconsistent on hyphens and
    leading zeros — strip both."""
    if s is None:
        return ""
    digits = re.sub(r"\D+", "", str(s))
    return digits.lstrip("0")


def _canon_insured(s) -> str:
    """Insured name canon. Handles comma-swap so 'Smith, John' ==
    'John Smith', collapses whitespace, casefolds, strips trailing
    carrier suffix ' - <Carrier>' so an APA row 'Acuna, David -
    Mercury' matches the bare 'Acuna David' that surfaces from a
    plain folder name."""
    s = _canon_text(s)
    if not s:
        return ""
    # Strip trailing " - <suffix>" (carrier or similar).
    s = re.sub(r"\s+-\s+.+$", "", s).strip()
    if "," in s:
        parts = [p.strip() for p in s.split(",", 1)]
        if len(parts) == 2 and parts[0] and parts[1]:
            s = f"{parts[1]} {parts[0]}"
    s = re.sub(r"[^a-z0-9& ]+", "", s)
    return " ".join(s.split())


def _canon_key(claim, insured) -> str:
    """Single dedup key. Prefers claim# (most specific); falls back to
    insured-canon when claim is missing. Empty when both are missing —
    callers should skip the upsert in that case."""
    c = _canon_claim(claim)
    if c:
        return f"claim:{c}"
    i = _canon_insured(insured)
    if i:
        return f"name:{i}"
    return ""


# ── Workbook I/O ───────────────────────────────────────────────────────────

def _ensure_workbook():
    """Open the workbook, creating it (minimal scaffold) if missing.
    Returns (workbook, path). Caller is responsible for saving."""
    p = path()
    if os.path.isfile(p):
        try:
            return openpyxl.load_workbook(p), p
        except Exception:
            # Corrupted — back it up so the user doesn't lose work, then
            # fall through to fresh-workbook creation.
            backup = p + f".corrupt-{int(time.time())}"
            try:
                os.replace(p, backup)
            except OSError:
                pass

    try:
        os.makedirs(os.path.dirname(p) or ".", exist_ok=True)
    except OSError:
        pass
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    ws = wb.create_sheet(SHEET_DATA)
    # Title rows mimic the template (rows 1+2 are merged title/blurb)
    ws.cell(1, 1, "Dispute / Claim Issue Tracker")
    ws.cell(2, 1, "")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=18)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=18)
    # Header row
    headers_for_default = {
        COL_RECEIVED: "Received Date", COL_CLAIM: "Claim #",
        COL_INSURED: "Insured / Customer", COL_CARRIER: "Carrier",
        COL_TYPE: "Dispute Type", COL_PRIORITY: "Priority",
        COL_INTAKE: "Intake Source", COL_SUMMARY: "Dispute Summary",
        COL_ACK: "Ack Email Sent?", COL_ESTIMATOR: "Assigned Estimator",
        COL_ASSIGNED_DATE: "Assigned Date",
        COL_TARGET_DATE: "Target Response Date",
        COL_STATUS: "Status",
        COL_NEXT_FOLLOWUP: "Next Follow-Up Date",
        COL_RESOLUTION: "Resolution Date", COL_OUTCOME: "Outcome",
        COL_AMOUNT: "Amount in Dispute", COL_NOTES: "Notes",
    }
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDEBF7")
    for i, key in enumerate(DEFAULT_COLUMN_ORDER):
        c = ws.cell(HEADER_ROW, i + 1, headers_for_default[key])
        c.font = bold
        c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True)
    # Options sheet — seed the default dropdowns the user already had.
    opts = wb.create_sheet(SHEET_OPTIONS)
    opts_rows = [
        ("Estimators", "Statuses", "Dispute Types", "Intake Sources",
         "Priorities", "Ack Options", "Outcomes"),
        ("Aaron", "New", "Scope", "Email", "Low", "No", "Pending"),
        ("Juan", "Under Review", "Coverage", "XA", "Medium", "Yes",
         "Approved"),
        ("Johnny", "Response Sent", "Other", "Other", "High", "",
         "Denied"),
        ("Kim", "Closed", "", "", "Urgent", "", "Withdrawn"),
        ("Nathan", "", "", "", "", "", "Escalated"),
    ]
    for r, row in enumerate(opts_rows, start=1):
        for c, val in enumerate(row, start=1):
            opts.cell(r, c, val)
    return wb, p


def _read_headers(ws) -> dict[str, int]:
    """Map canonical column key → 1-based column index by reading the
    workbook's actual header row. Lets the panel survive minor user
    edits to column ordering."""
    out: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(HEADER_ROW, col).value
        key = _HEADER_ALIASES.get(_canon_header(raw))
        if key and key not in out:
            out[key] = col
    return out


def _cell_value(ws, row, col):
    """Read a cell value, normalizing date/datetime to ISO strings so
    callers don't have to special-case openpyxl's mixed types."""
    v = ws.cell(row, col).value
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return v


def read_rows() -> list[dict]:
    """Return every data row (R5 onward) as a list of dicts. Empty rows
    (no claim, no insured, no summary) are skipped. Each dict carries
    a `row_number` so the GUI can issue per-row updates without
    re-querying."""
    p = path()
    if not os.path.isfile(p):
        return []
    try:
        wb = openpyxl.load_workbook(p, data_only=True)
    except Exception:
        return []
    try:
        if SHEET_DATA not in wb.sheetnames:
            return []
        ws = wb[SHEET_DATA]
        headers = _read_headers(ws)
        out: list[dict] = []
        for r in range(DATA_START, ws.max_row + 1):
            row: dict[str, object] = {"row_number": r}
            any_filled = False
            for key, col in headers.items():
                val = _cell_value(ws, r, col)
                if val not in (None, ""):
                    any_filled = True
                row[key] = val
            # An empty row stays empty — skip it. The Target formula
            # cell alone (column L) doesn't count, because its IF guard
            # returns blank when K is empty.
            meaningful = (row.get(COL_CLAIM) or row.get(COL_INSURED)
                           or row.get(COL_SUMMARY))
            if meaningful or any_filled and (
                    row.get(COL_RECEIVED) or row.get(COL_STATUS)):
                out.append(row)
        return out
    finally:
        # openpyxl leaks substantial memory until close() — every
        # background-thread scan was accumulating ~5MB per call.
        try: wb.close()
        except Exception: pass


def read_options() -> dict[str, list[str]]:
    """Return the dropdown lists from the Options sheet, keyed by the
    canonical column key they apply to. Re-read on each call so the
    user can edit Options in Excel and have the GUI pick it up after
    a refresh — no restart needed."""
    p = path()
    if not os.path.isfile(p):
        return _DEFAULT_OPTIONS.copy()
    try:
        wb = openpyxl.load_workbook(p, data_only=True)
    except Exception:
        return _DEFAULT_OPTIONS.copy()
    try:
        if SHEET_OPTIONS not in wb.sheetnames:
            return _DEFAULT_OPTIONS.copy()
        ws = wb[SHEET_OPTIONS]
        # Header → column key. Workbook uses friendly headers like
        # "Estimators"; map to the data column they populate.
        header_to_key = {
            "estimators":     COL_ESTIMATOR,
            "statuses":       COL_STATUS,
            "dispute types":  COL_TYPE,
            "intake sources": COL_INTAKE,
            "priorities":     COL_PRIORITY,
            "ack options":    COL_ACK,
            "outcomes":       COL_OUTCOME,
        }
        options: dict[str, list[str]] = {}
        for col in range(1, ws.max_column + 1):
            header_raw = ws.cell(1, col).value
            header_norm = (header_raw or "").strip().lower()
            key = header_to_key.get(header_norm)
            if not key:
                continue
            vals: list[str] = []
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, col).value
                if v is None:
                    continue
                sv = str(v).strip()
                if sv and sv not in vals:
                    vals.append(sv)
            if vals:
                options[key] = vals
        # Fall back to defaults for any column the user removed.
        for k, default in _DEFAULT_OPTIONS.items():
            options.setdefault(k, list(default))
        return options
    finally:
        try: wb.close()
        except Exception: pass


_DEFAULT_OPTIONS: dict[str, list[str]] = {
    COL_ESTIMATOR: ["Aaron", "Juan", "Johnny", "Kim", "Nathan"],
    COL_STATUS:    ["New", "Under Review", "Response Sent", "Closed"],
    COL_TYPE:      ["Inquiry", "Scope", "Coverage", "Other"],
    COL_INTAKE:    ["Email", "XA", "Other"],
    COL_PRIORITY:  ["Low", "Medium", "High", "Urgent"],
    COL_ACK:       ["No", "Yes"],
    COL_OUTCOME:   ["Pending", "Approved", "Denied", "Withdrawn",
                     "Escalated"],
}


# ── Write path (with locked-file fallback) ────────────────────────────────

def _is_locked(p: str) -> bool:
    """Cheap heuristic: try opening the file for append. Excel holds an
    exclusive write lock when the workbook is open. Read-only open
    succeeds either way."""
    if not os.path.isfile(p):
        return False
    try:
        with open(p, "a+b"):
            return False
    except (OSError, PermissionError):
        return True


def _queue_pending(action: str, payload: dict) -> None:
    """Drop an action to disk so it survives a wedge. Drained on the
    next save attempt. Pending entries are JSON for easy inspection."""
    pdir = _pending_dir()
    try:
        os.makedirs(pdir, exist_ok=True)
    except OSError:
        return
    entry = {
        "id":        uuid.uuid4().hex,
        "action":    action,
        "payload":   payload,
        "queued_at": datetime.now().isoformat(timespec="seconds"),
    }
    try:
        with open(os.path.join(pdir, f"{entry['id']}.json"),
                   "w", encoding="utf-8") as fh:
            json.dump(entry, fh, indent=2)
    except OSError:
        pass


def pending_count() -> int:
    """How many writes are waiting for Excel to release the file. The
    GUI shows this in the status bar so the user knows changes aren't
    landing yet."""
    pdir = _pending_dir()
    if not os.path.isdir(pdir):
        return 0
    try:
        return sum(1 for n in os.listdir(pdir) if n.endswith(".json"))
    except OSError:
        return 0


def _drain_pending(wb, ws, headers):
    """Replay any queued actions onto an open workbook in queued order.
    Returns the number drained so the GUI can toast `N applied`. Each
    file is deleted after its action runs (success or schema error —
    we don't want stale entries to keep retrying forever)."""
    pdir = _pending_dir()
    if not os.path.isdir(pdir):
        return 0
    files = sorted(n for n in os.listdir(pdir) if n.endswith(".json"))
    drained = 0
    for fname in files:
        fp = os.path.join(pdir, fname)
        try:
            with open(fp, encoding="utf-8") as fh:
                entry = json.load(fh)
        except (OSError, ValueError):
            try:
                os.remove(fp)
            except OSError:
                pass
            continue
        action = entry.get("action") or ""
        payload = entry.get("payload") or {}
        try:
            if action == "upsert":
                _apply_upsert(
                    ws, headers, payload,
                    merge_only_empty=bool(
                        payload.pop("_merge_only_empty", False)))
            elif action == "update_row":
                _apply_update_row(ws, headers, payload)
            elif action == "delete_row":
                _apply_delete_row(ws, headers, payload)
        except Exception as _replay_ex:
            # Replay errors shouldn't wedge the drain — but they also
            # shouldn't silently lose the queued write. Log so we can
            # debug why a write the user thought succeeded never made
            # it to the workbook. The .json sidecar file is preserved
            # on error so the user can recover by hand.
            try:
                import ems_log
                ems_log.error(
                    "dispute_tracker",
                    f"_drain_pending: {action} replay failed for "
                    f"{fname}: {_replay_ex!r} — sidecar preserved")
            except Exception:
                pass
            drained += 1
            continue
        try:
            os.remove(fp)
        except OSError:
            pass
        drained += 1
    return drained


def _save_workbook(wb, p: str) -> bool:
    """Atomic-ish save. Excel rejects half-written .xlsx files, so we
    write to <p>.tmp and os.replace into place — the worst-case outcome
    is the old file stays intact when the disk is full."""
    tmp = p + ".tmp-" + uuid.uuid4().hex[:8]
    try:
        wb.save(tmp)
    except Exception:
        try:
            if os.path.isfile(tmp):
                os.remove(tmp)
        except OSError:
            pass
        return False
    try:
        os.replace(tmp, p)
    except OSError:
        try:
            os.remove(tmp)
        except OSError:
            pass
        return False
    return True


def _format_date(val) -> object:
    """Normalize a date-ish input to a Python `date` for Excel — keeps
    Excel's date formatting working. Pass-through for empty strings so
    blank cells stay blank instead of becoming 1899-12-30 (Excel epoch
    surprise)."""
    if val in (None, "", " "):
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    # ISO YYYY-MM-DD or YYYY/MM/DD
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m-%d-%Y",
                 "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return s  # last-ditch — store as string so user can fix manually


# Date columns that should be coerced to real Excel dates.
_DATE_KEYS = {
    COL_RECEIVED, COL_ASSIGNED_DATE, COL_NEXT_FOLLOWUP,
    COL_RESOLUTION,
}


def _coerce_cell_value(key: str, val):
    """Per-column coercion before writing into the worksheet."""
    if key in _DATE_KEYS:
        return _format_date(val)
    if key == COL_AMOUNT:
        if val in (None, "", " "):
            return None
        try:
            # Strip $ and commas the user might paste in
            return float(re.sub(r"[$,\s]", "", str(val)))
        except (TypeError, ValueError):
            return val
    if val is None:
        return None
    return str(val)


def _write_target_formula(ws, row: int, col_target: int,
                            col_assigned: int):
    """Drop the workday-3 formula into the Target Response Date cell
    for the given row. Matches the seeded formula in the template."""
    col_letter = openpyxl.utils.get_column_letter(col_assigned)
    formula = (f'=IF({col_letter}{row}="","",'
               f'WORKDAY({col_letter}{row},3))')
    ws.cell(row, col_target, formula)


def _find_row_by_key(ws, headers, key: str) -> int:
    """Locate an existing row whose claim/insured canon matches `key`.
    Returns the 1-based row number, or 0 when no match."""
    col_claim = headers.get(COL_CLAIM)
    col_insured = headers.get(COL_INSURED)
    if not col_claim and not col_insured:
        return 0
    for r in range(DATA_START, ws.max_row + 1):
        cv = ws.cell(r, col_claim).value if col_claim else None
        iv = ws.cell(r, col_insured).value if col_insured else None
        if _canon_key(cv, iv) == key:
            return r
    return 0


def _first_empty_row(ws, headers) -> int:
    """Lowest row >= DATA_START with no claim/insured/summary."""
    col_claim   = headers.get(COL_CLAIM)
    col_insured = headers.get(COL_INSURED)
    col_summary = headers.get(COL_SUMMARY)
    r = DATA_START
    while r <= ws.max_row:
        bits = []
        for c in (col_claim, col_insured, col_summary):
            if c:
                v = ws.cell(r, c).value
                if v not in (None, ""):
                    bits.append(v)
        if not bits:
            return r
        r += 1
    return r  # one past max_row


def _apply_upsert(ws, headers, payload: dict,
                   merge_only_empty: bool = False) -> tuple[bool, int]:
    """Workbook-mutating side of upsert. Returns (was_new, row_number).
    Pure mutation — caller saves the workbook.

    `merge_only_empty=True` switches the existing-row branch from
    overwrite to merge-fill: a payload value only lands in the cell
    when the cell is currently empty (None or "" or whitespace). New
    rows always write every field — the merge guard only fires on
    re-touches of existing rows so user-edited cells stay intact.
    Used by the Trello sync so a re-pull with fresh Carrier / Summary
    data fills the blanks the spreadsheet entered the workbook with,
    without trampling a user-edited Estimator / Status / Priority etc.
    """
    key = _canon_key(payload.get(COL_CLAIM), payload.get(COL_INSURED))
    if not key:
        return False, 0
    row = _find_row_by_key(ws, headers, key)
    is_new = False
    if not row:
        row = _first_empty_row(ws, headers)
        is_new = True
        # Seed default Received Date + Status on first insertion.
        payload.setdefault(COL_RECEIVED, date.today().isoformat())
        payload.setdefault(COL_STATUS, "New")
        payload.setdefault(COL_ACK, "No")
    for key_name, value in payload.items():
        col = headers.get(key_name)
        if not col:
            continue
        # Merge guard: on an existing row, only write into cells the
        # user hasn't already populated. Without this, the Trello
        # re-pull stomped over manual edits — which is exactly why the
        # caller used to short-circuit with `_exists_for` and lose all
        # the new info instead. Now we keep the new info AND the edits.
        if merge_only_empty and not is_new:
            cur = ws.cell(row, col).value
            if cur not in (None, "") and not (
                    isinstance(cur, str) and not cur.strip()):
                continue
        coerced = _coerce_cell_value(key_name, value)
        # Attribute assignment, not the (row, col, value) form, so
        # `coerced is None` actually clears the cell instead of being
        # silently ignored by openpyxl.
        ws.cell(row, col).value = coerced
    # Always (re)write the Target formula — picks up Assigned Date
    # changes immediately, and survives on newly-created rows that the
    # template didn't pre-fill the formula for.
    col_target   = headers.get(COL_TARGET_DATE)
    col_assigned = headers.get(COL_ASSIGNED_DATE)
    if col_target and col_assigned:
        _write_target_formula(ws, row, col_target, col_assigned)
    return is_new, row


def _apply_update_row(ws, headers, payload: dict) -> bool:
    """Update a specific row by row_number. Used by the GUI when the
    user edits a single cell."""
    row = int(payload.get("row_number") or 0)
    if row < DATA_START:
        return False
    updates = payload.get("updates") or {}
    for key_name, value in updates.items():
        col = headers.get(key_name)
        if not col:
            continue
        # Attribute form so clearing (passing "" → None) actually
        # blanks the cell.
        ws.cell(row, col).value = _coerce_cell_value(key_name, value)
    col_target   = headers.get(COL_TARGET_DATE)
    col_assigned = headers.get(COL_ASSIGNED_DATE)
    if col_target and col_assigned:
        _write_target_formula(ws, row, col_target, col_assigned)
    return True


def _apply_delete_row(ws, headers, payload: dict) -> bool:
    """Blank a row's data cells (does NOT shift rows up — that would
    move the Target formulas around and confuse the user). Lets the
    user fully delete from Excel manually if they want.

    Note: openpyxl's `ws.cell(row, col, value=None)` is a no-op (the
    None positional is treated as "don't change" rather than "clear").
    Must use attribute assignment to actually blank the cell."""
    row = int(payload.get("row_number") or 0)
    if row < DATA_START:
        return False
    for col in headers.values():
        ws.cell(row, col).value = None
    return True


def _do_write(action: str, payload: dict) -> tuple[bool, int]:
    """Execute action against the workbook. Queues to pending on lock.
    Returns (success, row_number_when_applicable).

    For upsert, the returned `row_number` carries the inserted/updated
    row when the write succeeded synchronously; it's 0 when queued (we
    don't know the destination row until drain time)."""
    p = path()
    if _is_locked(p):
        _queue_pending(action, payload)
        return False, 0
    try:
        wb, p = _ensure_workbook()
    except Exception:
        _queue_pending(action, payload)
        return False, 0
    try:
        if SHEET_DATA not in wb.sheetnames:
            # Brand-new workbook — _ensure_workbook seeds it, but
            # defend against partial template upload (someone saved
            # without the data sheet).
            wb.create_sheet(SHEET_DATA)
        ws = wb[SHEET_DATA]
        headers = _read_headers(ws)
        # Drain any pending writes BEFORE the new one so order matches
        # the user's click order.
        _drain_pending(wb, ws, headers)
        row_no = 0
        ok = False
        if action == "upsert":
            is_new, row_no = _apply_upsert(
                ws, headers, payload,
                merge_only_empty=bool(payload.pop("_merge_only_empty", False)))
            ok = bool(row_no)
            if not ok:
                return False, 0
        elif action == "update_row":
            ok = _apply_update_row(ws, headers, payload)
            row_no = int(payload.get("row_number") or 0)
        elif action == "delete_row":
            ok = _apply_delete_row(ws, headers, payload)
            row_no = int(payload.get("row_number") or 0)
        else:
            return False, 0
        if not _save_workbook(wb, p):
            _queue_pending(action, payload)
            return False, 0
        return ok, row_no
    finally:
        # Close the openpyxl handle regardless of save outcome — the
        # underlying zip object holds onto a lot of memory until close.
        try: wb.close()
        except Exception: pass


def process_pending() -> int:
    """Drain any queued pending writes against the workbook. Returns
    the count drained. Safe to call when nothing's queued — quick
    no-op."""
    if pending_count() == 0:
        return 0
    p = path()
    if _is_locked(p):
        return 0
    try:
        wb, p = _ensure_workbook()
    except Exception:
        return 0
    try:
        if SHEET_DATA not in wb.sheetnames:
            return 0
        ws = wb[SHEET_DATA]
        headers = _read_headers(ws)
        drained = _drain_pending(wb, ws, headers)
        if drained:
            _save_workbook(wb, p)
        return drained
    finally:
        try: wb.close()
        except Exception: pass


# ── Public upsert / update helpers ────────────────────────────────────────

def upsert(payload: dict, *, source: str = "",
           merge_only_empty: bool = False) -> tuple[bool, int]:
    """Insert a new dispute row, or update fields on the existing row
    when one matches by (claim#, insured) canon.

    `payload` keys must use the `COL_*` constants. Any keys the
    workbook doesn't have headers for are silently dropped (lets
    callers pass extras without breaking).

    `source`, when set, becomes the COL_INTAKE value on insert (only —
    we don't re-overwrite Intake Source on a re-upsert, that's
    user-edited at that point).

    `merge_only_empty=True` enables the safe-merge mode: when the
    upsert finds an existing matching row, the new payload only fills
    cells that are currently empty. User-edited cells stay untouched.
    Used by the Trello sync so re-pulls combine with existing rows
    rather than overwriting hand-curated Estimator / Priority etc.

    Returns `(was_new, row_number)`. On a locked-file deferral the
    return is `(False, 0)` — the entry IS persisted to the pending
    queue, it just hasn't landed in the workbook yet."""
    if not payload.get(COL_CLAIM) and not payload.get(COL_INSURED):
        return False, 0
    if source:
        payload.setdefault(COL_INTAKE, source)
    # Check for an existing match before delegating, so we know whether
    # this is a real insert vs a re-touch. _do_write only reports the
    # row number, not the new/existing flag.
    p = path()
    pre_exists = False
    if os.path.isfile(p) and not _is_locked(p):
        try:
            wb_check = openpyxl.load_workbook(p, data_only=True)
            if SHEET_DATA in wb_check.sheetnames:
                ws_check = wb_check[SHEET_DATA]
                headers_check = _read_headers(ws_check)
                key = _canon_key(payload.get(COL_CLAIM),
                                  payload.get(COL_INSURED))
                if key and _find_row_by_key(ws_check, headers_check,
                                              key):
                    pre_exists = True
        except Exception:
            pass
    if merge_only_empty:
        # Travel the flag through the payload — it survives the lock-
        # deferral path (queued + replayed) without changing _do_write's
        # signature. _apply_upsert pops it before normal processing.
        payload["_merge_only_empty"] = True
    ok, row_no = _do_write("upsert", payload)
    return (not pre_exists) and ok, row_no


def update_row(row_number: int, updates: dict) -> bool:
    """Per-cell edits from the GUI. `updates` maps COL_* → new value."""
    if not row_number or not updates:
        return False
    ok, _ = _do_write("update_row",
                       {"row_number": row_number, "updates": updates})
    return ok


def delete_row(row_number: int) -> bool:
    """Blank the row's data cells. Row shape stays so Target formulas
    in adjacent rows aren't disturbed — user can fully delete from
    Excel manually if they prefer."""
    if not row_number:
        return False
    ok, _ = _do_write("delete_row", {"row_number": row_number})
    return ok


# ── Source-specific convenience wrappers ──────────────────────────────────

def _exists_for(claim: str, insured: str) -> bool:
    """True when a row already matches by (claim, insured) canon. Used
    by the first-seen wrappers so we don't keep overwriting user edits
    on subsequent re-syncs."""
    key = _canon_key(claim, insured)
    if not key:
        return False
    for r in read_rows():
        if _canon_key(r.get(COL_CLAIM), r.get(COL_INSURED)) == key:
            return True
    return False


def upsert_from_apa(*, insured: str = "", claim: str = "",
                     carrier: str = "", summary: str = "",
                     estimator: str = "") -> tuple[bool, int]:
    """APA Monitor first-seen path. Called when an Audit Dispute row is
    parsed for a job not already in the tracker. Intake Source = XA.

    Insert-if-missing semantics: once a row exists, subsequent calls
    are no-ops — the user's edits to Summary/Status/etc. won't get
    overwritten by a re-sync from APA. Returns (was_new, row_number);
    `was_new=False, row_number=0` when the row already existed and
    nothing was written."""
    if not insured and not claim:
        return False, 0
    if _exists_for(claim, insured):
        return False, 0
    return upsert({
        COL_INSURED:   insured,
        COL_CLAIM:     claim,
        COL_CARRIER:   carrier,
        COL_SUMMARY:   summary,
        COL_ESTIMATOR: estimator,
    }, source="XA")


def upsert_from_email(*, insured: str = "", claim: str = "",
                       carrier: str = "", subject: str = "",
                       preview: str = "") -> tuple[bool, int]:
    """Email scan first-seen path. Called when the inbox scan tags a
    message as a dispute. Intake Source = Email. Summary combines
    subject + preview snippet so the user can judge the row without
    opening Outlook.

    Same insert-if-missing semantics as `upsert_from_apa` — a second
    matching email won't re-stomp the Summary the user typed."""
    if not insured and not claim:
        return False, 0
    if _exists_for(claim, insured):
        return False, 0
    summary_parts = []
    if subject:
        summary_parts.append(subject.strip())
    if preview:
        snippet = (preview.strip()[:140]
                    + ("…" if len(preview.strip()) > 140 else ""))
        if snippet:
            summary_parts.append(snippet)
    return upsert({
        COL_INSURED: insured,
        COL_CLAIM:   claim,
        COL_CARRIER: carrier,
        COL_SUMMARY: " — ".join(summary_parts),
    }, source="Email")


# ── Hygiene helpers ───────────────────────────────────────────────────────

# ── Trello board ingestion ──────────────────────────────────────────────────
# The user maintains a shared Trello board ("EMS BILLING DISPUTES",
# shortLink `bnV8zbpJ`) where every active billing dispute lives as a
# card. Lanes on that board are named after the carrier ("AAA",
# "STATE FARM", …) with " - RESOLVED" appended once a dispute closes.
# Cards carry the standard customer / claim / carrier desc fields, so
# we can drop them into the tracker with no manual transcription.
DISPUTES_BOARD_SHORT_LINK = "bnV8zbpJ"


_FLAT_DESC_HEADER_RE = re.compile(r"\*\*([A-Z][A-Z0-9 /#&'\-]*?)\*\*")


def _parse_flat_card_desc(desc: str) -> dict[str, str]:
    """Parse the EMS BILLING DISPUTES board's flat ``**HEADER**\\nvalue``
    desc format. Different from ``trello_client.parse_card_desc`` —
    that one expects nested section headers (CUSTOMER INFORMATION →
    EMAIL); this board has flat single-line keys (CUSTOMER, CLAIM NUMBER,
    etc.) with the value sitting directly underneath.

    Returns ``{HEADER: stripped_value}``. Trello's zero-width markers
    (U+200C) and ``---`` divider lines are stripped from values."""
    if not desc:
        return {}
    out: dict[str, str] = {}
    matches = list(_FLAT_DESC_HEADER_RE.finditer(desc))
    for i, m in enumerate(matches):
        key = m.group(1).strip()
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(desc)
        raw = desc[start:end]
        # Strip the zero-width Trello marker + horizontal-rule dividers.
        raw = raw.replace("‌", "")
        raw = re.sub(r"^\s*---+\s*$", "", raw, flags=re.MULTILINE)
        # Collapse whitespace but preserve paragraph breaks.
        lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
        out[key] = "\n".join(lines)
    return out


def _strip_carrier_suffix(name: str) -> str:
    """Card titles look like ``Smith, John - AAA`` — strip the trailing
    ``- <CARRIER>`` so the row's Insured column is clean."""
    if not name:
        return ""
    s = name.strip()
    # Remove " - CARRIER" tail; keep commas inside the insured name.
    m = re.match(r"^(.*?)\s*[-–]\s*([A-Za-z0-9 &/]+)\s*$", s)
    if m and len(m.group(2)) <= 32:
        return m.group(1).strip()
    return s


def _clean_lane_to_carrier(lane: str) -> str:
    """``AAA - RESOLVED`` → ``AAA``. The board's RESOLVED suffix is a
    status marker, not part of the carrier name."""
    if not lane:
        return ""
    s = lane.strip()
    s = re.sub(r"\s*[-–]\s*RESOLVED\s*$", "", s, flags=re.IGNORECASE)
    return s.strip()


def _lane_status(lane: str) -> str:
    """Return the dispute status implied by the lane name. Lanes with
    ``RESOLVED`` map straight to **Closed** — per user direction,
    resolved-on-Trello and closed are the same thing here, and using
    the canonical "Closed" value means open_disputes() drops them
    automatically (it filters on status='closed' case-insensitive).
    Everything else stays Open."""
    if lane and re.search(r"\bresolved\b", lane, re.IGNORECASE):
        return "Closed"
    return "Open"


def _extract_dispute_from_card(card: dict, lane_name: str) -> dict:
    """Build the field bundle for one Trello card. Pulls insured from
    the card title, carrier from the lane (with desc fallback), and
    claim / adjuster / dates from the parsed flat-format desc."""
    card_name = (card or {}).get("name") or ""
    desc_fields = _parse_flat_card_desc(card.get("desc") or "")
    insured = _strip_carrier_suffix(card_name) or desc_fields.get(
        "CUSTOMER", "")
    carrier = _clean_lane_to_carrier(lane_name) or desc_fields.get(
        "INSURANCE", "")
    claim = desc_fields.get("CLAIM NUMBER", "")
    # Summary preview from FIELD NOTES + OFFICE NOTES (whichever has
    # content) so the spreadsheet row has at least a sentence-shaped
    # description.
    summary_parts = []
    for k in ("FIELD NOTES", "OFFICE NOTES"):
        v = desc_fields.get(k, "")
        if v:
            summary_parts.append(v)
    summary = "  ·  ".join(summary_parts)
    if len(summary) > 280:
        summary = summary[:280].rstrip() + "…"
    return {
        "insured": insured,
        "claim":   claim,
        "carrier": carrier,
        "summary": summary,
        "status":  _lane_status(lane_name),
        "card_url": card.get("shortUrl") or card.get("url") or "",
    }


def upsert_from_trello_card(card: dict, lane_name: str = "") -> tuple[bool, int]:
    """Insert OR merge-fill a Trello card into the tracker.

    Match key is (canonical claim#, canonical insured). When no match
    exists this is a clean insert tagged Intake=Trello. When a match
    exists, the new payload is merged in fill-empty mode — any cell
    the user has edited stays as-is; only blank cells get populated
    with what Trello has. That way a re-pull combines fresh Carrier /
    Summary data into a partial row the user started manually
    (Estimator / Priority / Status), without trampling those edits.

    Earlier behavior was insert-or-skip — once a row existed, Trello
    re-pulls were dropped on the floor and never enriched the row.
    User asked for the merge semantics: "it can grab/combine them with
    the existing ones if they match. it must check before doing the
    merge" (2026-05-28).

    Lane-derived status is still INSERT-ONLY (treated as a default,
    not authoritative — closing / moving cards on Trello doesn't
    retroactively change spreadsheet status; that's a Phase 2
    bidirectional sync concern).
    """
    fields = _extract_dispute_from_card(card, lane_name)
    if not fields["insured"] and not fields["claim"]:
        return False, 0
    existing = _exists_for(fields["claim"], fields["insured"])
    payload = {
        COL_INSURED: fields["insured"],
        COL_CLAIM:   fields["claim"],
        COL_CARRIER: fields["carrier"],
        COL_SUMMARY: fields["summary"],
    }
    # Status is insert-only — never overwrite a user's manual status
    # change. On merge runs we drop it entirely so the merge-fill path
    # can't even consider it (an empty cell + a Trello status would
    # otherwise repopulate Status to "Open" after the user resolved
    # the dispute manually).
    if not existing:
        payload[COL_STATUS] = fields["status"]
    return upsert(payload, source="Trello", merge_only_empty=existing)


def sync_from_trello_board(*, board_short_link: str = "",
                            progress_cb=None) -> dict:
    """Walk every open card on the disputes board, upsert each into
    the tracker. Cards in RESOLVED lanes get status=Resolved on first
    import; cards in active lanes get status=Open.

    Returns ``{"added": N, "merged": N, "skipped_existing": N,
    "errors": N}``. `merged` counts existing rows whose blank fields
    got filled with fresh Trello data; `skipped_existing` counts
    matches where every cell was already populated (nothing to fill).
    ``progress_cb(i, total, card_name)`` fires once per card if given.

    Network walk — call from a background thread.

    A franchise with no board of its own syncs NOTHING. Falling back to
    the default board would file another office's disputes into this
    office's tracker, which is exactly what the tracker is meant to
    keep straight.
    """
    link = (board_short_link or "").strip() or configured_board_link()
    if not link:
        return {"added": 0, "merged": 0, "skipped_existing": 0, "errors": 0,
                "no_board": True}
    import trello_client as _tc
    try:
        lists = _tc._call(
            f"/boards/{link}/lists",
            params={"fields": "id,name", "filter": "open"},
        ) or []
    except Exception:
        return {"added": 0, "merged": 0,
                "skipped_existing": 0, "errors": 1}
    lane_name_by_id = {lst["id"]: lst.get("name", "") for lst in lists}
    try:
        cards = _tc._call(
            f"/boards/{link}/cards",
            params={"fields": "id,name,desc,idList,shortUrl,closed",
                     "filter": "open"},
        ) or []
    except Exception:
        return {"added": 0, "merged": 0,
                "skipped_existing": 0, "errors": 1}

    added = 0
    merged = 0
    skipped = 0
    errors = 0
    total = len(cards)
    for i, card in enumerate(cards, 1):
        if progress_cb is not None:
            try:
                progress_cb(i, total, card.get("name", ""))
            except Exception:
                pass
        try:
            lane = lane_name_by_id.get(card.get("idList") or "", "")
            was_new, row_no = upsert_from_trello_card(card, lane_name=lane)
            if was_new:
                added += 1
            elif row_no > 0:
                # Existing row was touched by the merge-fill pass —
                # either one or more blanks were filled, or every
                # field was already populated and the write was a
                # no-op. Either way it counts as "merged" (the merge
                # ran successfully against an existing row).
                merged += 1
            else:
                skipped += 1
        except Exception:
            errors += 1
    return {"added": added, "merged": merged,
            "skipped_existing": skipped, "errors": errors}


_CLOSED_STATUSES = {"closed", "resolved"}


def open_disputes() -> list[dict]:
    """All rows whose Status isn't terminal. Used by the Hygiene panel
    + dashboards. Treats both 'Closed' and 'Resolved' as terminal —
    Resolved was used by an earlier Trello-sync run before the lane
    mapper was switched to write 'Closed' directly, so the filter
    needs to cover both to keep those rows out of the open list."""
    return [r for r in read_rows()
            if (r.get(COL_STATUS) or "").strip().lower()
               not in _CLOSED_STATUSES]


def disputes_needing_ack() -> list[dict]:
    """Open disputes where Ack Email Sent ≠ Yes. These are the rows
    most likely to slip — adjuster hasn't been acknowledged yet."""
    return [r for r in open_disputes()
            if (r.get(COL_ACK) or "").strip().lower() != "yes"]


def disputes_overdue(*, today: date | None = None) -> list[dict]:
    """Open disputes whose Target Response Date is in the past."""
    today = today or date.today()
    out: list[dict] = []
    for r in open_disputes():
        tgt = r.get(COL_TARGET_DATE)
        if not tgt:
            continue
        try:
            if isinstance(tgt, (datetime, date)):
                d = (tgt.date() if isinstance(tgt, datetime) else tgt)
            else:
                d = datetime.fromisoformat(str(tgt)[:10]).date()
        except (ValueError, TypeError):
            continue
        if d < today:
            out.append(r)
    return out
