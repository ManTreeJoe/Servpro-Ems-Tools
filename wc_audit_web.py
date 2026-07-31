"""WC Audit — Pywebview panel.

Monthly tool. Picks the WC export, classifies rows via Trello
(pending-appr > recon > estimating > ems/contents > not_sold),
writes a 5-sheet .xlsx to the shared share, posts to Sam via Teams,
and surfaces the last cached classification.
"""
from __future__ import annotations
import os, sys, datetime
import webview

_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path: sys.path.insert(0, _HERE)

import persistence

ASSETS_DIR = os.path.join(_HERE, "wc_audit_web_assets")
INDEX_HTML = os.path.join(ASSETS_DIR, "index.html")


class Api:
    def __init__(self): self._window = None
    def attach(self, w): self._window = w

    def last_result(self):
        try:
            raw = persistence.get("wc_audit_last_result") or {}
        except Exception:
            raw = {}
        if not isinstance(raw, dict): raw = {}
        return {
            "ran_at":     raw.get("ran_at") or "",
            "input_file": raw.get("input_file") or "",
            "output_file": raw.get("output_file") or "",
            "counts":     raw.get("counts") or {},
            "total":      raw.get("total") or 0,
        }

    def open_url(self, url=None):
        """Open a URL via the OS default browser. Used by the per-row
        🔗 Trello button so the user can jump straight from a WC row
        to its Trello card.

        Same defensive arg unpack + dict return shape as `open_file`
        so the JS click handler can surface any failure in the
        status bar instead of silently no-op'ing.
        """
        if isinstance(url, (list, tuple)) and url:
            url = url[0]
        if not url:
            return {"ok": False, "error": "no url given"}
        try:
            import webbrowser
            webbrowser.open(str(url))
            return {"ok": True, "url": str(url)}
        except Exception as ex:
            return {"ok": False, "error": f"{type(ex).__name__}: {ex}"}

    def open_file(self, path=None):
        """Open a file on disk via the OS default handler.

        Returns a dict so JS can surface useful errors instead of
        silently no-opping. Earlier this returned True/False and the
        click handler dropped the False on the floor, so the user
        saw "nothing happens" with no explanation.

        Pywebview occasionally bundles JS positional args into the
        first Python positional as a list (observed in production on
        audit_web's archive_month_apply). We unpack the bundle so a
        valid string path still flows through.
        """
        if isinstance(path, (list, tuple)) and path:
            path = path[0]
        if not path:
            return {"ok": False, "error": "no path given"}
        path = str(path)
        if not os.path.isfile(path):
            return {"ok": False, "error": f"not a file: {path}"}
        try:
            os.startfile(path)
            return {"ok": True, "path": path}
        except Exception as ex:
            return {"ok": False, "error": f"{type(ex).__name__}: {ex}",
                    "path": path}

    # ── Run classification + Save & Send ────────────────────────────
    def pick_source_file(self) -> str:
        """Open a native file picker for the WorkCenter source xlsx."""
        try:
            if self._window is None: return ""
            result = self._window.create_file_dialog(
                webview.OPEN_DIALOG,
                allow_multiple=False,
                file_types=("Excel files (*.xlsx;*.xls)",))
            if not result: return ""
            return result[0] if isinstance(result, (list, tuple)) else result
        except Exception:
            return ""

    def pick_saved_workbook(self) -> str:
        """Open a native file picker scoped to the OUTPUT_DIR so the
        user can re-open a previously-saved WC audit workbook (the
        ones written by Save & Send). Same picker as pick_source_file
        but rooted at the share path."""
        try:
            if self._window is None: return ""
            import wc_audit as _wc
            _od = _wc.output_dir()
            initial = _od if os.path.isdir(_od) else ""
            result = self._window.create_file_dialog(
                webview.OPEN_DIALOG,
                allow_multiple=False,
                directory=initial or "",
                file_types=("Excel files (*.xlsx;*.xls)",))
            if not result: return ""
            return result[0] if isinstance(result, (list, tuple)) else result
        except Exception:
            return ""

    # ── Output-folder location (change folder) ──────────────────────
    def audit_folder_location(self) -> dict:
        """Current WC-audit output directory + whether it exists."""
        try:
            import wc_audit as _wc
            d = _wc.output_dir()
            return {"dir": d, "exists": bool(d and os.path.isdir(d))}
        except Exception as ex:
            return {"error": str(ex)}

    def change_audit_folder(self) -> dict:
        """Folder picker → repoint the WC-audit output directory
        (persisted to config under `wc_audit_dir`). Does not move
        existing files — saved audits are dated, standalone workbooks."""
        try:
            if self._window is None:
                return {"ok": False, "error": "no window"}
            import wc_audit as _wc
            cur = _wc.output_dir()
            res = self._window.create_file_dialog(
                webview.FOLDER_DIALOG, directory=cur or "",
                allow_multiple=False)
            if not res:
                return {"ok": False, "canceled": True}
            new_dir = res[0] if isinstance(res, (list, tuple)) else res
            if not new_dir:
                return {"ok": False, "canceled": True}
            _wc.set_output_dir(new_dir)
            return {"ok": True, "dir": new_dir}
        except Exception as ex:
            return {"ok": False, "error": f"{type(ex).__name__}: {ex}"}

    def load_saved_workbook(self, path=None) -> dict:
        """Read a previously-saved WC audit workbook back into
        `_classify_rows` so the panel resumes editing the same set Sam
        will eventually see. Each sheet's name maps to a category key
        via CATEGORY_LABELS reversal; rows are reconstructed from the
        D-J columns + Assignee (for Estimating).

        Use when the user has hand-edited a saved workbook outside
        the panel and wants to bring it back in for further triage,
        OR when they want to re-open today's file after closing the
        panel.

        Returns `{ok, total, counts, input_file}`.
        """
        if isinstance(path, (list, tuple)) and path:
            path = path[0]
        path = str(path or "").strip()
        if not path:
            return {"ok": False, "error": "no path given"}
        if not os.path.isfile(path):
            return {"ok": False, "error": f"file not found: {path}"}
        try:
            import openpyxl
            import wc_audit as _wc
            # Reverse mapping: sheet title → category key.
            label_to_key = {lbl.strip().lower(): key
                            for key, lbl in _wc.CATEGORY_LABELS.items()}
            wb = openpyxl.load_workbook(path, data_only=True)
            new_rows = []
            counts = {}
            for ws in wb.worksheets:
                cat = label_to_key.get(ws.title.strip().lower())
                if cat is None:
                    continue  # unknown sheet — skip
                headers = [c.value for c in ws[1]]
                h_idx = {str(h or "").strip().lower(): i
                         for i, h in enumerate(headers)}
                # Position-fallback when headers don't match the
                # canonical names (sheet was edited and headers got
                # renamed). Falls back to xlsx column order D-J.
                col = lambda key, default: h_idx.get(key.lower(), default)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(c not in (None, "") for c in row):
                        continue
                    def _at(name, default):
                        i = col(name, default)
                        return row[i] if i < len(row) else None
                    new_rows.append({
                        "date_received": _at("date received", 0),
                        "corp_ref":      str(_at("corporate ref #", 1) or ""),
                        "project_num":   str(_at("project #", 2) or ""),
                        "property_type": str(_at("property type", 3) or ""),
                        "type":          str(_at("type", 4) or ""),
                        "progress":      str(_at("progress", 5) or ""),
                        "customer":      str(_at("customer", 6) or "").strip(),
                        "not_sold_det":  "",
                        "assignee":      (str(_at("assignee", 7) or "").strip()
                                          if cat == "estimating" else ""),
                        "_cat":          cat,
                        # Card metadata left empty — the load path
                        # doesn't include Trello state, so per-row
                        # 🔗 / 📌 buttons render as inactive until
                        # the user manually re-pins or re-classifies.
                        "card_id":       "",
                        "card_url":      "",
                        "card_name":     "",
                        "board_name":    "",
                        "list_name":     "",
                    })
                    counts[cat] = counts.get(cat, 0) + 1
            wb.close()
            if not new_rows:
                return {"ok": False,
                        "error": "no recognizable sheets in workbook"}
            self._classify_rows = new_rows
            # Mark as a read-only loaded workbook so list_rows_for_cat
            # preserves the original row order instead of re-sorting by
            # assignee. The user is viewing a completed workbook, not
            # building a new one.
            self._loaded_readonly = True
            # Stash input_file so the panel's status meta survives the
            # reload (matches what save_and_send writes).
            try:
                from persistence import set_value
                # last_result mirrors the schema run_classify uses.
                set_value("wc_audit_last_result", {
                    "ran_at":      datetime.datetime.now().strftime(
                        "%Y-%m-%d %H:%M"),
                    "input_file":  path,
                    "output_file": path,
                    "counts":      counts,
                    "total":       len(new_rows),
                })
            except Exception:
                pass
            return {"ok": True, "total": len(new_rows),
                    "counts": counts, "input_file": path}
        except Exception as ex:
            return {"ok": False,
                    "error": f"{type(ex).__name__}: {ex}"}

    def categories(self) -> list:
        """Category list + labels for the run-classification tabs."""
        try:
            import wc_audit as _wc
            return [{"key": c, "label": _wc.CATEGORY_LABELS.get(c, c)}
                    for c in _wc.CATEGORIES]
        except Exception:
            return []

    def run_classify(self, source_path: str) -> dict:
        """Run the WorkCenter classification pass. Spawns a background
        thread that loads the source xlsx, builds the Trello board
        index (slowest step), buckets every row by category, and
        emits progress events. Mirrors the Tk `_pick_and_load` flow."""
        # Fresh classification — clear the readonly flag so assignee
        # sorting works normally on new audits.
        self._loaded_readonly = False
        import threading as _t
        if getattr(self, "_classify_running", False):
            return {"started": False, "reason": "classify already running"}
        if not source_path or not os.path.isfile(source_path):
            return {"started": False, "reason": "source file not found"}
        self._classify_running = True
        self._classify_rows = []
        self._classify_source = source_path

        def _bg():
            try:
                import wc_audit as _wc
                import web_event
                web_event.event(self._window, "wc:classify-progress",
                                {"i": 0, "n": 0, "name": "Loading source…"})
                rows = _wc.load_source(source_path)
                # Trello index build is the slow step — stream progress,
                # throttled so the synchronous evaluate_js stream doesn't
                # freeze the UI.
                def _progress(i, n, board_name):
                    is_final = bool(n) and i >= n
                    web_event.throttled_event(
                        self._window, "wc:classify-progress",
                        {"i": i, "n": n, "name": board_name or "Indexing…"},
                        final=is_final)
                idx = _wc.build_trello_index(progress_cb=_progress)
                _wc.bucket_rows(rows, idx)
                self._classify_rows = rows
                # Counts per category for the summary tile
                by_cat = {c: len(_wc.rows_for_cat(rows, c))
                          for c in _wc.CATEGORIES}
                payload = {
                    "ok":         True,
                    "total":      len(rows),
                    "by_cat":     by_cat,
                    "source":     source_path,
                    "categories": [
                        {"key": c, "label": _wc.CATEGORY_LABELS.get(c, c),
                         "count": by_cat.get(c, 0)}
                        for c in _wc.CATEGORIES],
                }
                web_event.event(self._window, "wc:classify-done", payload)
            except Exception as ex:
                try:
                    import web_event
                    web_event.event(self._window, "wc:classify-done",
                                    {"ok": False,
                                     "error": f"{type(ex).__name__}: {ex}"})
                except Exception:
                    pass
            finally:
                self._classify_running = False

        _t.Thread(target=_bg, daemon=True).start()
        return {"started": True}

    def list_rows_for_cat(self, cat: str) -> list:
        """Return classified rows for a specific category — drives the
        per-tab table in the UI after run_classify completes.

        Each row dict from `wc_audit.load_source` carries the
        individual columns (date_received, corp_ref, project_num,
        property_type, type, progress, customer). We assemble the
        `cells` array on the way out in the same order the workbook
        writer uses, so the UI table columns match the output xlsx.
        Earlier this called `r.get("cells")` which doesn't exist on
        the row dict — the count was right but every customer/CAT/
        cell column rendered blank.
        """
        if not getattr(self, "_classify_rows", None): return []
        try:
            import wc_audit as _wc
            rows = _wc.rows_for_cat(self._classify_rows, cat) or []
            # Build a stable row_id = position in _classify_rows. Two
            # rows with the same customer (e.g., Rory Nowell on a CON
            # shell AND a WTR shell) need distinct identifiers so the
            # per-row actions (Move, Pin, Assignee) hit the row the
            # user actually clicked.
            all_rows = list(self._classify_rows)
            id_by_row = {id(r): i for i, r in enumerate(all_rows)}
            # Sort Estimating by assignee so same-estimator rows
            # group together. Skip when viewing a loaded saved workbook
            # — the file already has its own order which the user wants
            # to preserve as-is.
            if cat == "estimating" and not getattr(self, "_loaded_readonly", False):
                def _asg_key(r):
                    a = (r.get("assignee") or "").strip().lower()
                    return (a or "~", a)
                rows = sorted(rows, key=_asg_key)
            out = []
            for r in rows:
                cells = [
                    str(r.get("date_received") or ""),
                    str(r.get("corp_ref")      or ""),
                    str(r.get("project_num")   or ""),
                    str(r.get("property_type") or ""),
                    str(r.get("type")          or ""),
                    str(r.get("progress")      or ""),
                    str(r.get("customer")      or ""),
                ]
                out.append({
                    "row_id":   id_by_row.get(id(r), -1),
                    "customer": r.get("customer") or "",
                    "category": r.get("_cat") or r.get("category") or "",
                    "assignee": r.get("assignee") or "",
                    "cells":    cells,
                    "card_id":    r.get("card_id") or "",
                    "card_url":   r.get("card_url") or "",
                    "card_name":  r.get("card_name") or "",
                    "board_name": r.get("board_name") or "",
                    "list_name":  r.get("list_name") or "",
                })
            return out
        except Exception:
            return []

    def search_trello_for_pin(self, query: str = "") -> dict:
        """Fuzzy-search Trello for cards matching `query` so the user
        can manually pin one to a WC row when the auto-classifier
        missed the name. Mirrors job_widgets.open_trello_pin_dialog's
        backend — `trello_client.find_cards_by_name` returns the
        top matches across every open card on every in-scope board.

        Returns `{ok, results: [{card_id, card_name, card_url,
        board_name, list_name}], error?}`. Board and lane names are
        resolved best-effort via the workspace board / list metadata
        already cached during the last index build, so each result
        shows enough context for the user to pick the right card.
        """
        if isinstance(query, (list, tuple)) and query:
            query = query[0]
        query = str(query or "").strip()
        if not query:
            return {"ok": False, "error": "query is empty",
                    "results": []}
        try:
            import trello_client as tc
            cards = tc.find_cards_by_name(query, max_results=15) or []
        except Exception as ex:
            return {"ok": False, "error": str(ex), "results": []}
        # Cache board id → name and list id → name on the Api so we
        # don't refetch on every search. _board_name_cache lazily
        # populates as new ids come in.
        bcache = getattr(self, "_board_name_cache", None) or {}
        lcache = getattr(self, "_list_name_cache", None) or {}
        results = []
        for c in cards:
            cid = c.get("id") or ""
            cname = c.get("name") or ""
            url = c.get("shortUrl") or (
                f"https://trello.com/c/{cid}" if cid else "")
            bid = c.get("idBoard") or ""
            lid = c.get("idList") or ""
            board_name = bcache.get(bid, "")
            list_name = lcache.get(lid, "")
            # Lazy lookup if missing — single Trello fetch per
            # previously-unseen id, then cached.
            if bid and not board_name:
                try:
                    b = tc._call(f"/boards/{bid}",
                                  params={"fields": "name"}) or {}
                    board_name = b.get("name", "")
                    bcache[bid] = board_name
                except Exception:
                    pass
            if lid and not list_name:
                try:
                    l = tc._call(f"/lists/{lid}",
                                  params={"fields": "name"}) or {}
                    list_name = l.get("name", "")
                    lcache[lid] = list_name
                except Exception:
                    pass
            results.append({
                "card_id":    cid,
                "card_name":  cname,
                "card_url":   url,
                "board_name": board_name,
                "list_name":  list_name,
            })
        self._board_name_cache = bcache
        self._list_name_cache = lcache
        return {"ok": True, "results": results}

    def pin_trello_card(self, customer: str = "",
                          card_id: str = "", row_id=None) -> dict:
        """Manually attach a Trello card to a WC row. Persists the
        pin (so next month's classifier auto-finds it AND so audit /
        snapshot / job-notes see the same mapping) and re-classifies
        the row in memory based on the pinned card's board.

        Returns `{ok, row, error?}` so the JS can splice the updated
        row directly into the rendered table.
        """
        if isinstance(customer, (list, tuple)) and customer:
            customer = customer[0]
        if isinstance(card_id, (list, tuple)) and card_id:
            card_id = card_id[0]
        customer = str(customer or "").strip()
        card_id  = str(card_id  or "").strip()
        if not customer or not card_id:
            return {"ok": False, "error": "customer + card_id required"}
        if not getattr(self, "_classify_rows", None):
            return {"ok": False, "error": "no classification loaded"}
        try:
            import persistence as _per
            import trello_client as tc
            import wc_audit as _wc
            # Persist the pin — uses the canon-pin-key under the hood
            # so every other tool (audit / snapshot / job-notes / etc.)
            # picks it up automatically. Same helper the per-row pin
            # button uses elsewhere in the suite.
            _per.set_trello_card_id(customer, card_id)
            # Fetch full card details so we can update the row in
            # memory and re-classify without a full rebuild.
            card = tc.get_card(card_id) or {}
            bid = card.get("idBoard") or ""
            lid = card.get("idList")  or ""
            # Reuse the search method's name caches.
            bcache = getattr(self, "_board_name_cache", None) or {}
            lcache = getattr(self, "_list_name_cache", None) or {}
            board_name = bcache.get(bid) or ""
            list_name  = lcache.get(lid)  or ""
            if bid and not board_name:
                try:
                    b = tc._call(f"/boards/{bid}",
                                  params={"fields": "name"}) or {}
                    board_name = b.get("name", "")
                    bcache[bid] = board_name
                except Exception:
                    pass
            if lid and not list_name:
                try:
                    l = tc._call(f"/lists/{lid}",
                                  params={"fields": "name"}) or {}
                    list_name = l.get("name", "")
                    lcache[lid] = list_name
                except Exception:
                    pass
            self._board_name_cache = bcache
            self._list_name_cache  = lcache
            # Locate the row + apply the new bucket + card fields.
            target = self._find_row(row_id=row_id, customer=customer)
            if target is None:
                return {"ok": False, "error": "row not found"}
            target["card_id"]    = card_id
            target["card_url"]   = (card.get("shortUrl")
                                     or f"https://trello.com/c/{card_id}")
            target["card_name"]  = card.get("name") or ""
            target["board_name"] = board_name
            target["list_name"]  = list_name
            # Re-bucket using the same priority chain as classify_row:
            # pending-approval lane > board > suffix fallback.
            lane_low = (list_name or "").lower()
            if "pending" in lane_low and "appr" in lane_low:
                new_cat = "pending_approval"
                target["assignee"] = ""
            else:
                bucket = _wc._board_classifies_as(board_name)
                if bucket == "estimating":
                    new_cat = "estimating"
                    target["assignee"] = list_name or ""
                elif bucket:
                    new_cat = bucket
                    target["assignee"] = ""
                else:
                    # Fall back to project # suffix when board name
                    # doesn't classify.
                    new_cat = (_wc._category_from_loss_code(
                        _wc._suffix_from_project(
                            target.get("project_num") or ""))
                        or "not_sold")
                    target["assignee"] = ""
            target["_cat"] = new_cat
            return {"ok": True,
                    "category":  new_cat,
                    "card_id":    target["card_id"],
                    "card_url":   target["card_url"],
                    "card_name":  target["card_name"],
                    "board_name": target["board_name"],
                    "list_name":  target["list_name"],
                    "assignee":   target["assignee"]}
        except Exception as ex:
            return {"ok": False,
                    "error": f"{type(ex).__name__}: {ex}"}

    def update_row_fields(self, row_id=None, fields=None) -> dict:
        """Bulk-update editable column values on a single classified
        row. Mirrors the Disputes page's edit-modal save: receives a
        dict of `{field_name: new_value}` and writes them through to
        the in-memory `_classify_rows` so the next list_rows_for_cat
        / save_and_send picks them up.

        Editable fields:
          date_received, corp_ref, project_num, property_type, type,
          progress, customer, assignee.

        Defensive arg unpack — bridge sometimes bundles positionals.
        """
        if isinstance(row_id, (list, tuple)) and row_id and fields is None:
            # Bundled call: row_id arg actually carried both positionals.
            bundle = list(row_id)
            row_id = bundle[0] if len(bundle) > 0 else None
            fields = bundle[1] if len(bundle) > 1 else None
        if not getattr(self, "_classify_rows", None):
            return {"ok": False, "error": "no classification loaded"}
        if not isinstance(fields, dict):
            return {"ok": False, "error": "fields must be a dict"}
        try:
            r = self._find_row(row_id=row_id)
            if r is None:
                return {"ok": False, "error": "row not found"}
            # Allowlist — never overwrite system fields like _cat,
            # card_id, etc. via this generic editor.
            allowed = {"date_received", "corp_ref", "project_num",
                        "property_type", "type", "progress",
                        "customer", "assignee"}
            for k, v in fields.items():
                if k in allowed:
                    r[k] = ("" if v is None else str(v).strip())
            return {"ok": True}
        except Exception as ex:
            return {"ok": False,
                    "error": f"{type(ex).__name__}: {ex}"}

    def _find_row(self, row_id=None, customer=None):
        """Locate a row in `_classify_rows` by stable `row_id` first,
        falling back to the first customer-name match for backward
        compatibility. Returns `None` when nothing matches.

        Without the row_id path, two rows sharing a customer name
        (Rory Nowell with both a CON and a WTR shell) both routed
        every action to the first occurrence — the symptom the user
        reported as "the move says it worked but the row didn't
        change."
        """
        rows = getattr(self, "_classify_rows", None) or []
        if row_id is not None and row_id != "":
            try:
                idx = int(row_id)
            except (TypeError, ValueError):
                idx = -1
            if 0 <= idx < len(rows):
                return rows[idx]
        if customer:
            needle = str(customer).strip()
            for r in rows:
                if (r.get("customer") or "").strip() == needle:
                    return r
        return None

    def set_assignee(self, customer="", assignee="", row_id=None) -> dict:
        """Inline assignee edit on Estimating-tab rows. Mirrors Tk's
        double-click-cell flow (wc_audit_gui.py:289). Writes back into
        the in-memory classify_rows so save_and_send picks it up.

        Prefers `row_id` (stable position in _classify_rows) over
        `customer` so duplicate customer names route correctly.
        """
        if not getattr(self, "_classify_rows", None):
            return {"ok": False, "error": "no classification loaded"}
        try:
            r = self._find_row(row_id=row_id, customer=customer)
            if r is None:
                return {"ok": False, "error": "row not found"}
            r["assignee"] = (assignee or "").strip()
            return {"ok": True}
        except Exception as ex:
            return {"ok": False, "error": str(ex)}

    def move_row_category(self, customer="", target_cat="",
                            row_id=None) -> dict:
        """Manual reclassification — user override on top of the
        auto-bucket. Mirrors Tk's per-row 'Move to →' context menu.

        Writes the `_cat` key — the same key `wc_audit.bucket_rows`
        sets and `wc_audit.rows_for_cat` reads. Disambiguates by
        `row_id` first so duplicate-customer rows (Rory Nowell on
        two shells) move the row the user actually clicked instead
        of always retargeting the first occurrence.
        """
        if not getattr(self, "_classify_rows", None):
            return {"ok": False, "error": "no classification loaded"}
        try:
            r = self._find_row(row_id=row_id, customer=customer)
            if r is None:
                return {"ok": False, "error": "row not found"}
            r["_cat"] = target_cat
            return {"ok": True}
        except Exception as ex:
            return {"ok": False, "error": str(ex)}

    def save_and_send(self) -> dict:
        """Write the categorized workbook to the shared share + persist
        the run-result + open the output for Teams send. Mirrors the
        Tk 'Save & Send' workflow."""
        if not getattr(self, "_classify_rows", None):
            return {"ok": False, "error": "no classification to save"}
        try:
            import wc_audit as _wc
            output_path = _wc.default_output_path()
            _wc.write_workbook(self._classify_rows, output_path)
            counts = {c: len(_wc.rows_for_cat(self._classify_rows, c))
                      for c in _wc.CATEGORIES}
            try:
                import datetime as _dt
                persistence.set_value("wc_audit_last_result", {
                    "ran_at":      _dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "input_file":  getattr(self, "_classify_source", ""),
                    "output_file": output_path,
                    "counts":      counts,
                    "total":       len(self._classify_rows),
                })
            except Exception:
                pass
            # Open the workbook for Teams send
            try: os.startfile(output_path)
            except Exception: pass
            return {"ok": True, "output_file": output_path,
                    "counts": counts,
                    "total": len(self._classify_rows)}
        except Exception as ex:
            return {"ok": False, "error": str(ex)}

    def open_teams_to_sam(self) -> bool:
        """Launch a Teams chat to Sam (the typical recipient of the
        monthly WC Audit) so the user can drag the .xlsx in."""
        try:
            import persistence as _per
            sam_email = _per.get_escalation_email("Sam") or ""
            if not sam_email: return False
            import webbrowser as _wb
            _wb.open(f"msteams:/l/chat/0/0?users={sam_email}")
            return True
        except Exception:
            return False


def main(argv=None):
    api = Api()
    win = webview.create_window(
        title="WC Audit — Linguar Hub (web)",
        url=INDEX_HTML, js_api=api,
        width=900, height=720, min_size=(600, 400))
    api.attach(win)
    webview.start(debug=False)


if __name__ == "__main__":
    main()
