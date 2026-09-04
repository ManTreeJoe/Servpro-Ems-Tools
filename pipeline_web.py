"""Pipeline — Pywebview panel.

Renders every job lifecycle row (from `ems_db.job_lifecycle`) with
stage filter chips + search + sync button. Per-stage classification
via `pipeline_stages.derive_stage`. Per-row timeline modal pulls
`ems_db.list_transitions`. Export to per-stage Excel sheets.

Launch:
    python pipeline_web.py
    # or via launcher:
    python launcher.py --tool pipeline_web
"""
from __future__ import annotations

import datetime as _dt
import base64
from concurrent.futures import ThreadPoolExecutor
import io
import os
import re
import sys
import time
import webbrowser
from typing import Any

import webview

# Make sure the EMS Automation scripts dir is on sys.path. When this
# file runs standalone (double-clicked / from cmd) the cwd may be
# elsewhere; this anchors imports to the same dir as launcher.py.
_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

import ems_db
import pipeline_stages as ps
import pipeline_store
from job_settings_api import JobSettingsApi
from web_helpers import run_bg as _wh_run_bg


ASSETS_DIR = os.path.join(_HERE, "pipeline_web_assets")
INDEX_HTML = os.path.join(ASSETS_DIR, "index.html")


def _row_to_jsdict(r: dict, *, thresholds=None, history=None) -> dict:
    """Shape one lifecycle row for the frontend. Pre-computes the
    cheap derivations (days_in_stage, days_since_created, anomaly
    flag) here so the JS side stays declarative."""
    days_in = ps.days_in_stage(r)
    days_age = ps.days_since_created(r)
    stage = r.get("current_stage") or ""
    thresholds = thresholds if thresholds is not None else ps.get_thresholds()
    threshold = thresholds.get(stage, 9999)
    stall = "none"
    if days_in > threshold * 2:
        stall = "bad"
    elif days_in > threshold:
        stall = "warn"
    # Anomaly check — defaults guard already inside is_anomaly.
    try:
        is_anomaly = ps.is_anomaly(r, history=history)
    except Exception:
        is_anomaly = False
    return {
        "card_id":        r.get("card_id") or "",
        "client":         r.get("client_display") or "",
        "stage":          stage,
        "stage_label":    ps.STAGE_LABELS.get(stage, stage),
        "days_in_stage":  days_in,
        "age":            days_age,
        "last_activity":  (r.get("last_activity_at") or "")[:10],
        "owner":          r.get("owner") or "",
        "board":          r.get("board_name") or "",
        "lane":           r.get("list_name") or "",
        "card_url":       r.get("card_url") or "",
        "stall":          stall,
        "threshold":      threshold,
        "is_anomaly":     is_anomaly,
    }


# ── 🗂 Board view (live Trello kanban) ───────────────────────────────
# Boards the kanban mirrors, in display order. key = frontend section id;
# name must match the Trello board name (case-insensitive).
BOARD_SPECS = (
    ("wip", "WORK IN PROGRESS"),
    ("est", "ESTIMATING"),
    ("contents", "CONTENTS"),
    ("recon", "RECON WORK IN PROGRESS"),
)

# Stable Trello short links for boards explicitly assigned to Linguar Hub.
# Names remain the human-facing fallback, but a harmless rename must not make
# an entire division disappear from Jobs.
BOARD_SHORTLINKS = {
    "recon": "AmUodHrh",
}

# Historical EMS cards are deliberately outside BOARD_SPECS. Pulling 1,500+
# closed cards during normal startup made the active board slow again. The
# frontend requests this board only when the user opens Old Jobs.
ARCHIVE_BOARD_SPEC = ("logs", "THE LOGS")
# Franchise-neutral prefix: IE adds an EMS suffix while OC adds its office
# number. `_resolve_board` accepts one unique suffix match.


def _board_name_key(value: str) -> str:
    """Normalize harmless Trello naming differences without guessing boards."""
    return re.sub(r"[^A-Z0-9]+", " ", str(value or "").upper()).strip()


def _resolve_board(boards: list[dict], expected_name: str,
                   short_link: str = ""):
    if short_link:
        pinned = next((b for b in boards
                       if str(b.get("shortLink") or "").casefold()
                       == str(short_link).casefold()), None)
        if pinned:
            return pinned
    expected = _board_name_key(expected_name)
    exact = next((b for b in boards
                  if _board_name_key(b.get("name")) == expected), None)
    if exact:
        return exact
    # Permit a year/franchise suffix, e.g. "WORK IN PROGRESS - 2026", while
    # avoiding broad substring matching that could select an unrelated board.
    candidates = [b for b in boards
                  if _board_name_key(b.get("name")).startswith(expected + " ")]
    return candidates[0] if len(candidates) == 1 else None


def _complete_board_payload(payload: dict) -> bool:
    """A partial/missing projection is not a successful Pipeline cache."""
    boards = payload.get("boards") or []
    by_key = {b.get("key"): b for b in boards}
    return bool(payload.get("ok")) and all(
        key in by_key and not by_key[key].get("missing")
        for key, _name in BOARD_SPECS
    )


def _detected_work_environments(crm: dict, summary: dict,
                                division_cards: list[dict],
                                selected_division: str) -> list[dict]:
    """Add reliable division evidence without silently persisting a status.

    Manual states win. Folder shells and pinned/open cards only make an
    otherwise-unset division visible as Planned in the workspace.
    """
    from ems_db_common import DIVISIONS, normalize_division

    existing = {
        normalize_division(item.get("work_environment")): dict(item)
        for item in (crm.get("work_environments") or [])
        if isinstance(item, dict) and item.get("work_environment")
    }
    evidence = {division: [] for division in DIVISIONS}
    job_path = str(summary.get("path") or "").strip()
    if job_path and os.path.isdir(job_path):
        try:
            import job_folders
            shells = {str(name or "").upper()
                      for name in job_folders.shells_at(job_path)}
            for division in DIVISIONS:
                if division in shells:
                    evidence[division].append("job folder")
        except Exception:
            pass
    for card in division_cards or []:
        if not isinstance(card, dict) or not (card.get("pinned") or card.get("card_id")):
            continue
        evidence[normalize_division(card.get("division"))].append("Trello card")
    evidence[normalize_division(selected_division)].append("open board")

    merged = []
    for division in DIVISIONS:
        sources = list(dict.fromkeys(evidence[division]))
        item = existing.get(division)
        if item:
            item.setdefault("work_environment", division)
            item["detected_sources"] = sources
            merged.append(item)
        elif sources:
            merged.append({
                "work_environment": division, "stage": "planned", "owner": "",
                "inferred": True, "detected_sources": sources,
            })
    return merged

# Lanes never shown on the board — spacers + admin/template columns that
# aren't real jobs. (Per-CARD noise is handled by pipeline_stages.
# is_pipeline_skip.)
_NOISE_LANE_SUBSTRINGS = (
    "spacer", "template", "templet", "marketing team",
    "on call", "on-call", "collections process", "disposal",
    "garments", "labels",
)

# Days-in-lane stall thresholds for the board chip — a cheap, lane-
# agnostic proxy off last activity. (The Stages table keeps the precise
# per-stage thresholds.)
_BOARD_STALL_WARN = 7
_BOARD_STALL_BAD = 14


def _is_noise_lane(lane_name):
    low = (lane_name or "").strip().lower()
    if not low:
        return True
    return any(s in low for s in _NOISE_LANE_SUBSTRINGS)


def _days_since_iso(iso):
    """Whole days between an ISO date/datetime string and today; 0 on
    parse failure."""
    if not iso:
        return 0
    try:
        d = _dt.date.fromisoformat(str(iso)[:10])
        return max(0, (_dt.date.today() - d).days)
    except Exception:
        return 0


_CARD_FIELDS = ("name,desc,shortUrl,idBoard,idList,labels,due,"
                "dueComplete,dateLastActivity,closed")


def _build_board(tc, ps, key, bname, board_obj):
    """Pull + shape ONE Trello board into its lanes/cards payload.
    Shared by board_view (both boards) and board_view_one (single-board
    refresh). `board_obj` is the raw board dict from list_boards (or None
    when the board name isn't found)."""
    if not board_obj:
        return {"key": key, "name": bname, "board_id": "",
                "lanes": [], "missing": True}
    bid = board_obj.get("id")
    try:
        lists = tc._call(f"/boards/{bid}/lists",
                         params={"fields": "name,pos", "filter": "open"}) or []
    except Exception:
        lists = []
    # One board-level card request replaces one request per lane. Large WIP
    # boards previously needed dozens of serial round trips before the first
    # paint. Trello can inline every checklist in the same response.
    cards_by_list = None
    try:
        raw_cards = tc._call(f"/boards/{bid}/cards", params={
            "fields": _CARD_FIELDS, "filter": "open", "checklists": "all",
            "checklist_fields": "name,pos",
        }) or []
        cards_by_list = {}
        for card in raw_cards:
            if card.get("closed"):
                continue
            try:
                card = tc.order_checklists(card)
            except Exception:
                pass
            cards_by_list.setdefault(card.get("idList"), []).append(card)
    except Exception:
        cards_by_list = None
    lanes = []
    for l in lists:
        lname = (l.get("name") or "").strip()
        if _is_noise_lane(lname):
            continue
        if cards_by_list is not None:
            cards = cards_by_list.get(l.get("id"), [])
        else:
            try:
                cards = tc.cards_in_list_with_checklists(
                    l.get("id"), fields=_CARD_FIELDS)
            except Exception:
                cards = []
        shaped = []
        for c in cards:
            try:
                if ps.is_pipeline_skip(c.get("name") or "", lname):
                    continue
            except Exception:
                pass
            shaped.append(_card_to_board_dict(c, lname))
        lanes.append({"list_id": l.get("id"), "name": lname,
                      "pos": l.get("pos") or 0,
                      "count": len(shaped), "cards": shaped})
    return {"key": key, "name": board_obj.get("name") or bname,
            "board_id": bid, "lanes": lanes}


def _card_to_board_dict(card, lane_name):
    """Shape one Trello card for the board UI: client name + loss-type
    chips + checklist progress + due/overdue + days-in-lane stall.
    JSON-only primitives."""
    import trello_client as tc
    name = (card.get("name") or "").strip()
    days = _days_since_iso(card.get("dateLastActivity"))
    stall = "bad" if days > _BOARD_STALL_BAD else (
        "warn" if days > _BOARD_STALL_WARN else "none")
    try:
        done, total = tc.checklist_progress(card)
    except Exception:
        done, total = 0, 0
    try:
        loss = tc.card_loss_type(card)
    except Exception:
        loss = ""
    due = card.get("due") or ""
    due_complete = bool(card.get("dueComplete"))
    overdue = False
    if due and not due_complete:
        try:
            overdue = _dt.date.fromisoformat(str(due)[:10]) < _dt.date.today()
        except Exception:
            overdue = False
    return {
        "card_id":      card.get("id") or "",
        "name":         name,
        "client":       name,
        "url":          card.get("shortUrl") or "",
        "list_id":      card.get("idList") or "",
        "lane":         lane_name,
        "loss_types":   [s.strip() for s in
                         (loss.split(",") if loss else []) if s.strip()],
        "checklist":    {"done": done, "total": total},
        "due":          str(due)[:10] if due else "",
        "due_complete": due_complete,
        "overdue":      overdue,
        "days_in_lane": days,
        "stall":        stall,
        "last_activity_at": card.get("dateLastActivity") or "",
    }


def _trello_board_payload():
    """Pull all configured transition boards from the Trello adapter."""
    import trello_client as tc
    try:
        available_boards = tc.list_boards() or []
    except Exception as ex:
        return {"ok": False, "error": str(ex), "boards": []}
    with ThreadPoolExecutor(max_workers=len(BOARD_SPECS),
                            thread_name_prefix="trello-board") as pool:
        futures = [pool.submit(_build_board, tc, ps, key, bname,
                               _resolve_board(available_boards, bname,
                                              BOARD_SHORTLINKS.get(key, "")))
                   for key, bname in BOARD_SPECS]
        out = [future.result() for future in futures]
    return {"ok": True, "boards": out, "source": "trello"}


class Api(JobSettingsApi):
    """Methods exposed to JS via `pywebview.api`.

    Every call is async on the JS side; pywebview marshals across the
    process boundary automatically. Long-running operations (sync,
    enrichment) are spawned on background threads so the WebView UI
    stays responsive.
    """

    def __init__(self):
        self._window = None
        self._last_rows = []
        self._sync_running = False
        self._audit = None  # lazily-built audit_web.Api for card audits
        self._audit_card_cache = {}
        self._division_reconcile_cache = {}
        self._document_cache = {}
        self._board_view_cache = None
        self._lifecycle_view_cache = None
        self._lifecycle_housekeeping_done = False
        self._workspace_cache = {}
        self._crm_workspace_cache = {}
        self._old_jobs_cache = {}
        self._companycam_report_window = None

    def _department_changed(self):
        """Drop API-instance state that belongs to the previous franchise."""
        self._board_view_cache = None
        self._lifecycle_view_cache = None
        self._workspace_cache.clear()
        self._crm_workspace_cache.clear()
        self._audit_card_cache.clear()
        self._division_reconcile_cache.clear()
        self._document_cache.clear()
        self._old_jobs_cache.clear()
        self._audit = None

    def _workspace_cache_key(self, client: str, card_id: str = "",
                             division: str = ""):
        return ((client or "").strip().casefold(),
                (card_id or "").strip().casefold(),
                (division or "EMS").strip().upper())

    def _invalidate_workspace(self, client: str = "", card_id: str = ""):
        client_key = (client or "").strip().casefold()
        card_key = (card_id or "").strip().casefold()
        for key in list(self._workspace_cache):
            if ((client_key and key[0] == client_key) or
                    (card_key and key[1] == card_key) or
                    (not client_key and not card_key)):
                self._workspace_cache.pop(key, None)
        if client_key:
            self._crm_workspace_cache.pop(client_key, None)
        elif not card_key:
            self._crm_workspace_cache.clear()

    def _crm_workspace(self, client: str, summary: dict) -> dict:
        """Reuse the shared CRM hydration between fast and full opens."""
        key = (client or "").strip().casefold()
        cached = self._crm_workspace_cache.get(key)
        if cached and time.monotonic() - cached[0] < 45:
            return cached[1]
        result = self._audit_api().crm_job_workspace(client, summary)
        self._crm_workspace_cache[key] = (time.monotonic(), result)
        return result

    def personal_preferences(self) -> dict:
        """Safe presentation-only choices for this Windows user."""
        import config
        cfg = config.load() or {}
        return {
            "density": (cfg.get("ui_density") or "comfortable"),
            "default_view": (cfg.get("pipeline_default_view") or "board"),
            "reduce_motion": bool(cfg.get("reduce_motion", False)),
        }

    def choose_board_background(self) -> dict:
        """Choose and prepare one machine-local board background.

        The database stores only the file path in per-user UI state. Image
        bytes never enter Supabase or the text job records.
        """
        if self._window is None:
            return {"ok": False, "error": "The board window is not ready."}
        try:
            result = self._window.create_file_dialog(
                webview.OPEN_DIALOG, allow_multiple=False,
                file_types=("Board images (*.jpg;*.jpeg;*.png;*.webp;*.bmp)",))
            if not result:
                return {"ok": True, "cancelled": True}
            path = result[0] if isinstance(result, (list, tuple)) else result
            return self.load_board_background(str(path))
        except Exception as ex:
            return {"ok": False, "error": str(ex)}

    def load_board_background(self, path: str) -> dict:
        """Return a screen-sized JPEG data URL for a previously chosen file."""
        path = os.path.abspath(os.path.expanduser(str(path or "").strip()))
        if not os.path.isfile(path):
            return {"ok": False, "error": "That background image is no longer available."}
        try:
            from PIL import Image, ImageOps
            with Image.open(path) as source:
                image = ImageOps.exif_transpose(source).convert("RGB")
                image.thumbnail((2560, 1600), Image.Resampling.LANCZOS)
                out = io.BytesIO()
                image.save(out, format="JPEG", quality=82, optimize=True)
            encoded = base64.b64encode(out.getvalue()).decode("ascii")
            return {"ok": True, "path": path,
                    "name": os.path.basename(path),
                    "data_url": "data:image/jpeg;base64," + encoded}
        except Exception as ex:
            return {"ok": False, "error": f"Could not use that image: {ex}"}

    # ── 🗂 Board view + drag-move + per-card audit actions ───────────
    def board_view(self, force_trello: bool = False) -> dict:
        """Load the Linguar-owned shared Pipeline first.

        Trello is now an adapter: it seeds an empty Pipeline and refreshes it
        only when requested.  Missing migration 011 fails soft to the legacy
        live-Trello behaviour so an employee is never locked out by rollout.
        """
        if not force_trello and self._board_view_cache:
            cached_at, cached_payload = self._board_view_cache
            if time.monotonic() - cached_at < 30:
                return {**cached_payload, "cached": True}
        shared_safe = pipeline_store.shared_scope_safe()
        if not force_trello:
            saved = pipeline_store.load_board_cache()
            if _complete_board_payload(saved):
                self._board_view_cache = (time.monotonic(), saved)
                return saved
            if shared_safe:
                shared = pipeline_store.load_boards(BOARD_SPECS)
                if _complete_board_payload(shared):
                    self._board_view_cache = (time.monotonic(), shared)
                    return shared
        live = _trello_board_payload()
        if not live.get("ok"):
            if shared_safe:
                shared = pipeline_store.load_boards(BOARD_SPECS)
                if shared.get("ok"):
                    shared["warning"] = "Trello unavailable; showing shared Pipeline"
                    self._board_view_cache = (time.monotonic(), shared)
                    return shared
            return live
        live.setdefault("source", "trello")
        if shared_safe:
            mirrored = pipeline_store.mirror_boards(live)
            live["mirrored"] = bool(mirrored.get("ok"))
            live["schema_missing"] = bool(mirrored.get("schema_missing"))
        else:
            live["mirrored"] = False
            live["shared_scope_deferred"] = True
        pipeline_store.save_board_cache(live)
        self._board_view_cache = (time.monotonic(), live)
        return live

    def board_view_shared_refresh(self) -> dict:
        """Refresh the saved projection after the UI has already painted."""
        if pipeline_store.shared_scope_safe():
            shared = pipeline_store.load_boards(BOARD_SPECS)
            if _complete_board_payload(shared):
                self._board_view_cache = (time.monotonic(), shared)
                return shared
        # Until the shared Pipeline migration is installed, keep the instant
        # disk-cache paint but refresh that projection from Trello in the
        # background.  This avoids leaving users on a stale board indefinitely.
        return self.board_view(force_trello=True)

    def board_view_one(self, key: str) -> dict:
        """Re-pull a SINGLE board (the per-board ↻ refresh) so the user
        can refresh one division board without waiting on all three."""
        spec = next((s for s in (*BOARD_SPECS, ARCHIVE_BOARD_SPEC)
                     if s[0] == key), None)
        if not spec:
            return {"ok": False, "error": f"unknown board '{key}'"}
        import trello_client as tc
        import pipeline_stages as ps
        try:
            available_boards = tc.list_boards() or []
        except Exception as ex:
            return {"ok": False, "error": str(ex)}
        bname = spec[1]
        board = _build_board(tc, ps, key, bname,
                             _resolve_board(available_boards, bname))
        if key == ARCHIVE_BOARD_SPEC[0]:
            # Read-only historical adapter for now. Do not mirror thousands
            # of legacy cards into the active shared Pipeline projection.
            return {"ok": True, "board": board, "source": "trello",
                    "historical": True, "mirrored": False}
        mirrored = (pipeline_store.mirror_boards({"boards": [board]})
                    if pipeline_store.shared_scope_safe()
                    else {"ok": False, "scope_deferred": True})
        return {"ok": True, "board": board, "source": "trello",
                "mirrored": bool(mirrored.get("ok"))}

    def global_card_search(self, query: str, limit: int = 24) -> dict:
        """Search beyond the three board projections shown on Jobs.

        The local lifecycle mirror supplies closed/archive history instantly;
        Trello supplies open cards on any board the signed-in user can access.
        Results open in the same job workspace as an on-board card.
        """
        query = str(query or "").strip()
        if len(query) < 2:
            return {"ok": True, "cards": []}
        limit = max(1, min(int(limit or 24), 60))
        try:
            import card_search
            import trello_client as tc

            local = card_search.search_local(query, limit=limit * 2)
            try:
                remote = tc.find_accessible_cards_by_name(
                    query, max_results=limit * 2, include_closed=True)
            except Exception:
                remote = []
            merged = card_search.merge(local, remote, query)
            merged.sort(key=lambda row: (-float(row.get("_score") or 0),
                                         str(row.get("name") or "").casefold()))
            cards = []
            for row in merged[:limit]:
                card_id = row.get("card_id") or row.get("id") or ""
                cards.append({
                    "card_id": card_id,
                    "name": row.get("name") or "",
                    "url": row.get("url") or tc.card_url_from_id(card_id),
                    "board": row.get("board") or "",
                    "list_name": row.get("list_name") or "",
                    "source_label": ("Job history" if row.get("_source") == "local"
                                     else "Archived Trello card" if row.get("closed")
                                     else "Trello"),
                })
            return {"ok": True, "cards": cards, "count": len(cards)}
        except Exception as ex:
            return {"ok": False, "cards": [],
                    "error": f"{type(ex).__name__}: {ex}"}

    def move_card(self, card_id: str, list_id: str) -> dict:
        """Move a card to another lane on the REAL Trello board. The
        frontend confirms before calling this (drag-to-move with a
        'Move X to <lane>?' prompt)."""
        if not card_id or not list_id:
            return {"ok": False, "error": "card_id + list_id required"}
        shared = pipeline_store.move_card(card_id, list_id)
        self._board_view_cache = None
        self._invalidate_workspace(card_id=card_id)
        try:
            import trello_client as tc
            ok = tc.move_card(card_id, list_id)
            pipeline_store.mark_card_sync(
                card_id, ok=bool(ok),
                error="Trello rejected the move" if not ok else "")
            if ok:
                return {"ok": True, "synced": True,
                        "stored": bool(shared.get("ok"))}
            if shared.get("ok"):
                return {"ok": True, "synced": False,
                        "warning": "Saved in Linguar Hub; Trello needs review"}
            return {"ok": False, "error": "Trello rejected the move"}
        except Exception as ex:
            pipeline_store.mark_card_sync(card_id, ok=False, error=str(ex))
            if shared.get("ok"):
                return {"ok": True, "synced": False,
                        "warning": "Saved in Linguar Hub; Trello is unavailable"}
            return {"ok": False, "error": str(ex)}

    def create_lane(self, board_key: str, name: str) -> dict:
        name = (name or "").strip()
        if not name:
            return {"ok": False, "error": "Enter a lane name."}
        board = next((b for b in (self.board_view(force_trello=True).get("boards") or [])
                      if b.get("key") == board_key), None)
        if not board or not board.get("board_id"):
            return {"ok": False, "error": "That Trello board is not available."}
        try:
            import trello_client as tc
            created = tc.create_list(board["board_id"], name)
            if not created:
                return {"ok": False, "error": "Trello did not create the lane."}
            self._board_view_cache = None
            return {"ok": True, "lane": {"list_id": created.get("id"),
                    "name": created.get("name") or name, "pos": created.get("pos") or 0,
                    "count": 0, "cards": []}}
        except Exception as ex:
            return {"ok": False, "error": str(ex)}

    def rename_lane(self, list_id: str, name: str) -> dict:
        name = (name or "").strip()
        if not list_id or not name:
            return {"ok": False, "error": "Enter a lane name."}
        try:
            import trello_client as tc
            updated = tc.update_list(list_id, name=name)
            self._board_view_cache = None
            return {"ok": bool(updated), "name": (updated or {}).get("name") or name,
                    "error": "Trello did not rename the lane." if not updated else ""}
        except Exception as ex:
            return {"ok": False, "error": str(ex)}

    def archive_lane(self, list_id: str) -> dict:
        if not list_id:
            return {"ok": False, "error": "Lane ID is missing."}
        try:
            import trello_client as tc
            updated = tc.update_list(list_id, closed=True)
            self._board_view_cache = None
            return {"ok": bool(updated),
                    "error": "Trello did not archive the lane." if not updated else ""}
        except Exception as ex:
            return {"ok": False, "error": str(ex)}

    def reorder_lane(self, list_id: str, previous_id: str = "",
                     next_id: str = "") -> dict:
        """Place a lane between its new neighbours using Trello positions."""
        if not list_id:
            return {"ok": False, "error": "Lane ID is missing."}
        try:
            import trello_client as tc
            if previous_id and next_id:
                prev = tc.get_list(previous_id, fields="id,pos") or {}
                nxt = tc.get_list(next_id, fields="id,pos") or {}
                a, b = float(prev.get("pos") or 0), float(nxt.get("pos") or 0)
                pos = (a + b) / 2 if a < b else "bottom"
            elif next_id:
                pos = "top"
            else:
                pos = "bottom"
            updated = tc.update_list(list_id, pos=pos)
            self._board_view_cache = None
            return {"ok": bool(updated),
                    "error": "Trello did not move the lane." if not updated else ""}
        except Exception as ex:
            return {"ok": False, "error": str(ex)}

    def _audit_api(self):
        if self._audit is None:
            import audit_web
            self._audit = audit_web.Api()
        return self._audit

    # New Loss belongs to the Jobs surface. Keep one provisioning engine in
    # audit_web while the Main/Trial UI migration is in progress, but expose
    # that engine directly to Pipeline so opening intake never has to navigate
    # through Daily Run.
    def parse_new_loss(self, text: str) -> dict:
        return self._audit_api().parse_new_loss(text)

    def new_loss_templates(self) -> dict:
        return self._audit_api().new_loss_templates()

    def search_client_folders(self, query: str = "", limit: int = 40) -> dict:
        return self._audit_api().search_client_folders(query, limit)

    def plan_new_loss_folder(self, fields: dict, child: str = "",
                             second_claim: bool = False,
                             parent: str = "") -> dict:
        return self._audit_api().plan_new_loss_folder(
            fields, child, second_claim, parent)

    def create_new_loss(self, fields: dict, child: str = "",
                        second_claim: bool = False,
                        promote_first: bool = False,
                        make_folder: bool = True,
                        make_companycam: bool = True,
                        parent: str = "") -> dict:
        result = self._audit_api().create_new_loss(
            fields, child, second_claim, promote_first,
            make_folder, make_companycam, parent)
        if result.get("ok"):
            self._board_view_cache = None
            self._lifecycle_view_cache = None
            self._old_jobs_cache.clear()
        return result

    def audit_card(self, client: str) -> dict:
        """Run the single-job audit for this card's client and return a
        compact pass/fail summary for the card popover. Delegates to
        audit_web — same engine the Audit panel uses."""
        if not client:
            return {"ok": False, "error": "client required"}
        cache_key = client.strip().casefold()
        cached = self._audit_card_cache.get(cache_key)
        if cached and time.monotonic() - cached[0] < 90:
            return dict(cached[1])
        try:
            res = self._audit_api().audit_one_job(client)
        except Exception as ex:
            return {"ok": False, "error": str(ex)}
        if not res or res.get("ok") is False:
            return {"ok": False,
                    "error": (res or {}).get("error") or "audit failed"}
        row = res.get("row") or {}
        shaped = {
            "ok":            True,
            "client":        res.get("canonical") or client,
            "found":         bool(row.get("found")),
            "flagged":       bool(row.get("flagged")),
            "form_issues":   list(row.get("form_issues") or []),
            "photo_issues":  list(row.get("photo_issues") or []),
            "requirements":  [r.get("label") for r
                              in (row.get("requirements") or [])],
            "aging":         int(row.get("aging") or 0),
            "activity":      list(row.get("activity") or []),
            "techs":         list(row.get("techs") or []),
            "carrier":       row.get("carrier") or "",
            "folder":        row.get("folder") or "",
            "path":          row.get("path") or "",
            "last_seen":     row.get("last_seen") or "",
            "misplaced_forms": list(row.get("misplaced_forms") or []),
            "misplaced_photos": list(row.get("misplaced_photos") or []),
            "trello_card_id": row.get("trello_card_id") or "",
        }
        self._audit_card_cache[cache_key] = (time.monotonic(), shaped)
        return shaped

    def job_card_workspace(self, client: str, card_id: str = "",
                           division: str = "") -> dict:
        """Full Pipeline card: audit + CRM + Trello transition material."""
        started = time.monotonic()
        workspace_key = self._workspace_cache_key(client, card_id, division)
        cached_workspace = self._workspace_cache.get(workspace_key)
        if cached_workspace and time.monotonic() - cached_workspace[0] < 45:
            return {**cached_workspace[1], "cached": True,
                    "load_ms": round((time.monotonic() - started) * 1000)}
        audit_api = self._audit_api()
        reconcile_divisions = getattr(
            audit_api, "reconcile_crm_division_trello_cards", None)
        reconcile_key = (client.strip().casefold(),
                         (card_id or "").strip().casefold(),
                         (division or "EMS").strip().upper())

        def load_division_reconciliation():
            cached = self._division_reconcile_cache.get(reconcile_key)
            if cached and time.monotonic() - cached[0] < 300:
                return cached[1]
            result = (reconcile_divisions(client, card_id, division)
                      if reconcile_divisions else
                      {"ok": True, "divisions": [], "has_conflict": False})
            self._division_reconcile_cache[reconcile_key] = (
                time.monotonic(), result)
            return result

        # Card reconciliation searches Trello and is independent of the
        # folder audit. Start it first so its network time overlaps the audit
        # and CRM reads instead of adding ~5 seconds after them.
        reconcile_pool = ThreadPoolExecutor(
            max_workers=2, thread_name_prefix="workspace-identity")
        reconcile_future = reconcile_pool.submit(load_division_reconciliation)
        old_jobs_future = reconcile_pool.submit(
            self._old_ems_jobs, client, card_id)
        summary = self.audit_card(client)
        if not summary.get("ok"):
            reconcile_pool.shutdown(wait=False, cancel_futures=True)
            return summary
        # The board click carries a stronger identity than a name-derived
        # audit pin. Pass that exact card through CRM hydration as well, so a
        # legacy collapsed name cannot attach this workspace to its sibling.
        if (card_id or "").strip():
            summary = {**summary, "trello_card_id": card_id.strip()}
        crm = self._crm_workspace(client, summary)
        try:
            division_reconciliation = reconcile_future.result()
            old_jobs = old_jobs_future.result()
        finally:
            reconcile_pool.shutdown(wait=False)
        # crm_job_workspace already fetched these pins. Re-reading them here
        # used to add another shared-database round trip to every open.
        division_trello_cards = {
            "ok": True, "cards": list(crm.get("division_trello_cards") or [])}
        newly_linked = any(
            item.get("state") == "auto_pinned"
            for item in (division_reconciliation.get("divisions") or [])
            if isinstance(item, dict)
        )
        if newly_linked or not division_trello_cards["cards"]:
            division_trello_cards = audit_api.crm_division_trello_cards(client)
        division_cards = division_trello_cards.get("cards", [])
        info_sections, checklists, comments, attachments, members = [], [], [], [], []
        current_user_id = ""
        try:
            import supabase_client
            current_user_id = str((supabase_client.current_user() or {}).get("id") or "")
        except Exception:
            pass
        try:
            import ems_db
            import job_settings
            job = ems_db.find_job_by_name(client) or {}
            values = job_settings.stored_values(job)
            grouped = {}
            order = []
            for fid, section, _key, label, core in job_settings.FIELDS:
                value = str(values.get(fid) or "").strip()
                if not value:
                    continue
                if section not in grouped:
                    grouped[section] = []
                    order.append(section)
                grouped[section].append({"id": fid, "label": label,
                                         "value": value, "core": bool(core)})
            info_sections = [{"name": section.title(), "fields": grouped[section]}
                             for section in order]
        except Exception:
            pass
        from ems_db_common import normalize_division
        requested_division = normalize_division(division) if division else ""
        opened_id = (card_id or "").strip()
        if not requested_division and opened_id:
            requested_division = next((
                str(card.get("division") or "") for card in division_cards
                if str(card.get("card_id") or "").casefold() == opened_id.casefold()
            ), "")
        selected_division = requested_division or "EMS"
        selected_card = next((card for card in division_cards
                              if card.get("division") == selected_division), {})
        # Opening a board card is an exact, user-selected identity. Never
        # replace it with a name-derived/saved division pin: legacy keys could
        # collapse sibling claims (notably every ``PCM - ...`` card) and that
        # displayed another job's comments. Division-tab navigation passes no
        # opened id, so its saved division card still wins as intended.
        cid = opened_id or str(selected_card.get("card_id") or "").strip()
        if not cid and selected_division == "EMS":
            cid = (summary.get("trello_card_id") or "").strip()
        crm = dict(crm)
        crm["work_environments"] = _detected_work_environments(
            crm, summary, division_cards, selected_division)
        # Linguar Hub is the durable source. Trello below is now an import/
        # compatibility adapter and refreshes this local copy when available.
        # Trello, shared activity, shared checklists, and the document-folder
        # walk are independent. Running them together removes three stacked
        # network/disk waits from every uncached card open.
        trello_card, local_activity, trello_me = {}, [], {}
        with ThreadPoolExecutor(max_workers=5,
                                thread_name_prefix="workspace") as pool:
            checklist_future = pool.submit(pipeline_store.list_checklists, cid)
            activity_future = pool.submit(pipeline_store.list_activity, cid)
            document_future = pool.submit(
                self._document_signature_workspace,
                client, cid, summary.get("path") or "")
            if cid:
                import trello_client as tc
                trello_future = pool.submit(tc.get_card, cid)
                trello_me_future = pool.submit(tc.get_member_me)
            else:
                trello_future = pool.submit(lambda: {})
                trello_me_future = pool.submit(lambda: {})
            checklists = checklist_future.result()
            local_activity = activity_future.result()
            documents = document_future.result()
            try:
                trello_card = trello_future.result() or {}
            except Exception as ex:
                summary["trello_error"] = str(ex)
            try:
                trello_me = trello_me_future.result() or {}
            except Exception:
                trello_me = {}
        if cid:
            try:
                imported_checklists = [{
                    "id": cl.get("id") or "", "name": cl.get("name") or "Checklist",
                    "items": [{"id": item.get("id") or "",
                               "name": item.get("name") or "",
                               "complete": item.get("state") == "complete"}
                              for item in (cl.get("checkItems") or [])],
                } for cl in (trello_card.get("checklists") or [])]
                if imported_checklists:
                    saved = pipeline_store.save_checklists(
                        cid, imported_checklists, source="trello")
                    checklists = (saved.get("checklists")
                                  if saved.get("ok") else imported_checklists)
                attachments = [{"name": a.get("name") or "Attachment",
                                "url": a.get("url") or "",
                                "date": a.get("date") or ""}
                               for a in (trello_card.get("attachments") or [])]
                members = [m.get("fullName") or m.get("username") or ""
                           for m in (trello_card.get("members") or [])]
                # get_card already includes the latest 50 activity actions.
                # Re-fetching every historical comment here used as many as
                # 20 additional network calls each time a card opened. Older
                # comments remain in the shared activity table once imported.
                trello_comments = [action for action in (trello_card.get("actions") or [])
                                   if action.get("type") == "commentCard"]
                activity_import = []
                for action in trello_comments:
                    text = str((action.get("data") or {}).get("text") or "").strip()
                    if not text:
                        continue
                    creator = action.get("memberCreator") or {}
                    actor = creator.get("fullName") or "Trello"
                    comment = {"id": action.get("id") or "", "text": text,
                               "actor": actor, "at": action.get("date") or "",
                               "source": "trello",
                               "can_manage": bool(trello_me.get("id") and
                                                  creator.get("id") == trello_me.get("id"))}
                    comments.append(comment)
                    activity_import.append({
                        "action_type": "comment", "body": text,
                        "actor_name": actor, "source": "trello",
                        "external_id": comment["id"],
                        "happened_at": comment["at"],
                    })
                pipeline_store.add_activities(cid, activity_import)
            except Exception as ex:
                summary["trello_error"] = str(ex)
        # Preserve Linguar-only comments when Trello was unavailable. Avoid
        # duplicates for comments already mirrored by external action id.
        seen = {c.get("id") for c in comments if c.get("id")}
        for activity in local_activity:
            ext = activity.get("external_id") or ""
            if ext and ext in seen:
                continue
            if activity.get("action_type") == "comment":
                metadata = activity.get("metadata_json") or {}
                if isinstance(metadata, str):
                    try:
                        metadata = json.loads(metadata)
                    except (TypeError, ValueError):
                        metadata = {}
                owner_id = str(metadata.get("actor_id") or "")
                comments.append({"id": ext or activity.get("activity_key") or "",
                                 "text": activity.get("body") or "",
                                 "actor": activity.get("actor_name") or "Linguar Hub",
                                 "at": activity.get("happened_at") or "",
                                 "source": activity.get("source") or "linguar",
                                 "can_manage": bool(current_user_id and
                                                    owner_id == current_user_id)})
        comments.sort(key=lambda c: c.get("at") or "", reverse=True)
        result = {"ok": True, "client": client, "card_id": cid,
                "selected_division": selected_division,
                "selected_trello_url": (f"https://trello.com/c/{cid}"
                                        if cid else ""),
                "audit": summary, "crm": crm, "info_sections": info_sections,
                "division_trello_cards": division_cards,
                "division_card_reconciliation": division_reconciliation,
                "checklists": checklists, "comments": comments,
                "attachments": attachments, "members": members,
                "old_jobs": old_jobs,
                "documents": documents,
                "load_ms": round((time.monotonic() - started) * 1000)}
        if len(self._workspace_cache) >= 80:
            oldest = min(self._workspace_cache,
                         key=lambda key: self._workspace_cache[key][0])
            self._workspace_cache.pop(oldest, None)
        self._workspace_cache[workspace_key] = (time.monotonic(), result)
        return result

    def _old_ems_jobs(self, client: str, current_card_id: str = "") -> list:
        """Closed LOGS-EMS cards belonging to a currently-open client.

        A person can have several separate losses.  Their old EMS card is a
        previous job/claim, not more history for the active card.  Keep the
        Trello card identity intact so opening it can never mix its comments
        or claim data into the current loss.
        """
        from ems_db_sqlite import canon_key
        key = canon_key(client or "")
        cached = self._old_jobs_cache.get(key)
        if cached and time.monotonic() - cached[0] < 600:
            return [dict(row) for row in cached[1]
                    if row.get("card_id") != current_card_id]
        try:
            import trello_client as tc
            candidates = [row for row in tc.find_cards_by_name(
                client, max_results=50, with_lists=True)
                if str(row.get("board") or "").strip().casefold()
                   == tc.LOGS_BOARD_NAME.strip().casefold()
                and canon_key(row.get("name") or "") == key]

            def enrich(row):
                card = tc.get_card_lite(row.get("card_id") or "") or {}
                sections = tc.parse_card_desc(card.get("desc") or "") or {}
                claim = ""
                loss_date = ""
                received = ""
                for fields in sections.values():
                    if not isinstance(fields, dict):
                        continue
                    claim = claim or str(fields.get("CLAIM NUMBER") or "").strip()
                    loss_date = loss_date or str(fields.get("DATE OF LOSS") or "").strip()
                    received = received or str(fields.get("DATE RECEIVED") or "").strip()
                return {**row, "claim_number": claim, "loss_date": loss_date,
                        "date_received": received, "status": "Closed"}

            with ThreadPoolExecutor(max_workers=min(4, len(candidates) or 1),
                                    thread_name_prefix="old-ems-job") as pool:
                rows = list(pool.map(enrich, candidates))
            rows.sort(key=lambda row: (row.get("date_received") or "",
                                       row.get("name") or ""), reverse=True)
            self._old_jobs_cache[key] = (time.monotonic(), rows)
            return [dict(row) for row in rows
                    if row.get("card_id") != current_card_id]
        except Exception:
            return []

    def refresh_job_card_workspace(self, client: str, card_id: str = "",
                                   division: str = "") -> dict:
        """Explicit deep refresh; normal opens remain cache-friendly."""
        key = (client or "").strip().casefold()
        self._invalidate_workspace(client, card_id)
        self._audit_card_cache.pop(key, None)
        for reconcile_key in list(self._division_reconcile_cache):
            if (reconcile_key == key or
                    (isinstance(reconcile_key, tuple) and
                     reconcile_key[0] == key)):
                self._division_reconcile_cache.pop(reconcile_key, None)
        try:
            from ems_db_sqlite import canon_key
            self._old_jobs_cache.pop(canon_key(client or ""), None)
        except Exception:
            pass
        for cache_key in list(self._document_cache):
            if not card_id or cache_key[0] == card_id:
                self._document_cache.pop(cache_key, None)
        return self.job_card_workspace(client, card_id, division)

    def job_card_workspace_fast(self, client: str, card_id: str = "",
                                division: str = "") -> dict:
        """Shared-data-first workspace used for an immediate card open.

        This deliberately avoids the live folder audit, Trello search/card
        calls, and network-drive walk. The frontend follows it with the full
        workspace request while the user can already read or edit CRM data.
        """
        started = time.monotonic()
        workspace_key = self._workspace_cache_key(client, card_id, division)
        cached_workspace = self._workspace_cache.get(workspace_key)
        if cached_workspace and time.monotonic() - cached_workspace[0] < 45:
            return {**cached_workspace[1], "cached": True,
                    "load_ms": round((time.monotonic() - started) * 1000)}
        if not (client or "").strip():
            return {"ok": False, "error": "client required"}
        summary = {"ok": True, "client": client, "found": True,
                   "form_issues": [], "photo_issues": [], "requirements": [],
                   "activity": [], "path": "", "trello_card_id": card_id or ""}
        # First paint must stay one cheap identity read. The previous "fast"
        # path called crm_job_workspace, which waited on master-job joins,
        # logs, events, audit history and division-card queries (~3 seconds
        # on the office connection) before showing basic facts. The full
        # request that follows still hydrates all of that in the background.
        job = ems_db.find_job_by_name(client) or {}
        from ems_db_common import normalize_division
        selected_division = normalize_division(division) if division else "EMS"
        cid = (card_id or "").strip()
        division_cards = [{"division": selected_division, "card_id": cid,
                           "url": f"https://trello.com/c/{cid}" if cid else "",
                           "pinned": bool(cid)}]
        from job_progress import evaluate as evaluate_job_progress
        crm = {
            "ok": True,
            "job_id": job.get("job_id") or "",
            "canon_key": job.get("canon_key") or "",
            "lifecycle_stage": job.get("lifecycle_stage") or "intake",
            "job_type": job.get("job_type") or "",
            "priority": job.get("priority") or "normal",
            "work_environments": [], "division_trello_cards": division_cards,
            "relationships": [], "job_log": [], "timeline": [],
            "progress": evaluate_job_progress(job, summary, []),
        }
        crm["work_environments"] = _detected_work_environments(
            crm, summary, division_cards, selected_division)
        crm["progress"]["review_mode"] = True
        info_sections = []
        try:
            import job_settings
            values, grouped, order = job_settings.stored_values(job), {}, []
            for fid, section, _key, label, _core in job_settings.FIELDS:
                value = str(values.get(fid) or "").strip()
                if not value:
                    continue
                if section not in grouped:
                    grouped[section] = []
                    order.append(section)
                grouped[section].append({"id": fid, "label": label, "value": value})
            info_sections = [{"name": section.title(), "fields": grouped[section]}
                             for section in order]
        except Exception:
            pass
        return {"ok": True, "client": client, "card_id": cid,
                "selected_division": selected_division,
                "selected_trello_url": (f"https://trello.com/c/{cid}" if cid else ""),
                "audit": summary, "crm": crm, "info_sections": info_sections,
                "division_trello_cards": division_cards,
                "division_card_reconciliation": {"ok": True, "divisions": []},
                "checklists": [], "comments": [], "attachments": [], "members": [],
                "documents": {"provider": "DocuSign", "request": {}, "files": [],
                              "connected": False},
                "deferred_loading": True,
                "load_ms": round((time.monotonic() - started) * 1000)}

    def _document_signature_workspace(self, client: str, card_id: str,
                                      job_path: str) -> dict:
        """Text-only view of signature state and files already on disk."""
        cache_key = (str(card_id or ""), os.path.abspath(job_path) if job_path else "")
        cached = self._document_cache.get(cache_key)
        if cached and time.monotonic() - cached[0] < 120:
            return cached[1]
        pending = None
        try:
            import docusign_requests as dsr
            for entry in dsr.pending_requests():
                if ((card_id and entry.get("card_id") == card_id) or
                        str(entry.get("client") or "").casefold() ==
                        str(client or "").casefold()):
                    pending = entry
                    break
        except Exception:
            pass
        files = []
        root = os.path.abspath(job_path) if job_path else ""
        if root and os.path.isdir(root):
            for dirname, _dirs, names in os.walk(root):
                if "docs" not in dirname.casefold() and "paperwork" not in dirname.casefold():
                    continue
                for name in names:
                    if not name.casefold().endswith((".pdf", ".doc", ".docx")):
                        continue
                    full = os.path.join(dirname, name)
                    try:
                        modified = _dt.datetime.fromtimestamp(
                            os.path.getmtime(full)).isoformat(timespec="seconds")
                    except OSError:
                        modified = ""
                    low = (dirname + " " + name).casefold()
                    files.append({"name": name, "path": full,
                                  "modified_at": modified,
                                  "signed": ("signed" in low or
                                             "final paperwork" in low)})
                    if len(files) >= 100:
                        break
                if len(files) >= 100:
                    break
        files.sort(key=lambda item: item.get("modified_at") or "", reverse=True)
        result = {"provider": "DocuSign", "connected": False,
                  "connection_note": "Direct DocuSign connection is not configured yet",
                  "request": pending or {}, "files": files[:20],
                  "storage": "Official signed file stays in the X: OD job folder; only status and path are indexed"}
        self._document_cache[cache_key] = (time.monotonic(), result)
        return result

    def mark_docusign_sent(self, client: str, card_id: str = "",
                           customer_email: str = "") -> dict:
        """Record a manually-sent envelope using Job Information email."""
        try:
            import docusign_requests as dsr
            cid = (card_id or "").strip()
            if not cid:
                return {"ok": False, "error": "This job needs a linked Trello card for the current tracker"}
            entry = dsr.request(cid, client_name=client,
                                email_override=customer_email)
            if not entry:
                return {"ok": False, "error": "Could not record the DocuSign request"}
            return {"ok": True, "entry": entry}
        except Exception as ex:
            return {"ok": False, "error": str(ex)}

    def open_docusign(self) -> bool:
        return self.open_url("https://app.docusign.com/")

    @staticmethod
    def _job_docs_folder(job_path: str, division: str = "EMS") -> str:
        """Best existing DOCS folder for one division, without creating it."""
        root = os.path.abspath(str(job_path or "").strip()) if job_path else ""
        if not root or not os.path.isdir(root):
            return ""
        wanted = str(division or "EMS").strip().casefold()
        candidates = []
        for dirname, dirs, _names in os.walk(root):
            depth = os.path.relpath(dirname, root).count(os.sep)
            if depth > 4:
                dirs[:] = []
                continue
            if os.path.basename(dirname).casefold() not in ("docs", "documents", "paperwork"):
                continue
            parts = [part.casefold() for part in os.path.relpath(dirname, root).split(os.sep)]
            division_match = any(wanted == part or wanted in part for part in parts)
            candidates.append((0 if division_match else 1, len(parts), dirname))
        return min(candidates)[2] if candidates else ""

    def companycam_report_context(self, client: str, job_path: str = "",
                                  division: str = "EMS") -> dict:
        """Resolve the CompanyCam project used by the report workspace."""
        try:
            import companycam_api as cc
            if not cc.is_configured():
                return {"ok": False, "error": "CompanyCam is not connected in Settings."}
            job = ems_db.find_job_by_name(client) or {}
            address = str(job.get("address") or job.get("loss_address") or "")
            project_id = cc.find_project_id(
                client, address_hint=address, folder_path=job_path or "")
            if not project_id:
                found = cc.find_project(client, address_hint=address)
                return {"ok": False,
                        "error": found.get("reason") or "No matching CompanyCam project was found.",
                        "candidates": found.get("candidates") or []}
            project = cc.get_project(project_id) or {"id": project_id, "name": client}
            project_url = str(project.get("photo_url") or "").strip()
            if not project_url:
                project_url = f"https://app.companycam.com/projects/{project_id}"
            return {"ok": True, "client": client, "division": division or "EMS",
                    "project_id": str(project_id), "project_name": project.get("name") or client,
                    "project_url": project_url,
                    "docs_folder": self._job_docs_folder(job_path, division)}
        except Exception as ex:
            return {"ok": False, "error": f"CompanyCam lookup failed: {ex}"}

    def open_companycam_report_editor(self, client: str, job_path: str = "",
                                      division: str = "EMS",
                                      project_id: str = "") -> dict:
        """Open CompanyCam as a dedicated in-app report editing window."""
        context = self.companycam_report_context(client, job_path, division)
        if not context.get("ok"):
            return context
        if project_id:
            context["project_id"] = str(project_id)
            context["project_url"] = f"https://app.companycam.com/projects/{project_id}"
        try:
            title = f"CompanyCam Report — {client}"
            win = self._companycam_report_window
            if win is not None:
                try:
                    win.load_url(context["project_url"])
                    win.show()
                    return {**context, "opened": True, "reused": True}
                except Exception:
                    self._companycam_report_window = None
            self._companycam_report_window = webview.create_window(
                title=title, url=context["project_url"], width=1280, height=840,
                min_size=(760, 560))
            return {**context, "opened": True, "reused": False}
        except Exception as ex:
            return {**context, "ok": False,
                    "error": f"Could not open the CompanyCam report window: {ex}"}

    def companycam_quick_report_plan(self, client: str, job_path: str = "",
                                     division: str = "EMS",
                                     start_date: str = "", end_date: str = "",
                                     tag: str = "", offset: int = 0,
                                     limit: int = 120) -> dict:
        context = self.companycam_report_context(client, job_path, division)
        if not context.get("ok"):
            return context
        import companycam_report
        result = companycam_report.plan(context["project_id"],
                                        start_date=start_date,
                                        end_date=end_date, tag=tag,
                                        offset=offset, limit=limit)
        return {**context, **result}

    def generate_companycam_quick_report(self, client: str, job_path: str,
                                         division: str, report_type: str,
                                         photo_ids: list, start_date: str = "",
                                         end_date: str = "", tag: str = "") -> dict:
        context = self.companycam_report_context(client, job_path, division)
        if not context.get("ok"):
            return context
        if not context.get("docs_folder"):
            return {**context, "ok": False,
                    "error": f"No {division} DOCS folder is available for this job."}
        import companycam_report
        result = companycam_report.generate(
            context["project_id"], client, context["docs_folder"], report_type,
            photo_ids or [], start_date=start_date, end_date=end_date, tag=tag)
        if result.get("ok"):
            self._document_cache.clear()
        return {**context, **result}

    def open_job_folder(self, client: str, path: str = "") -> dict:
        return self._audit_api().open_od_for_client(client, path)

    def list_job_folder_candidates(self, client: str, year: str = "") -> dict:
        """Visible folder-link flow for Pipeline cards; no context menu required."""
        return self._audit_api().list_folder_candidates(client, year)

    def link_job_folder(self, client: str, path: str,
                        confirm: bool = False) -> dict:
        result = self._audit_api().set_folder_path(client, path, confirm)
        if result.get("ok"):
            self._invalidate_workspace(client=client)
            self._board_view_cache = None
        return result

    def open_xa_link(self, client: str, card_id: str = "") -> bool:
        return self._audit_api().open_xa_link(client, card_id)

    def post_xa_note(self, client: str, note: str, tag: str = "",
                     card_id: str = "") -> dict:
        result = self._audit_api().post_xa_note(client, note, tag, card_id)
        if result.get("ok"):
            self._invalidate_workspace(client, card_id)
        return result

    def import_initial_notes(self, client: str, card_id: str = "") -> dict:
        return self._audit_api().import_initial_notes(client, card_id)

    def open_companycam_link(self, client: str) -> bool:
        return self._audit_api().open_companycam_link(client)

    def companycam_plan_pull(self, client: str, tech: str = "",
                             card_id: str = "",
                             dest_subfolder: str = "") -> dict:
        return self._audit_api().companycam_plan_pull(
            client, tech, card_id, dest_subfolder)

    def companycam_pull_assigned_bg(self, client: str, assignments: list,
                                    tech: str = "",
                                    card_id: str = "") -> dict:
        return self._audit_api().companycam_pull_assigned_bg(
            client, assignments, tech, card_id)

    def open_document(self, path: str) -> dict:
        path = os.path.abspath(path or "")
        if not path or not os.path.isfile(path):
            return {"ok": False, "error": "document is no longer in that folder"}
        try:
            os.startfile(path)
            return {"ok": True}
        except OSError as ex:
            return {"ok": False, "error": str(ex)}

    def list_pics_stages(self, client: str) -> dict:
        return self._audit_api().list_pics_stages(client)

    def copy_pics_to_clipboard(self, client: str, stage: str = "") -> dict:
        return self._audit_api().copy_pics_to_clipboard(client, stage)

    def save_crm_work_environment(self, client: str, work_environment: str,
                                  stage: str, owner: str = "") -> dict:
        result = self._audit_api().save_crm_work_environment(
            client, work_environment, stage, owner)
        if result.get("ok"):
            self._invalidate_workspace(client=client)
        return result

    def crm_division_trello_cards(self, client: str) -> dict:
        return self._audit_api().crm_division_trello_cards(client)

    def pin_crm_division_trello(self, client: str, division: str,
                                card_id_or_url: str) -> dict:
        result = self._audit_api().pin_crm_division_trello(
            client, division, card_id_or_url)
        if result.get("ok"):
            self._invalidate_workspace(client=client)
        return result

    def unpin_crm_division_trello(self, client: str, division: str) -> dict:
        result = self._audit_api().unpin_crm_division_trello(client, division)
        if result.get("ok"):
            self._invalidate_workspace(client=client)
        return result

    def set_job_check_item(self, card_id: str, item_id: str,
                           complete: bool) -> dict:
        local = pipeline_store.set_check_item(card_id, item_id, complete)
        synced = False
        error = ""
        try:
            import trello_client as tc
            synced = bool(tc.set_check_item_state(
                card_id, item_id, "complete" if complete else "incomplete"))
            if not synced:
                error = "Trello did not accept the checklist update"
        except Exception as ex:
            error = str(ex)
        if synced:
            pipeline_store.mark_card_sync(card_id, ok=True)
        elif local.get("ok"):
            pipeline_store.mark_card_sync(card_id, ok=False, error=error)
        result = {"ok": bool(local.get("ok")) or synced, "saved_local": bool(local.get("ok")),
                "synced": synced, "warning": error if local.get("ok") and not synced else "",
                "error": "" if local.get("ok") or synced else (local.get("error") or error)}
        if result.get("ok"):
            self._invalidate_workspace(card_id=card_id)
            self._board_view_cache = None
        return result

    def set_job_requirement(self, client: str, requirement_key: str,
                            state: str, note: str = "",
                            details: dict | None = None) -> dict:
        result = self._audit_api().set_job_requirement(
            client, requirement_key, state, note, details)
        if result.get("ok"):
            self._invalidate_workspace(client=client)
        return result

    def post_job_comment(self, client: str, card_id: str, text: str) -> dict:
        text = (text or "").strip()
        if not text:
            return {"ok": False, "error": "write a comment first"}
        actor = "Linguar Hub"
        actor_id = ""
        try:
            import supabase_client
            user = supabase_client.current_user() or {}
            actor_id = str(user.get("id") or "")
            if user.get("id") and not user.get("display_name"):
                return {"ok": False, "error":
                        "Add your name in My Settings before commenting."}
            actor = supabase_client.actor_name(actor)
        except Exception:
            pass
        posted_action = None
        error = ""
        if card_id:
            try:
                import trello_client as tc
                posted_action = tc.post_comment(card_id, text)
                if not posted_action:
                    error = "Trello did not accept the comment"
            except Exception as ex:
                error = str(ex)
        external_id = str(posted_action.get("id") or "") if isinstance(
            posted_action, dict) else ""
        local = pipeline_store.add_activity(
            card_id, "comment", text, actor, external_id=external_id,
            actor_id=actor_id)
        posted = bool(posted_action)
        result = {"ok": bool(local) or posted, "posted_trello": posted,
                "warning": error if local and not posted else "",
                "comment": {"id": local.get("activity_key") or "",
                            "external_id": external_id,
                            "text": text, "actor": actor,
                            "at": local.get("happened_at") or _dt.datetime.now().isoformat(),
                            "source": "linguar", "can_manage": bool(actor_id)}}
        if result.get("ok"):
            self._invalidate_workspace(client, card_id)
        return result

    def edit_job_comment(self, client: str, comment_id: str,
                         source: str, text: str,
                         external_id: str = "") -> dict:
        clean = (text or "").strip()
        if not clean:
            return {"ok": False, "error": "a comment cannot be empty"}
        actor_id = ""
        try:
            import supabase_client
            actor_id = str((supabase_client.current_user() or {}).get("id") or "")
        except Exception:
            pass
        if source == "trello":
            result = self._audit_api().update_card_comment(client, comment_id, clean)
            if result.get("ok"):
                self._invalidate_workspace(client=client)
            return result
        result = pipeline_store.update_activity(comment_id, clean, actor_id=actor_id)
        if not result.get("ok"):
            return result
        trello_id = external_id or result.get("external_id") or ""
        if trello_id:
            synced = self._audit_api().update_card_comment(client, trello_id, clean)
            if not synced.get("ok"):
                self._invalidate_workspace(client=client)
                return {**result, "text": clean, "warning":
                        "Saved in Linguar Hub, but Trello did not update"}
        self._invalidate_workspace(client=client)
        return {**result, "text": clean, "synced_trello": bool(trello_id)}

    def delete_job_comment(self, client: str, comment_id: str,
                           source: str, external_id: str = "") -> dict:
        actor_id = ""
        try:
            import supabase_client
            actor_id = str((supabase_client.current_user() or {}).get("id") or "")
        except Exception:
            pass
        if source == "trello":
            result = self._audit_api().delete_card_comment(client, comment_id)
            if result.get("ok"):
                self._invalidate_workspace(client=client)
            return result
        result = pipeline_store.delete_activity(comment_id, actor_id=actor_id)
        if not result.get("ok"):
            return result
        trello_id = external_id or result.get("external_id") or ""
        if trello_id:
            synced = self._audit_api().delete_card_comment(client, trello_id)
            if not synced.get("ok"):
                self._invalidate_workspace(client=client)
                return {**result, "warning":
                        "Deleted in Linguar Hub, but Trello did not delete"}
        self._invalidate_workspace(client=client)
        return {**result, "synced_trello": bool(trello_id)}

    def save_job_log_update(self, client: str, entry: dict,
                            card_id: str = "") -> dict:
        result = self._audit_api().save_crm_job_log(client, entry, card_id)
        if result.get("ok"):
            self._invalidate_workspace(client=client)
        return result

    def import_job_log_from_trello(self, client: str, card_id: str) -> dict:
        """Import recognized work events from the selected division card."""
        result = self._audit_api().import_crm_job_log_from_trello(client, card_id)
        if result.get("ok"):
            self._invalidate_workspace(client, card_id)
        return result

    def delete_job_log_update(self, client: str, entry_id: str,
                              card_id: str = "") -> dict:
        result = self._audit_api().delete_crm_job_log(client, entry_id, card_id)
        if result.get("ok"):
            self._invalidate_workspace(client=client)
        return result

    def job_log_update_history(self, entry_id: str) -> dict:
        return self._audit_api().crm_job_log_history(entry_id)

    def flag_missing_card(self, card_id: str, client: str,
                          item_text: str, note: str = "") -> dict:
        """Record a missing-item flag + post a Trello comment straight
        to this card (card id known — no pin lookup). Mirrors the
        audit/snapshot flag-missing flow; we post our own 🚩 comment so
        the tracker's auto-comment is suppressed (post_comment=False)."""
        if not client or not item_text:
            return {"ok": False, "error": "client + item required"}
        try:
            try:
                from missing_items_tracker import capture_missing_items
                capture_missing_items(
                    client, missing=[item_text], stage="audit", note=note,
                    card_id=card_id or "", post_comment=False)
            except Exception:
                pass
            posted = False
            if card_id:
                try:
                    import trello_client as tc
                    body = f"🚩 **Flagged missing**: {item_text}"
                    if note:
                        body += f"\n\n{note}"
                    tc.post_comment(card_id, body)
                    posted = True
                except Exception:
                    posted = False
            self._invalidate_workspace(client, card_id)
            return {"ok": True, "posted_trello": posted}
        except Exception as ex:
            return {"ok": False, "error": str(ex)}

    def attach(self, window):
        """Stash the window handle so background tasks can push events
        back to JS via window.evaluate_js()."""
        self._window = window

    # ── Threshold editor (P2) ────────────────────────────────────────
    def get_thresholds(self) -> dict:
        """Live thresholds dict + per-stage default for the editor UI."""
        live = ps.get_thresholds()
        return {
            "stages": [
                {"key": s, "label": ps.STAGE_LABELS.get(s, s),
                 "days": int(live.get(s) or 0),
                 "default": int(ps.DEFAULT_STAGE_THRESHOLDS.get(s) or 0)}
                for s in ps.STAGES
            ],
        }

    def set_threshold(self, stage: str, days) -> dict:
        """Save a per-stage override. `days=null` resets to default."""
        try:
            d = None if days in (None, "") else int(days)
            ps.set_threshold(stage, d)
            return {"ok": True}
        except Exception as ex:
            return {"ok": False, "error": str(ex)}

    def reset_thresholds(self) -> dict:
        """Reset all overrides — every stage falls back to its default."""
        try:
            for s in ps.STAGES:
                ps.set_threshold(s, None)
            return {"ok": True}
        except Exception as ex:
            return {"ok": False, "error": str(ex)}

    # ── 🕒 Per-card timeline (mirrors Tk _show_timeline) ────────────
    def card_timeline(self, card_id: str) -> dict:
        """Return the stage-transition history for a card, ordered
        oldest → newest. Each entry: from_stage, to_stage, when
        (ISO), days_in_from. Used by the per-row timeline modal."""
        if not card_id:
            return {"ok": False, "error": "card_id required",
                    "transitions": []}
        try:
            import ems_db as _db
            history = _db.list_transitions(card_id=card_id, order="ASC") or []
        except Exception as ex:
            return {"ok": False, "error": str(ex), "transitions": []}
        out = []
        for h in history:
            out.append({
                "from_stage":     h.get("from_stage") or "",
                "to_stage":       h.get("to_stage") or "",
                "when":           h.get("when") or h.get("transition_at") or "",
                "days_in_from":   h.get("days_in_from") or 0,
            })
        return {"ok": True, "card_id": card_id,
                "transitions": out,
                "total": len(out)}

    # ── 📊 Export to Excel (mirrors Tk _export_to_excel) ────────────
    def export_to_excel(self) -> dict:
        """Write the currently-loaded pipeline rows to a .xlsx in the
        user's downloads folder, one sheet per stage. Mirrors Tk's
        _export_to_excel — same column order, same stage-sheet split.
        Opens the file in Excel on success.
        """
        try:
            import datetime as _dt2
            import os as _os
            import pipeline_stages as _ps
            rows = self._last_rows or []
            if not rows:
                return {"ok": False,
                        "error": "no rows — sync from Trello first"}
            try:
                import openpyxl
            except ImportError:
                return {"ok": False,
                        "error": "openpyxl not installed"}
            today = _dt2.date.today()
            default_name = (f"Pipeline {today.month}-{today.day}-"
                            f"{today.year % 100}.xlsx")
            out_dir = _os.path.join(_os.path.expanduser("~"), "Downloads")
            if not _os.path.isdir(out_dir):
                out_dir = _os.path.expanduser("~")
            path = _os.path.join(out_dir, default_name)
            # Disambiguate if same-day file exists already
            i = 1
            while _os.path.isfile(path):
                stem = default_name[:-5]
                path = _os.path.join(out_dir, f"{stem} ({i}).xlsx")
                i += 1
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            headers = ("Client", "Stage", "Days in Stage", "Age",
                        "Last Activity", "Owner", "Board", "Lane", "URL")
            for stage in _ps.STAGES:
                sheet_rows = [r for r in rows
                               if (r.get("current_stage") or "") == stage]
                label = _ps.STAGE_LABELS.get(stage, stage)[:31] or stage[:31]
                ws = wb.create_sheet(title=label)
                ws.append(headers)
                for r in sheet_rows:
                    ws.append([
                        r.get("client_display") or r.get("client") or "",
                        _ps.STAGE_LABELS.get(stage, stage),
                        r.get("days_in_stage") or 0,
                        r.get("age_days") or 0,
                        r.get("last_activity_iso") or "",
                        r.get("owner") or "",
                        r.get("board_name") or "",
                        r.get("lane") or "",
                        r.get("card_url") or "",
                    ])
            if not wb.sheetnames:
                wb.create_sheet(title="Empty")
            wb.save(path)
            try: _os.startfile(path)
            except Exception: pass
            return {"ok": True, "path": path,
                    "rows": len(rows),
                    "stages": len(wb.sheetnames)}
        except Exception as ex:
            return {"ok": False, "error": f"{type(ex).__name__}: {ex}"}

    # ── Read-only data API ───────────────────────────────────────────
    def stages(self) -> list[dict]:
        """Stage labels in lifecycle order — drives the filter chips."""
        return [{"key": s, "label": ps.STAGE_LABELS.get(s, s)}
                for s in ps.STAGES]

    def lifecycle_view(self, force: bool = False) -> dict:
        """Return the complete Stages projection in one bounded DB pass.

        The old UI made three bridge calls and shaped every row by querying
        transition history again for each individual job.  On a hosted DB
        that turned a 2,000-row view into thousands of requests.  Cache the
        finished projection briefly; explicit lifecycle sync bypasses it.
        """
        now = time.monotonic()
        cached = self._lifecycle_view_cache
        if (not force and cached and now - cached[0] < 45):
            return cached[1]
        try:
            # Legacy cleanup is maintenance, not a prerequisite for paint.
            # The Trello sync already performs it; never put table-wide
            # repair/delete work in the user's read path.
            if not self._lifecycle_housekeeping_done:
                self._lifecycle_housekeeping_done = True

            rows = ems_db.lifecycle_list(paid_window_days=30)
            thresholds = ps.get_thresholds()
            try:
                history = ps.historical_stage_stats()
            except Exception:
                history = {}
            shaped = [_row_to_jsdict(r, thresholds=thresholds,
                                     history=history) for r in rows]
            counts = {}
            for row in shaped:
                stage = row.get("stage") or ""
                counts[stage] = counts.get(stage, 0) + 1
            result = {
                "ok": True,
                "stages": self.stages(),
                "rows": shaped,
                "counts": counts,
            }
            self._last_rows = shaped
            self._lifecycle_view_cache = (now, result)
            return result
        except Exception as ex:
            return {"ok": False,
                    "error": f"{type(ex).__name__}: {ex}",
                    "stages": self.stages(), "rows": [], "counts": {}}

    def lifecycle_rows(self) -> list[dict]:
        """Every active lifecycle row (plus paid within 30d), shaped
        for the frontend table. Read-only — no DB writes happen here.

        Also runs the self-heal backfill so legacy 0d-in-stage rows
        get their stage_entered_at corrected before the JS side
        renders day counts."""
        result = self.lifecycle_view()
        return result.get("rows", []) if result.get("ok") else []

    def stage_counts(self) -> dict:
        """{stage_key: count} for the filter chip badges."""
        result = self.lifecycle_view()
        return result.get("counts", {}) if result.get("ok") else {}

    # ── Actions ──────────────────────────────────────────────────────
    def open_url(self, url: str) -> bool:
        """Open `url` in the user's default browser. Used for the
        Trello card jump on row double-click."""
        if not url:
            return False
        try:
            webbrowser.open(url)
            return True
        except Exception:
            return False

    def copy_to_clipboard(self, text: str) -> bool:
        """Clipboard write via the Win32 clipboard. NOT a throwaway
        tk.Tk() — Tk's delayed-rendering clipboard left a dead owner that
        froze the next paste in any app ('clipboard freezes constantly')."""
        if not text:
            return False
        from web_helpers import set_clipboard_text
        return set_clipboard_text(text)

    def subcontractor_dispatch_draft(self, fields: dict,
                                     options: dict) -> dict:
        """Create—but never send—a subcontractor dispatch email draft."""
        try:
            import subcontractor_dispatch
            return subcontractor_dispatch.compose(fields, options)
        except Exception as ex:
            return {"ok": False, "error": f"Could not build draft: {ex}"}

    def sync_from_trello(self) -> dict:
        """Kick off a Trello workspace sync on a background thread.
        Returns immediately so the UI can show a spinner; progress +
        completion are pushed via window.evaluate_js() events."""
        if self._sync_running:
            return {"started": False, "reason": "sync already in progress"}
        self._sync_running = True

        def _bg():
            try:
                def _progress(i, n, board):
                    self._emit_js(
                        f"window.dispatchEvent(new CustomEvent("
                        f"'pipeline:sync-progress', "
                        f"{{detail: {{i: {i}, n: {n}, "
                        f"board: {self._jsstr(board)}}}}}));")
                result = ps.sync_workspace(progress_cb=_progress)
                self._emit_js(
                    f"window.dispatchEvent(new CustomEvent("
                    f"'pipeline:sync-done', "
                    f"{{detail: {{ok: true, "
                    f"cards: {result.get('cards', 0)}, "
                    f"boards: {result.get('boards', 0)}}}}}));")
            except Exception as ex:
                msg = self._jsstr(f"{type(ex).__name__}: {ex}")
                self._emit_js(
                    f"window.dispatchEvent(new CustomEvent("
                    f"'pipeline:sync-done', "
                    f"{{detail: {{ok: false, error: {msg}}}}}));")
            finally:
                self._sync_running = False
        _wh_run_bg(_bg)
        return {"started": True}

    # ── Bridge helpers ───────────────────────────────────────────────
    def _emit_js(self, js: str) -> None:
        # Forwards to BOTH the home shell window AND the iframe's
        # contentWindow so event listeners inside the embedded panel
        # actually receive the dispatch. Without this, sync events
        # fire silently and the UI looks frozen.
        try:
            import web_event
            web_event.dispatch(self._window, js)
        except Exception:
            pass

    def _jsstr(self, s: str) -> str:
        """Tiny JSON-string encoder for inline JS templates above."""
        import json
        return json.dumps(str(s))


def main(argv=None):
    api = Api()
    window = webview.create_window(
        title="Pipeline — Linguar Hub",
        url=INDEX_HTML,
        js_api=api,
        width=1280, height=820,
        min_size=(720, 500),
    )
    api.attach(window)
    webview.start(debug=False)


if __name__ == "__main__":
    main()
