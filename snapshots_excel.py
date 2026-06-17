"""SharePoint snapshots spreadsheet — source of truth for job state.

Per-year workbook at ``<root>\\Snapshots <YYYY>.xlsx`` with three tabs
matching the existing ``ServPro SS.xlsx`` template:

    NEW LOSS YY        — open jobs not yet closed
    Completed YY       — close-out done (audit shows OK)
    Incomplete YY      — cancelled / archived / 3rd-party / SVC-only
    Needs Attention YY — reconciler couldn't resolve the job (no matching
                         Trello card, or the card couldn't be classified)

Triggered from the EMS Snapshot tool's audit completion (one row per
audit result, idempotent upsert by case-insensitive Name match within
the year). Acts as a fail-safe + cross-team source of truth so anyone
can pull the .xlsx directly without driving the Snapshot tool.

Public entry points:
    sync_jobs(audit_results, *, year=None)         — run from snapshot
    sync_one(audit_row, *, year=None)              — single-job convenience
    read_jobs(year)                                — for the viewer dialog
    workbook_path(year)                            — UI display + open-in-Excel
    pending_count(year)                            — sidebar/badge

File-locked-in-Excel handling: writes go to a ``.pending\\`` sidecar
JSON queue; ``sync_jobs`` drains the queue first on each run, so a
locked file just means the row gets applied next time.
"""
import json
import os
import re
import threading
import time
from datetime import datetime

# openpyxl is the only realistic option on Windows without Office —
# pure-Python, reads + writes .xlsx, preserves formatting on round-trip.
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Constants ─────────────────────────────────────────────────────────────

# Default root — created by the user as agreed (X:\IE_Public\Snapshots).
# Override with snapshots_excel.set_root(path) if the share moves; the
# override persists in config.json under `snapshots_root`.
_DEFAULT_ROOT = r"X:\IE_Public\Front Operation\EMS Admin\Snapshots"
# In-process override (tests set this directly via set_root with no config
# available). When unset, get_root() resolves from config then default.
_root = None


def set_root(path):
    """Override the workbook root directory and PERSIST it to config so
    the change survives a restart. Used by the Snapshot panel's location
    control and by tests. Pass None / empty to revert to the default."""
    global _root
    _root = (path or "").strip() or None
    try:
        import config
        cfg = config.load()
        cfg["snapshots_root"] = _root or _DEFAULT_ROOT
        config.save(cfg)
    except Exception:
        # No config available (e.g. unit test) — the in-process override
        # above still applies for this run.
        pass


def get_root():
    """Resolve the workbook root: in-process override first, then the
    persisted `snapshots_root` config value, then the default."""
    if _root:
        return _root
    try:
        import config
        p = (config.load().get("snapshots_root") or "").strip()
        if p:
            return p
    except Exception:
        pass
    return _DEFAULT_ROOT


# Column order matches the existing ServPro SS template. Keep stable —
# changing this requires migrating every existing row.
COLUMNS = [
    "Date Received",
    "Name",
    "Closing Date",
    "Comment",
    "Type of Loss",
    "Claim#",
    "Carrier",
    "Folder",
    "Scheduled Ins.",
    "Lead",
    "Inspection",
    "Initial Photos",
    "Sketch",
    "Docusketch ordered?",
    "Demo Start",
    "Demo Photos",
    "Scope",
    "Final Photos",
    "ATP",
    "CIF",
    "CER",
    "COS",
    "Add'l Docs",
]
_COL_INDEX = {name: i + 1 for i, name in enumerate(COLUMNS)}

# Sheets are addressed as "<base> YY" — the template uses two-digit year
# suffixes (NEW LOSS 25, Completed 25, Incomplete 25). Stick to that so
# legacy rolls of the workbook keep working without name churn.
_SHEET_NEW        = "NEW LOSS"
_SHEET_COMPLETED  = "Completed"
_SHEET_INCOMPLETE = "Incomplete"
_SHEET_ATTENTION  = "Needs Attention"
_ALL_SHEETS = (_SHEET_NEW, _SHEET_COMPLETED, _SHEET_INCOMPLETE,
               _SHEET_ATTENTION)


def _sheet_name(base, year):
    """Build the per-year sheet title (e.g., "NEW LOSS 25"). Two-digit
    suffix matches the existing template; switching to four would
    invalidate every legacy reference in the workbook."""
    return f"{base} {year % 100:02d}"


# ── Path helpers ─────────────────────────────────────────────────────────

def workbook_path(year):
    """Absolute path to the workbook for the given year. The directory
    is NOT created here — that's `_ensure_dir`'s job at write time so
    callers can ask for the path even when the share is unavailable."""
    return os.path.join(get_root(), f"Snapshots {year}.xlsx")


def _pending_dir(year):
    return os.path.join(get_root(), ".pending")


def _ensure_dir(path):
    try:
        os.makedirs(path, exist_ok=True)
    except OSError:
        pass


def pending_count(year=None):
    """Number of queued rows waiting for the file to unlock. Surfaces
    in the viewer's status bar so the user knows when a write is
    deferred. Year arg unused for now (single queue dir) but kept so
    the signature can shard later without breaking callers."""
    pdir = _pending_dir(year)
    if not os.path.isdir(pdir):
        return 0
    try:
        return sum(1 for n in os.listdir(pdir) if n.endswith(".json"))
    except OSError:
        return 0


# ── Workbook plumbing ────────────────────────────────────────────────────

def _ensure_workbook(year):
    """Open the year's workbook, creating it (with the three tabs and
    a header row each) if it doesn't exist yet. Returns the
    (workbook, path) pair. Caller is responsible for saving."""
    path = workbook_path(year)
    _ensure_dir(os.path.dirname(path))
    if os.path.isfile(path):
        try:
            return openpyxl.load_workbook(path), path
        except Exception:
            # Corrupted workbook — back it up, start fresh. Rare but if
            # the file is truncated (interrupted save mid-flush) we
            # don't want to wedge the whole sync.
            backup = path + f".corrupt-{int(time.time())}"
            try:
                os.replace(path, backup)
            except OSError:
                pass

    wb = openpyxl.Workbook()
    # Workbook() seeds a "Sheet" — repurpose it as the first tab to
    # avoid a stray empty sheet at index 0.
    default_ws = wb.active
    wb.remove(default_ws)
    for base in _ALL_SHEETS:
        ws = wb.create_sheet(_sheet_name(base, year))
        _write_header(ws)
    return wb, path


def _write_header(ws):
    """Write the column header row + apply visual styling matching the
    feel of the template (bold black header on light fill)."""
    fill = PatternFill("solid", fgColor="DDEBF7")  # pale blue, easy on eyes
    font = Font(bold=True)
    for col_name, col_idx in _COL_INDEX.items():
        c = ws.cell(1, col_idx, col_name)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
    # Reasonable default widths — Name + Comment get more room since
    # they hold the longest free-text.
    widths = {"Name": 28, "Comment": 30, "Folder": 18, "Carrier": 14,
              "Type of Loss": 12, "Claim#": 18, "Lead": 12,
              "Date Received": 13, "Closing Date": 13, "Inspection": 13,
              "Scheduled Ins.": 13, "Demo Start": 13}
    for col_name, w in widths.items():
        idx = _COL_INDEX.get(col_name)
        if idx:
            ws.column_dimensions[get_column_letter(idx)].width = w
    # All the yes/no columns — narrow.
    for col_name in ("Initial Photos", "Sketch", "Docusketch ordered?",
                      "Demo Photos", "Scope", "Final Photos",
                      "ATP", "CIF", "CER", "COS", "Add'l Docs"):
        idx = _COL_INDEX.get(col_name)
        if idx:
            ws.column_dimensions[get_column_letter(idx)].width = 11
    ws.freeze_panes = "A2"


def _canon_name_key(name):
    """Canonical key for Name-cell matching. Handles the form
    mismatches that produced duplicate rows in the wild:
      • Case folding ('Smith, John' == 'smith, john')
      • Whitespace collapse ('Rand  Melissa' == 'Rand Melissa')
      • Comma-swap ('Jacqueline Sanchez' == 'Sanchez, Jacqueline')

    Deliberately does NOT strip a trailing " - X" suffix the way the
    Trello-pin canonicalization does. In the spreadsheet, " - X" is
    almost always a sub-job identifier (commercial unit, tenant
    name) rather than a carrier suffix — "Action Property Management"
    and "Action Property Management - Mendiola, Mary (97820)" are
    distinct jobs and must NOT collapse. The Trello-pin canon can be
    aggressive because it's matching against a single-job dict; the
    spreadsheet canon has to preserve sub-job distinctness.

    Single source of truth for `_find_row_by_name`, `_build_name_index`,
    and `dedupe_workbook` so a row written from any audit / snapshot /
    IUQ surface lands on the same existing record.

    ── Sibling canon functions (don't confuse) ──────────────────────
    `persistence._canon_pin_key` → strips " - Carrier" suffix; use for
        Trello-pin / folder-path / APA-cross-reference lookups.
    `ems_db.canon_key` → alias of `_canon_pin_key`.
    `initial_upload_queue._norm_client_for_run_doc` → comma-swap only,
        used for fuzzy matching against today's run-doc text.
    Use THIS function when: writing a row into Snapshots.xlsx /
    Dispute Tracker / Estimate Requests workbooks.
    """
    s = (name or "").strip()
    if not s:
        return ""
    # Collapse internal whitespace runs to a single space + case fold.
    s = " ".join(s.split()).lower()
    if "," in s:
        parts = [p.strip() for p in s.split(",", 1)]
        if len(parts) == 2 and parts[0] and parts[1]:
            s = " ".join(f"{parts[1]} {parts[0]}".split())
    return s


def _find_row_by_name(ws, name):
    """Match a Name cell to an existing row via `_canon_name_key`.
    Returns the row number or None. Skips the header.

    Earlier versions did only a case-folded + trimmed compare, which
    missed comma-swap variants ("Sanchez, Jacqueline" vs "Jacqueline
    Sanchez") and internal-whitespace variants ("Rand  Melissa" vs
    "Rand Melissa") — both produced duplicate rows.
    """
    needle = _canon_name_key(name)
    if not needle:
        return None
    name_col = _COL_INDEX["Name"]
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, name_col).value
        if v is None:
            continue
        if _canon_name_key(str(v)) == needle:
            return r
    return None


def _build_name_index(wb, year):
    """One-pass index of {canon_name: (sheet_title, row_num)} across the
    year's three sheets. Reconcile-style operations look up the same
    rows repeatedly; `_find_row_by_name` is O(N) per call and the loop
    becomes O(N*M). Build this once at the top, refresh affected sheets
    after `_move_row_between_sheets` (which deletes a row and so
    shifts indices).

    Keys use `_canon_name_key` so comma-swap / whitespace-variant rows
    hash to the same slot — looking up "Sanchez, Jacqueline" finds
    "Jacqueline Sanchez" and vice versa.
    """
    index = {}
    name_col = _COL_INDEX["Name"]
    for base in _ALL_SHEETS:
        title = _sheet_name(base, year)
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, name_col).value
            if v is None:
                continue
            key = _canon_name_key(str(v))
            if key:
                index[key] = (title, r)
    return index


def _refresh_index_for_sheet(wb, year, base, index):
    """Rebuild the index entries for one sheet after a row delete /
    insert. Cheaper than a full `_build_name_index` rebuild."""
    title = _sheet_name(base, year)
    if title not in wb.sheetnames:
        # Drop any stale entries that pointed at this title
        for k in [k for k, (t, _) in index.items() if t == title]:
            index.pop(k, None)
        return
    ws = wb[title]
    name_col = _COL_INDEX["Name"]
    # Purge old entries for this sheet first
    for k in [k for k, (t, _) in index.items() if t == title]:
        index.pop(k, None)
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, name_col).value
        if v is None:
            continue
        key = _canon_name_key(str(v))
        if key:
            index[key] = (title, r)


def _move_row_between_sheets(wb, year, name, target_base):
    """If `name` exists in any sheet OTHER than the target, copy its
    cells into the target sheet (preserving any free-text comments the
    user typed) and delete the source row. Returns True on a move.

    Without this, routing a job from NEW LOSS → Completed would leave
    a stale row in NEW LOSS and a fresh row in Completed — so the
    spreadsheet would say two contradictory things about the same job.
    """
    target_title = _sheet_name(target_base, year)
    # Create the target sheet on demand so callers (e.g. the Tracked-tab
    # manual move into a brand-new "Needs Attention" sheet) don't have to
    # pre-create it. No-op when it already exists.
    if target_title not in wb.sheetnames:
        _write_header(wb.create_sheet(target_title))
    moved = False
    for base in _ALL_SHEETS:
        if base == target_base:
            continue
        title = _sheet_name(base, year)
        if title not in wb.sheetnames:
            continue
        src = wb[title]
        r = _find_row_by_name(src, name)
        if r is None:
            continue
        # Pull the row into a dict so we can re-stamp into the target.
        existing = {COLUMNS[i]: src.cell(r, i + 1).value
                    for i in range(len(COLUMNS))}
        dst = wb[target_title]
        dst_r = _find_row_by_name(dst, name)
        if dst_r is None:
            dst_r = (dst.max_row or 1) + 1
            for col_name, col_idx in _COL_INDEX.items():
                dst.cell(dst_r, col_idx, existing.get(col_name))
        else:
            # Target already had a row too (rare race) — fill missing
            # cells from the source row but don't overwrite anything
            # the target already knows.
            for col_name, col_idx in _COL_INDEX.items():
                if dst.cell(dst_r, col_idx).value in (None, "", "-"):
                    dst.cell(dst_r, col_idx, existing.get(col_name))
        src.delete_rows(r, 1)
        moved = True
    return moved


# ── Audit row → spreadsheet row mapping ──────────────────────────────────
#
# The audit dict has photo_issues / form_issues as strings; we map them
# to yes/no/- per-column. Best-effort — preserve any user-typed value
# already in the cell when our evidence is weaker than theirs.

# Substrings (lowercased) that mark a tracked column's asset as MISSING.
# CRITICAL: these must match the EXACT issue labels audit_logic emits, not
# the column abbreviation. `check_forms` returns display names like
# "Auth to Perform" / "Customer Info Form" (never "ATP"/"CIF"), and
# `check_photos` returns "Demo pics" / "Initial pics" (never "demo photo").
# The old hints ("ATP", "initial photo", …) matched NONE of those labels,
# so every column silently defaulted to "yes" — the source of the
# "it says we have demo/scope/final photos when we don't" bug.
_MISSING_MARKERS = {
    "Initial Photos": ("initial pics", "initial photo", "missing initial"),
    "Demo Photos":    ("demo pics", "demo photo", "missing demo"),
    "Final Photos":   ("final pics", "final photo", "missing final"),
    "Sketch":         ("docusketch", "sketch"),
    # check_forms emits the full display name; match those plus the
    # abbreviation as a belt-and-suspenders fallback.
    "Scope":          ("scope",),
    "ATP":            ("auth to perform", "auth. to perform", "atp"),
    "CIF":            ("customer info", "cif"),
    # NOTE: no bare "cer" — it's a substring of "Cert of Satisfaction"
    # (COS) and would false-flag CER whenever COS is missing.
    "CER":            ("customer equip", "equip resp"),
    "COS":            ("cert of satisf", "cert of satisfaction"),
}

# Folder-name patterns under the job's PICS dir that POSITIVELY prove a
# photo stage exists on disk. Only these three columns get a "yes" from
# real files — the audit never emits a reliable "present" signal for them
# (it only flags a stage missing when the run-doc mentioned it, so "not
# flagged" never means "we have it").
_PHOTO_PRESENCE_PATTERNS = {
    "Initial Photos": r'\b(initial|inspection)\b',
    "Demo Photos":    r'\bdemo\b',
    "Final Photos":   r'\bfinal\b',
}

# Columns the audit can affirmatively VERIFY present (check_forms /
# check_docusketch actually look for the asset), so when an audit ran and
# didn't flag them, they're genuinely present → "yes".
_AUDIT_VERIFIED_COLS = ("Sketch", "Scope", "ATP", "CIF", "CER", "COS")


def _flagged_missing(column, photo_issues, form_issues):
    """True when the audit flagged this column's asset as missing."""
    markers = _MISSING_MARKERS.get(column, ())
    for s in ((photo_issues or []) + (form_issues or [])):
        s = str(s).lower()
        if any(m in s for m in markers):
            return True
    return False


def _photo_evidence(r):
    """Scan the job's actual photos folder and return the set of photo
    columns ("Initial Photos" / "Demo Photos" / "Final Photos") that have
    REAL evidence on disk — a non-empty stage folder or a matching loose
    photo. "...get that from the files..." — these cells must reflect what
    was actually collected, never a default "yes".

    Resolves the folder from r['path'] (audit rows carry it) or, failing
    that, a pinned OD folder for the client (mark_completed rows). Returns
    an empty set when no folder resolves, so the caller leaves those cells
    to the existing value rather than asserting anything."""
    path = (r.get("path") or "").strip()
    if not path:
        try:
            import persistence
            path = persistence.get_folder_path(r.get("client") or "") or ""
        except Exception:
            path = ""
    if not path or not os.path.isdir(path):
        return set()
    # Mirror audit_logic: photos live under EMS/PICS, CONTENTS/PICS, or a
    # bare PICS/Photos at the job root.
    base = path
    for sub in ("EMS", "CONTENTS"):
        cand = os.path.join(path, sub)
        if os.path.isdir(cand):
            base = cand
            break
    pics = os.path.join(base, "PICS")
    if not os.path.isdir(pics):
        alt = os.path.join(base, "Photos")
        if os.path.isdir(alt):
            pics = alt
    try:
        entries = os.listdir(pics)
    except OSError:
        return set()
    present = set()
    for col, pat in _PHOTO_PRESENCE_PATTERNS.items():
        for e in entries:
            if not re.search(pat, e, re.IGNORECASE):
                continue
            full = os.path.join(pics, e)
            if os.path.isdir(full):
                try:
                    with os.scandir(full) as it:
                        if any(True for _ in it):
                            present.add(col)
                            break
                except OSError:
                    pass
            elif os.path.isfile(full):
                present.add(col)
                break
    return present


def _evidence_cell(column, r, existing, *, present, audited):
    """Decide one tracked yes/no cell from real evidence — never a default.

       - column proven on disk / audit-verified present → "yes"
       - audit flagged it required-but-missing           → "no"
       - no evidence either way → keep the user's existing cell, else "-"

    This replaces the old "absence of a flagged issue = yes" rule, which
    fabricated "yes" for stages a job never reached and for every column
    on rows (e.g. snapshot-generate / mark_completed) that carry no audit
    data at all."""
    photo_issues = r.get("photo_issues") or []
    form_issues  = r.get("form_issues") or []
    if column in _PHOTO_PRESENCE_PATTERNS:
        if column in present:
            return "yes"
        if _flagged_missing(column, photo_issues, form_issues):
            return "no"
    else:  # audit-verifiable column (Sketch / Scope / ATP / CIF / CER / COS)
        if _flagged_missing(column, photo_issues, form_issues):
            return "no"
        if audited:
            return "yes"
    prev = existing.get(column)
    return prev if (isinstance(prev, str) and prev.strip()) else "-"


_ROUTE_HINT_TO_SHEET = {
    "completed":       _SHEET_COMPLETED,
    "incomplete":      _SHEET_INCOMPLETE,
    "new_loss":        _SHEET_NEW,
    "needs_attention": _SHEET_ATTENTION,
}


def _route_for(r):
    """Decide which sheet the job belongs in.

    Priority order:
      1. r["_route_override"]  — explicit hint set by the caller
         (e.g., snapshot generate-PDF sets "completed"; the Trello
         reconciler sets the result of `card_route_status`).
      2. Comment marks it cancelled / archived       → Incomplete
      3. r["flagged"] OR r["new_loss"]               → NEW LOSS
      4. would-be Completed AND no Claim# filed      → Incomplete
      5. else                                        → Completed

    The "cancelled" check looks at the audit's free-text fields plus
    the existing spreadsheet Comment cell (from `_existing_comment`)
    so a hand-typed "Cancelled" in Excel routes future syncs to the
    Incomplete tab without us re-flagging it as a new loss.

    The Claim# guard (rule 4) catches rows that would otherwise drift
    into Completed despite never having a claim filed — the user wants
    those surfaced in Incomplete as a follow-up queue, not buried in
    Completed. NEW LOSS rows are exempt because a missing claim is the
    *expected* state at intake; only the would-be-Completed path gets
    demoted.
    """
    override = (r.get("_route_override") or "").strip().lower()
    if override in _ROUTE_HINT_TO_SHEET:
        return _ROUTE_HINT_TO_SHEET[override]
    comment = (r.get("_existing_comment") or "").lower()
    if any(k in comment for k in ("cancel", "3rd party", "third party",
                                    "archived", "archive")):
        return _SHEET_INCOMPLETE
    if r.get("flagged"):
        return _SHEET_NEW
    if r.get("new_loss"):
        return _SHEET_NEW
    has_claim = bool((str(r.get("_claim") or "").strip())
                     or (str(r.get("_existing_claim") or "").strip()))
    if not has_claim:
        return _SHEET_INCOMPLETE
    return _SHEET_COMPLETED


def _row_to_cells(r, *, existing=None):
    """Build the {column → value} dict to write. Pulls audit data and,
    when an `existing` row dict is provided (read from the workbook),
    keeps any user-typed free-text the user has put in cells we can't
    confidently overwrite."""
    existing = existing or {}
    folder = (r.get("folder") or os.path.basename(r.get("path") or "")).strip()
    cells = {}
    # Date Received: existing cell wins; else the caller-provided
    # `_date_received` (Trello desc PROPERTY DETAILS, falling back to
    # card creation date). The old `_date_received_for(r)` heuristic
    # was removed 2026-05-07 — it was scanning folder-name prefixes
    # like "1-2-25 Smith" and slapping random dates on rows whose
    # folder happened to start with a date pattern.
    cells["Date Received"] = (existing.get("Date Received")
                               or r.get("_date_received")
                               or "")
    cells["Name"] = (r.get("client") or "").strip()
    # Closing Date: existing cell wins; else caller-provided
    # `_closing_date` (set when the Trello card was archived or moved
    # to LOGS - EMS). When neither source has a value we leave the
    # cell blank — better an empty cell than a fabricated date the
    # user has to hunt down and correct.
    if existing.get("Closing Date"):
        cells["Closing Date"] = existing.get("Closing Date")
    elif r.get("_closing_date"):
        cells["Closing Date"] = r["_closing_date"]
    else:
        cells["Closing Date"] = ""
    # Free-text: never clobber what the user typed.
    cells["Comment"]      = existing.get("Comment", "") or ""
    # Type of Loss: prefer existing cell; fall back to a caller-provided
    # value (e.g., the Trello card label resolved by the reconciler).
    cells["Type of Loss"] = (existing.get("Type of Loss")
                              or r.get("_type_of_loss")
                              or "")
    # Carrier / Claim# also accept caller-provided fallbacks (parsed
    # from the Trello card desc) so a fresh row gets populated without
    # waiting for the user to type them. Existing cell wins so manual
    # corrections stick.
    cells["Claim#"]       = (existing.get("Claim#")
                              or r.get("_claim")
                              or "")
    cells["Carrier"]      = (existing.get("Carrier")
                              or r.get("_carrier")
                              or "")
    cells["Folder"]       = folder or existing.get("Folder", "")
    cells["Scheduled Ins."] = existing.get("Scheduled Ins.", "") or ""
    cells["Lead"]         = existing.get("Lead", "") or ""
    cells["Inspection"]   = existing.get("Inspection", "") or ""
    # Audit-derived yes/no columns. Decided from real evidence — photos
    # from the files on disk, forms from what the audit actually verified
    # — never a blanket "yes". A row that ran no audit (e.g. snapshot
    # generate → mark_completed) carries no form_issues key; we treat it
    # as un-audited so its yes/no cells are preserved/"-" rather than
    # fabricated. See _evidence_cell / _photo_evidence.
    audited = ("form_issues" in r) or ("photo_issues" in r) or bool(r.get("path"))
    present = _photo_evidence(r)
    cells["Initial Photos"]      = _evidence_cell("Initial Photos", r, existing,
                                                  present=present, audited=audited)
    cells["Sketch"]              = _evidence_cell("Sketch", r, existing,
                                                  present=present, audited=audited)
    cells["Docusketch ordered?"] = existing.get("Docusketch ordered?", "") or ""
    cells["Demo Start"]          = existing.get("Demo Start", "") or ""
    cells["Demo Photos"]         = _evidence_cell("Demo Photos", r, existing,
                                                  present=present, audited=audited)
    cells["Scope"]               = _evidence_cell("Scope", r, existing,
                                                  present=present, audited=audited)
    cells["Final Photos"]        = _evidence_cell("Final Photos", r, existing,
                                                  present=present, audited=audited)
    cells["ATP"] = _evidence_cell("ATP", r, existing, present=present, audited=audited)
    cells["CIF"] = _evidence_cell("CIF", r, existing, present=present, audited=audited)
    cells["CER"] = _evidence_cell("CER", r, existing, present=present, audited=audited)
    cells["COS"] = _evidence_cell("COS", r, existing, present=present, audited=audited)
    cells["Add'l Docs"] = existing.get("Add'l Docs", "") or ""
    return cells


# ── Pending sidecar (Excel-locked fallback) ──────────────────────────────

def _is_locked(path):
    """Best-effort check: if Excel has the workbook open it holds an
    advisory write lock. We can't acquire it on Windows for a quick
    test without third-party libs, so we use os.rename to a same-name
    no-op as a write-permission probe — fails fast on locked files."""
    if not os.path.isfile(path):
        return False
    try:
        # Try to open for append. Excel's lock blocks this.
        with open(path, "ab"):
            return False
    except OSError:
        return True


def _enqueue_pending(year, audit_rows):
    """Persist the pending rows to a sidecar JSON. Each call writes a
    new file (timestamp-named) so multiple deferred syncs don't
    overwrite each other's queues."""
    pdir = _pending_dir(year)
    _ensure_dir(pdir)
    fname = f"{int(time.time() * 1000)}-{year}.json"
    payload = {"year": year, "rows": [_row_for_persist(r) for r in audit_rows]}
    try:
        with open(os.path.join(pdir, fname), "w", encoding="utf-8") as f:
            json.dump(payload, f, default=str)
    except OSError:
        pass


def _row_for_persist(r):
    """Compact subset of the audit row needed to replay a sync. We
    don't need the full dict — just the fields `_row_to_cells` and
    `_route_for` consult."""
    return {
        "client":       r.get("client"),
        "folder":       r.get("folder"),
        "path":         r.get("path"),
        "flagged":      bool(r.get("flagged")),
        "new_loss":     bool(r.get("new_loss")),
        "form_issues":  list(r.get("form_issues") or []),
        "photo_issues": list(r.get("photo_issues") or []),
        "last_active":  r.get("last_active"),
    }


def _drain_pending(year):
    """Replay any queued pending JSONs whose year matches. Called at
    the top of `sync_jobs` so the recovery path is automatic — no
    user intervention needed when the file unlocks."""
    pdir = _pending_dir(year)
    if not os.path.isdir(pdir):
        return []
    drained = []
    try:
        names = sorted(os.listdir(pdir))
    except OSError:
        return []
    for n in names:
        if not n.endswith(".json"):
            continue
        full = os.path.join(pdir, n)
        try:
            with open(full, "r", encoding="utf-8") as f:
                payload = json.load(f)
        except (OSError, json.JSONDecodeError):
            continue
        if payload.get("year") != year:
            continue
        drained.extend(payload.get("rows") or [])
        try:
            os.remove(full)
        except OSError:
            pass
    return drained


# ── Public API ───────────────────────────────────────────────────────────

_LOCK = threading.Lock()


def sync_jobs(audit_results, *, year=None, on_done=None):
    """Sync a list of audit results into the year's workbook. Runs in
    a background thread — never blocks the caller's UI. Idempotent by
    (year, name): existing rows update in place, missing rows insert,
    rows can move between sheets when their state changes.

    `on_done(applied, deferred)` — optional callback fired on the UI
    thread (Tk-style `after(0, ...)` style not available here, but
    callers can wrap the callback themselves) once the sync settles.
    `applied` is the count of rows actually written; `deferred` is the
    count queued for later because the file was locked.
    """
    rows = list(audit_results or [])
    yr = year or datetime.today().year

    def _go():
        applied, deferred = 0, 0
        with _LOCK:
            try:
                applied, deferred = _apply_sync(yr, rows)
            except Exception:
                # Don't let a bad sync take down the snapshot tool —
                # log via ems_log if available, otherwise swallow.
                try:
                    import ems_log
                    ems_log.warn("snapshots_excel",
                                  f"sync_jobs failed for year {yr}",
                                  exc_info=True)
                except Exception:
                    pass
        if on_done:
            try:
                on_done(applied, deferred)
            except Exception:
                pass

    threading.Thread(target=_go, daemon=True).start()


def sync_one(audit_row, *, year=None, on_done=None):
    """Convenience for the snapshot tool: sync just the current
    insured's audit row. Equivalent to `sync_jobs([row])`."""
    sync_jobs([audit_row], year=year, on_done=on_done)


def _apply_sync(year, audit_rows):
    """The actual workbook mutation. Caller must hold `_LOCK`. Returns
    (applied_count, deferred_count). When the workbook file is locked,
    every row is enqueued — `applied` is 0 in that case."""
    # Drain any prior deferrals first so they get replayed before
    # today's rows. Pending rows look like persisted dicts; today's
    # rows are full audit dicts; both flow through the same write
    # path because `_row_for_persist`'s output is a strict subset.
    drained = _drain_pending(year)
    queue = drained + list(audit_rows or [])
    if not queue:
        return (0, 0)

    path = workbook_path(year)
    if _is_locked(path):
        # Re-enqueue the full queue (drained + new), not just `audit_rows`
        # — otherwise the drained-but-not-yet-applied rows get dropped on
        # the floor each time the file is locked.
        _enqueue_pending(year, queue)
        return (0, len(queue))

    wb, path = _ensure_workbook(year)
    applied = 0
    for r in queue:
        try:
            _apply_one(wb, year, r)
            applied += 1
        except Exception:
            # Skip the bad row — don't abort the whole batch.
            continue
    try:
        wb.save(path)
    except PermissionError:
        # Race: file got opened in Excel between our probe and save.
        # Re-queue the FULL queue (drained + new) — the save was never
        # durable, so drained rows still need replay too.
        _enqueue_pending(year, queue)
        return (0, len(queue))
    return (applied, 0)


def _apply_one(wb, year, r):
    """Upsert a single row into the right sheet, moving from another
    sheet if the routing changed. Read-then-write so user free-text
    is preserved."""
    name = (r.get("client") or "").strip()
    if not name:
        return

    # First, learn what the spreadsheet already knows about this name —
    # so we can keep user-typed Comment/Carrier/Claim#/etc. The route
    # decision can also depend on the existing comment ("Cancelled").
    existing = {}
    for base in _ALL_SHEETS:
        title = _sheet_name(base, year)
        if title not in wb.sheetnames:
            ws = wb.create_sheet(title)
            _write_header(ws)
            continue
        ws = wb[title]
        rr = _find_row_by_name(ws, name)
        if rr is not None:
            for i, col in enumerate(COLUMNS):
                v = ws.cell(rr, i + 1).value
                if v not in (None, "") and col not in existing:
                    existing[col] = v
            break  # first hit wins — name should be unique across sheets
    r["_existing_comment"] = existing.get("Comment", "")
    # Surface the workbook's Claim# to _route_for so the missing-claim
    # demotion to Incomplete can fire without us re-reading the workbook.
    r["_existing_claim"]   = existing.get("Claim#", "") or ""

    target_base = _route_for(r)
    target_title = _sheet_name(target_base, year)
    if target_title not in wb.sheetnames:
        ws = wb.create_sheet(target_title)
        _write_header(ws)

    # Move the row if it lived in a different sheet.
    _move_row_between_sheets(wb, year, name, target_base)

    ws = wb[target_title]
    cells = _row_to_cells(r, existing=existing)
    rr = _find_row_by_name(ws, name)
    if rr is None:
        rr = (ws.max_row or 1) + 1
    for col_name, col_idx in _COL_INDEX.items():
        v = cells.get(col_name, "")
        ws.cell(rr, col_idx, v)


def read_jobs(year):
    """Read every row across the three sheets for the year. Returns
    a list of dicts with columns + a `_sheet` key naming the source
    sheet (so the viewer can colour-code by sheet)."""
    path = workbook_path(year)
    out = []
    if not os.path.isfile(path):
        return out
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return out
    for base in _ALL_SHEETS:
        title = _sheet_name(base, year)
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            name = row[_COL_INDEX["Name"] - 1] if len(row) >= _COL_INDEX["Name"] else None
            if not name or not str(name).strip():
                continue
            d = {"_sheet": base}
            for i, col in enumerate(COLUMNS):
                d[col] = row[i] if i < len(row) else None
            out.append(d)
    try:
        wb.close()
    except Exception:
        pass
    return out


def _resolve_card_for_client(name):
    """Look up the Trello card id for a client. Pinned ids win; fall back
    to fuzzy-search across `client_search_terms` (canonical name plus
    user-registered aliases). Returns `(card_id_or_None, error_str_or_None)`
    — the error string is non-None only when a search call raised, so
    callers like the reconciler can collect them for the summary.

    Extracted so `reconcile_with_trello` and `_build_trello_enrichment`
    don't drift apart on lookup logic.
    """
    name = (name or "").strip()
    if not name:
        return (None, None)
    try:
        import trello_client as _tc
        import persistence as _p
    except Exception:
        return (None, None)

    try:
        pinned = _p.get_trello_card_ids(name) or []
        if pinned:
            return (pinned[0], None)
    except Exception:
        pass

    try:
        terms = _p.client_search_terms(name)
    except Exception:
        terms = [name]
    last_err = None
    for term in terms:
        try:
            hits = _tc.find_cards_by_name(term, max_results=1)
        except Exception as ex:
            last_err = f"{name}: search failed ({ex})"
            continue
        if hits:
            return (hits[0]["card_id"], None)
    return (None, last_err)


def reconcile_with_trello(year, *, progress_cb=None):
    """One-shot reconciliation: walk every row in the year's workbook,
    look up the linked Trello card per row (pinned card id first, then
    fuzzy name search), and rewrite each row to:
      - the sheet that matches `trello_client.card_route_status`
      - the Type of Loss derived from `trello_client.card_loss_type`
        (only when the cell was empty — never overwrites a manual entry).

    Returns a dict:
        {"moved": int, "loss_type_set": int,
         "skipped_no_match": int, "errors": [str, ...]}

    Idempotent — safe to re-run. Rows whose Trello card can't be found
    are left in their current sheet (skipped). Free-text columns are
    preserved; only the route + Type of Loss change.

    Workbook is loaded for write, so Excel must be closed. When the file
    is locked we abort with an error rather than queuing — the user
    explicitly requested this action and should retry after closing.
    """
    import trello_client as _tc

    summary = {"moved": 0, "loss_type_set": 0,
               "skipped_no_match": 0, "errors": [],
               # Diagnostic breakdown so the user can SEE the routing
               # outcome — e.g. "12 stayed NEW LOSS because their card
               # is still open / on a non-LOGS board".
               "by_target": {"completed": 0, "incomplete": 0,
                             "new_loss": 0, "needs_attention": 0},
               "stayed_new_loss": [],
               # Jobs the reconciler couldn't resolve (no matching card /
               # unclassifiable) — moved to the Needs Attention sheet.
               "needs_attention": [],
               # Jobs newly inserted into NEW LOSS from the Initial Upload
               # Queue (weren't in the workbook yet).
               "added_new_loss": 0, "added_names": []}
    # idList → lane name, cached so the terminal-lane close check costs at
    # most one extra Trello call per distinct lane across the whole run.
    list_name_cache = {}

    yr = year or datetime.today().year

    path = workbook_path(yr)
    if _is_locked(path):
        summary["errors"].append(
            "Workbook is open in Excel — close it and retry.")
        return summary

    with _LOCK:
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as ex:
            summary["errors"].append(f"Open failed: {ex}")
            return summary

        # Flush any pending sidecar rows into the workbook first — otherwise
        # they're invisible to the reconcile loop and the next sync_jobs
        # would re-route them through the old routing rules. The drain is
        # done inside the lock + before reading rows, so the in-memory wb
        # picks up the writes.
        for r in _drain_pending(yr):
            try:
                _apply_one(wb, yr, r)
            except Exception as ex:
                summary["errors"].append(f"pending replay: {ex}")

        # Read rows from the IN-MEMORY workbook (not read_jobs(), which
        # re-opens from disk and misses anything we just drained).
        rows = []
        for base in _ALL_SHEETS:
            title = _sheet_name(base, yr)
            if title not in wb.sheetnames:
                continue
            ws = wb[title]
            for r in range(2, ws.max_row + 1):
                name_val = ws.cell(r, _COL_INDEX["Name"]).value
                if not name_val or not str(name_val).strip():
                    continue
                rows.append({"Name": str(name_val).strip(), "_sheet": base})
        if not rows:
            summary["errors"].append(f"No rows found in {yr} workbook.")
            return summary

        # Single-pass index so we don't re-scan the workbook for every row.
        name_index = _build_name_index(wb, yr)

        # ── Add-missing: every job in the Initial Upload Queue that isn't
        # already in the workbook gets inserted into NEW LOSS, so every
        # job we do the initial upload for is tracked from the start
        # (user-requested 2026-06-11). Dedup by canonical name against
        # what's already in any sheet. Best-effort — a Trello/IUQ hiccup
        # never blocks the move-closed reconcile below.
        try:
            import initial_upload_queue as _iuq
            iuq_cards = _iuq._fetch_queue_from_trello() or []
        except Exception as ex:
            iuq_cards = []
            summary["errors"].append(f"IUQ fetch failed: {ex}")
        seen_canon = set(name_index.keys())
        for c in iuq_cards:
            nm = ((c.get("client") or c.get("name") or "")
                  if isinstance(c, dict) else "").strip()
            if not nm:
                continue
            ck = _canon_name_key(nm)
            if ck in seen_canon:
                continue
            try:
                _apply_one(wb, yr, {"client": nm,
                                    "_route_override": "new_loss"})
                seen_canon.add(ck)
                summary["added_new_loss"] = summary.get("added_new_loss", 0) + 1
                if len(summary.setdefault("added_names", [])) < 50:
                    summary["added_names"].append(nm)
            except Exception as ex:
                summary["errors"].append(f"add {nm}: {ex}")
        # New rows aren't in `rows` (read before the inserts), so the
        # move-closed loop below won't reprocess them — they correctly
        # stay on NEW LOSS as fresh, still-open jobs.

        def _to_needs_attention(row, name, reason):
            """Move an unresolvable row to the Needs Attention sheet (no
            matching Trello card, or a card we couldn't classify). Keeps
            the diagnostics + index in sync, mirroring the main move path.
            Idempotent — a row already on the sheet is left in place."""
            summary["by_target"]["needs_attention"] = (
                summary["by_target"].get("needs_attention", 0) + 1)
            if len(summary["needs_attention"]) < 50:
                summary["needs_attention"].append(f"{name} — {reason}")
            if row.get("_sheet") == _SHEET_ATTENTION:
                return
            target_title = _sheet_name(_SHEET_ATTENTION, yr)
            if target_title not in wb.sheetnames:
                _write_header(wb.create_sheet(target_title))
            source_base = row.get("_sheet")
            _move_row_between_sheets(wb, yr, name, _SHEET_ATTENTION)
            summary["moved"] += 1
            if source_base:
                _refresh_index_for_sheet(wb, yr, source_base, name_index)
            _refresh_index_for_sheet(wb, yr, _SHEET_ATTENTION, name_index)

        total = len(rows)
        for i, row in enumerate(rows, 1):
            name = (row.get("Name") or "").strip()
            if not name:
                continue
            if progress_cb:
                try:
                    progress_cb(i, total, name)
                except Exception:
                    pass

            card_id, err = _resolve_card_for_client(name)
            if err:
                summary["errors"].append(err)
            if not card_id:
                summary["skipped_no_match"] += 1
                _to_needs_attention(row, name, "no Trello card matched")
                continue

            try:
                card = _tc.get_card(card_id, actions_limit=50)
                if not card:
                    summary["skipped_no_match"] += 1
                    _to_needs_attention(row, name, "Trello card not found")
                    continue
                # Pull EVERY comment via the paged API, not just the
                # last 50 the card payload bundles. Long-running cards
                # (audit-rejection chains, multi-month restoration jobs)
                # often have the "comp service" / "no final estimate" /
                # "cancelled" closeout comment older than the 50-item
                # cap, which caused them to misroute as Completed. The
                # paged helper has its own internal max_pages guard so
                # this still terminates on cards with thousands of
                # comments.
                try:
                    all_comments = _tc.get_all_comments(card_id) or []
                except Exception:
                    all_comments = []
                comments = "\n".join(
                    (a.get("data", {}).get("text") or "")
                    for a in all_comments
                )
                # Resolve the card's lane name (cached) so a job closed
                # by moving its card to a "Closed out" / "Logs" / "Paid"
                # lane — not just the LOGS board — is recognized too.
                lid = card.get("idList")
                if lid and lid not in list_name_cache:
                    try:
                        _l = _tc.get_list(lid)
                        list_name_cache[lid] = (_l or {}).get("name", "")
                    except Exception:
                        list_name_cache[lid] = ""
                list_name = list_name_cache.get(lid, "")
                status = _tc.card_route_status(
                    card, comments_text=comments, list_name=list_name)
                loss_type = _tc.card_loss_type(card)
            except Exception as ex:
                summary["errors"].append(f"{name}: trello fetch failed ({ex})")
                _to_needs_attention(row, name, f"card unclassifiable ({ex})")
                continue

            summary["by_target"][status] = summary["by_target"].get(status, 0) + 1
            if status == "new_loss" and len(summary["stayed_new_loss"]) < 50:
                summary["stayed_new_loss"].append(name)
            target_base = _ROUTE_HINT_TO_SHEET.get(status, _SHEET_NEW)
            target_title = _sheet_name(target_base, yr)
            if target_title not in wb.sheetnames:
                ws = wb.create_sheet(target_title)
                _write_header(ws)

            # Move row if its sheet changed. delete_rows shifts indices,
            # so refresh the affected sheets in the index after a move.
            if row.get("_sheet") != target_base:
                source_base = row.get("_sheet")
                _move_row_between_sheets(wb, yr, name, target_base)
                summary["moved"] += 1
                if source_base:
                    _refresh_index_for_sheet(wb, yr, source_base, name_index)
                _refresh_index_for_sheet(wb, yr, target_base, name_index)

            # Look up the row once via the index for every per-cell update
            # below. The previous code called `_find_row_by_name` twice
            # per iteration (O(N) each) which made the reconciler walk the
            # workbook ~2*M*N times for M rows.
            entry = name_index.get(_canon_name_key(name))
            rr = entry[1] if entry else None
            ws = wb[target_title]

            # Set Type of Loss from Trello labels. The reconciler is
            # explicitly a re-pull action, so we DO overwrite here when
            # the derived value differs — that's the only way a previous
            # single-label "Mold" gets corrected to "Water and Mold"
            # after a second cause is added. (`sync_jobs` keeps the
            # don't-clobber behavior for normal ongoing syncs.)
            if loss_type and rr is not None:
                col = _COL_INDEX["Type of Loss"]
                cur = ws.cell(rr, col).value
                cur_str = (cur or "").strip() if isinstance(cur, str) else (
                    "" if cur is None else str(cur).strip())
                if cur_str.lower() != loss_type.lower():
                    ws.cell(rr, col, loss_type)
                    summary["loss_type_set"] += 1

            # Pull Carrier / Claim# / dates from the Trello card and
            # write them into empty cells. We DON'T overwrite existing
            # values — Type of Loss is derived data we own, but Carrier
            # / Claim# / dates are often hand-edited by the user (claim
            # # corrections, special closing dates) and clobbering
            # them would erase that work.
            try:
                fields = _tc.parse_card_desc(card.get("desc"))
                ins = fields.get("INSURANCE INFORMATION") or {}
                trello_carrier = (ins.get("INSURANCE COMPANY") or "").strip()
                trello_claim   = (ins.get("CLAIM NUMBER") or "").strip()
            except Exception:
                trello_carrier = trello_claim = ""
            # Date Received: PROPERTY DETAILS > DATE RECEIVED in the
            # card desc is the source of truth (the user typed it when
            # the job came in) — always overwrite the cell with it,
            # because earlier sync runs may have stamped a wrong date
            # (e.g. card creation timestamp) before this rule existed.
            # Card creation date is a softer fallback: only fills the
            # cell when it's empty, never overwrites.
            try:
                received_from_desc = _tc.card_received_date(card)
            except Exception:
                received_from_desc = None
            try:
                received_from_creation = _tc.card_creation_date(card_id)
            except Exception:
                received_from_creation = None
            try:
                trello_closing = _tc.card_closing_date(card_id)
            except Exception:
                trello_closing = None

            if rr is not None:
                def _fill_if_empty(col_name, value):
                    if value is None or value == "":
                        return
                    col = _COL_INDEX[col_name]
                    cur = ws.cell(rr, col).value
                    cur_empty = (
                        cur is None
                        or (isinstance(cur, str) and not cur.strip())
                    )
                    if cur_empty:
                        ws.cell(rr, col, value)

                def _overwrite_if_different(col_name, value):
                    """Set the cell unconditionally when the Trello-
                    derived value differs from what's there. Used for
                    fields where Trello is the source of truth (Type of
                    Loss, Date Received from desc, Closing Date)."""
                    if value is None or value == "":
                        return
                    col = _COL_INDEX[col_name]
                    cur = ws.cell(rr, col).value
                    if cur != value:
                        ws.cell(rr, col, value)
                _fill_if_empty("Carrier", trello_carrier)
                _fill_if_empty("Claim#",  trello_claim)
                if received_from_desc is not None:
                    _overwrite_if_different("Date Received",
                                              received_from_desc)
                else:
                    _fill_if_empty("Date Received", received_from_creation)
                _overwrite_if_different("Closing Date", trello_closing)

        try:
            wb.save(path)
        except PermissionError:
            summary["errors"].append(
                "Workbook locked at save time — partial updates lost.")
        except Exception as ex:
            summary["errors"].append(f"Save failed: {ex}")

    return summary


def _build_trello_enrichment(client):
    """Look up the linked Trello card for `client` and pull every
    spreadsheet field Trello can answer that the local files can't:
        - Carrier (from card desc INSURANCE INFORMATION)
        - Claim# (same)
        - Type of Loss (from card labels)
        - Date Received (card creation date)
        - Closing Date (most recent move into LOGS - EMS, when present)

    Returns a dict shaped for `_row_to_cells` (`_carrier`, `_claim`,
    `_type_of_loss`, `_date_received`, `_closing_date`). Empty dict
    when no card matches or Trello is unreachable. Best-effort —
    callers should treat the result as additive.
    """
    enrich = {}
    try:
        import trello_client as _tc
    except Exception:
        return enrich

    # Shared lookup helper — same pinned-first / alias-fuzzy ordering the
    # reconciler uses, so both paths resolve to the same card.
    card_id, _err = _resolve_card_for_client(client)
    if not card_id:
        return enrich

    try:
        card = _tc.get_card(card_id, actions_limit=10)
    except Exception:
        card = None
    if not card:
        return enrich

    # Carrier + Claim# from the templated desc fields.
    try:
        fields = _tc.parse_card_desc(card.get("desc"))
        ins = fields.get("INSURANCE INFORMATION") or {}
        carrier = (ins.get("INSURANCE COMPANY") or "").strip()
        claim   = (ins.get("CLAIM NUMBER") or "").strip()
        if carrier:
            enrich["_carrier"] = carrier
        if claim:
            enrich["_claim"] = claim
    except Exception:
        pass

    # Type of Loss from the card's labels (multi-cause aware).
    try:
        loss = _tc.card_loss_type(card)
        if loss:
            enrich["_type_of_loss"] = loss
    except Exception:
        pass

    # Date Received: PROPERTY DETAILS > DATE RECEIVED on the card desc
    # is what the user typed when intake happened — that's the truth.
    # Card creation date is a fallback for cards copied from templates
    # where the field never got filled in.
    try:
        received = (_tc.card_received_date(card)
                     or _tc.card_creation_date(card_id))
        if received is not None:
            enrich["_date_received"] = received
    except Exception:
        pass

    # Closing Date = most recent move into LOGS - EMS board. Only set
    # when the card actually moved there; otherwise leave the cell to
    # the existing Date Received + 7 fallback in `_row_to_cells`.
    try:
        closed = _tc.card_closing_date(card_id)
        if closed is not None:
            enrich["_closing_date"] = closed
    except Exception:
        pass

    return enrich


def mark_completed(client, *, year=None, type_of_loss=None):
    """Force-route this client's row to the Completed sheet, creating
    the row if it doesn't exist. Used by the snapshot tool when a PDF
    is generated — that's the user's signal the job is being handed
    off, regardless of any audit flags.

    Pulls Carrier / Claim# / dates from the linked Trello card so the
    spreadsheet row gets populated from the authoritative source. The
    caller's `type_of_loss` overrides the Trello-derived one when both
    are present (snapshot's local cause field is more specific than
    raw card labels in some edge cases).
    """
    row = {
        "client": (client or "").strip(),
        "_route_override": "completed",
    }
    enrich = _build_trello_enrichment(client)
    row.update(enrich)
    if type_of_loss:
        row["_type_of_loss"] = type_of_loss
    sync_one(row, year=year)


def set_comment(client, comment, *, year=None):
    """Write the Comment cell for `client`. Searches all three sheets
    via `_find_row_by_name` (canonical name match) and updates the
    first hit. Returns True on write, False when the client wasn't
    found or the workbook is locked.

    Used by the spreadsheet's 📝 Generate notes flow — the user
    confirms a generated summary, this writes it to the row's free-
    text Comment column. Permits empty `comment` (clears the cell)."""
    if not client:
        return False
    yr = year or datetime.today().year
    path = workbook_path(yr)
    if _is_locked(path):
        return False
    with _LOCK:
        try:
            wb = openpyxl.load_workbook(path)
        except Exception:
            return False
        comment_col = _COL_INDEX["Comment"]
        wrote = False
        for base in _ALL_SHEETS:
            title = _sheet_name(base, yr)
            if title not in wb.sheetnames:
                continue
            ws = wb[title]
            r = _find_row_by_name(ws, client)
            if r is None:
                continue
            ws.cell(r, comment_col, comment or "")
            wrote = True
            break
        if not wrote:
            try:
                wb.close()
            except Exception:
                pass
            return False
        try:
            wb.save(path)
        except Exception:
            return False
    return True


def move_existing_row(client, target_sheet_base, *, year=None):
    """Public mover used by the state reconciler — finds `client`'s
    existing row (across all three sheets via _canon_name_key) and
    relocates it to `target_sheet_base` ("NEW LOSS" / "Completed" /
    "Incomplete"). Returns True on move, False when the client wasn't
    found anywhere.

    Honors the workbook lock — returns False without raising when
    Excel has the file open."""
    if not client or target_sheet_base not in _ALL_SHEETS:
        return False
    yr = year or datetime.today().year
    path = workbook_path(yr)
    if _is_locked(path):
        return False
    with _LOCK:
        try:
            wb = openpyxl.load_workbook(path)
        except Exception:
            return False
        moved = _move_row_between_sheets(wb, yr, client, target_sheet_base)
        if not moved:
            try:
                wb.close()
            except Exception:
                pass
            return False
        try:
            wb.save(path)
        except Exception:
            return False
    return True


# Columns a user may hand-edit from the Snapshot Tracked tab. Explicit
# allowlist so a stray/typo'd key can't write to an unintended column.
_EDITABLE_COLUMNS = (
    "Date Received", "Name", "Closing Date", "Comment", "Type of Loss",
    "Claim#", "Carrier", "Folder", "Scheduled Ins.", "Lead",
    "Inspection", "Docusketch ordered?", "Demo Start", "Add'l Docs",
    "Initial Photos", "Sketch", "Demo Photos", "Scope", "Final Photos",
    "ATP", "CIF", "CER", "COS",
)


def update_row_fields(client, fields, *, year=None):
    """Hand-edit one or more cells on `client`'s existing row.

    Finds the row across ALL sheets by canonical name and writes each
    `{column: value}` in `fields` — but only columns in
    `_EDITABLE_COLUMNS` (unknown keys ignored). `None` values clear the
    cell. Returns ``{"ok", "updated":[cols], "error"}``. Honors the
    workbook lock — returns ok=False (not raising) when Excel has it open.

    Used by the Tracked tab's per-row ✎ Edit dialog so the user can fix
    Carrier / Claim# / Type of Loss / dates / Comment etc. directly,
    without round-tripping through Trello.
    """
    if not client:
        return {"ok": False, "error": "no client"}
    if not isinstance(fields, dict) or not fields:
        return {"ok": False, "error": "no fields"}
    yr = year or datetime.today().year
    path = workbook_path(yr)
    if _is_locked(path):
        return {"ok": False,
                "error": "Workbook is open in Excel — close it and retry."}
    updated = []
    with _LOCK:
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as ex:
            return {"ok": False, "error": f"Open failed: {ex}"}
        target_ws = rr = None
        for base in _ALL_SHEETS:
            title = _sheet_name(base, yr)
            if title not in wb.sheetnames:
                continue
            r = _find_row_by_name(wb[title], client)
            if r is not None:
                target_ws, rr = wb[title], r
                break
        if target_ws is None:
            try: wb.close()
            except Exception: pass
            return {"ok": False,
                    "error": f"{client!r} not found in {yr} workbook"}
        for col_name, value in fields.items():
            if col_name not in _EDITABLE_COLUMNS:
                continue
            ci = _COL_INDEX.get(col_name)
            if not ci:
                continue
            target_ws.cell(rr, ci, "" if value is None else value)
            updated.append(col_name)
        if not updated:
            try: wb.close()
            except Exception: pass
            return {"ok": False, "error": "no editable fields given"}
        try:
            wb.save(path)
        except Exception as ex:
            return {"ok": False, "error": f"Save failed: {ex}"}
    return {"ok": True, "updated": updated}


def mark_new_loss(client, *, year=None, type_of_loss=None):
    """Force-route this client's row to the NEW LOSS sheet, creating
    the row if it doesn't exist. Used by the Initial Upload flow —
    when the user ticks INITIAL UPLOAD on a Trello card the job has
    officially been started, so it should appear on the NEW LOSS
    sheet of the snapshots workbook even before the first audit.

    Same Trello-enrichment as `mark_completed` — Carrier / Claim# /
    dates / loss type all come along for the ride if the card is
    pinned or fuzzy-matchable.
    """
    row = {
        "client": (client or "").strip(),
        "_route_override": "new_loss",
    }
    enrich = _build_trello_enrichment(client)
    row.update(enrich)
    if type_of_loss:
        row["_type_of_loss"] = type_of_loss
    sync_one(row, year=year)


def dedupe_workbook(year, *, dry_run=False):
    """Walk every row in the year's workbook, group by `_canon_name_key`,
    and merge duplicate rows into a single keeper.

    Catches three classes of dupes:
      • Internal-whitespace variants ("Rand  Melissa" + "Rand Melissa")
      • Comma-swap variants ("Sanchez, Jacqueline" + "Jacqueline Sanchez")
      • Case variants ("DOE, JOHN" + "Doe, John")

    Keeper selection per group:
      1. Prefer the comma form ("Last, First") — matches SERVPRO filing
         convention so the Name cell reads consistently after merge.
      2. Otherwise the first occurrence (top-to-bottom, sheet order
         NEW LOSS → Completed → Incomplete).

    Merge rule: keeper cell wins when it's non-empty; otherwise the
    duplicate's value fills the gap. So a row with a typed Comment +
    a row with a Lead value collapse into one row with BOTH.

    Returns a dict:
        {"groups": [
            {"canonical": "...", "keeper": (sheet, row, name),
             "merged": [(sheet, row, name), ...],
             "filled_fields": [col1, ...]},
            ...
         ],
         "rows_removed": int, "errors": [str, ...]}

    `dry_run=True` walks + reports without saving. Safe to call
    repeatedly — second call on a deduped workbook returns 0 groups.
    """
    summary = {"groups": [], "rows_removed": 0, "errors": []}
    yr = year or datetime.today().year
    path = workbook_path(yr)
    if _is_locked(path):
        summary["errors"].append(
            "Workbook is open in Excel — close it and retry.")
        return summary

    with _LOCK:
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as ex:
            summary["errors"].append(f"Open failed: {ex}")
            return summary

        # Snapshot every row across every sheet with its canonical key.
        name_col = _COL_INDEX["Name"]
        rows: list[dict] = []
        for base in _ALL_SHEETS:
            title = _sheet_name(base, yr)
            if title not in wb.sheetnames:
                continue
            ws = wb[title]
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, name_col).value
                if v is None or not str(v).strip():
                    continue
                canon = _canon_name_key(str(v))
                if not canon:
                    continue
                rows.append({
                    "sheet": base,
                    "title": title,
                    "row":   r,
                    "name":  str(v),
                    "canon": canon,
                })

        # Group by canonical key. Single-row groups need no work.
        by_canon: dict[str, list[dict]] = {}
        for entry in rows:
            by_canon.setdefault(entry["canon"], []).append(entry)
        dup_groups = {k: v for k, v in by_canon.items() if len(v) > 1}
        if not dup_groups:
            try:
                wb.close()
            except Exception:
                pass
            return summary

        # Plan deletions BEFORE applying — delete_rows() shifts indices,
        # so we apply per-sheet from highest row down to keep earlier
        # row numbers stable.
        deletions_by_title: dict[str, list[int]] = {}

        for canon, group in dup_groups.items():
            # Pick keeper. Priority: comma form (SERVPRO filing standard)
            # → no-double-whitespace form (clean cell) → first occurrence.
            comma_forms = [g for g in group if "," in g["name"]]
            if comma_forms:
                keeper = comma_forms[0]
            else:
                clean_forms = [g for g in group
                               if "  " not in g["name"].strip()]
                keeper = clean_forms[0] if clean_forms else group[0]
            others = [g for g in group if g is not keeper]

            ws_keeper = wb[keeper["title"]]
            filled: list[str] = []
            for col_name, col_idx in _COL_INDEX.items():
                cur = ws_keeper.cell(keeper["row"], col_idx).value
                cur_empty = (cur is None
                             or (isinstance(cur, str) and not cur.strip()))
                if not cur_empty:
                    continue
                # Pull the first non-empty value from any duplicate.
                for dup in others:
                    ws_dup = wb[dup["title"]]
                    v = ws_dup.cell(dup["row"], col_idx).value
                    v_empty = (v is None
                               or (isinstance(v, str) and not v.strip()))
                    if not v_empty:
                        ws_keeper.cell(keeper["row"], col_idx, v)
                        filled.append(col_name)
                        break

            for dup in others:
                deletions_by_title.setdefault(dup["title"], []).append(
                    dup["row"])

            summary["groups"].append({
                "canonical":    canon,
                "keeper":       (keeper["sheet"], keeper["row"],
                                 keeper["name"]),
                "merged":       [(o["sheet"], o["row"], o["name"])
                                 for o in others],
                "filled_fields": filled,
            })
            summary["rows_removed"] += len(others)

        if dry_run:
            try:
                wb.close()
            except Exception:
                pass
            return summary

        # Apply deletions per-sheet, highest-row-first.
        for title, row_list in deletions_by_title.items():
            ws = wb[title]
            for r in sorted(set(row_list), reverse=True):
                try:
                    ws.delete_rows(r, 1)
                except Exception as ex:
                    summary["errors"].append(
                        f"{title} row {r} delete failed: {ex}")

        try:
            wb.save(path)
        except PermissionError:
            summary["errors"].append(
                "Workbook locked at save time — dedupe rolled back.")
        except Exception as ex:
            summary["errors"].append(f"Save failed: {ex}")

    return summary


def open_in_excel(year):
    """Best-effort open the workbook in Excel via the OS handler.
    Returns True on success, False if the file doesn't exist yet
    (no rows synced yet for this year)."""
    path = workbook_path(year)
    if not os.path.isfile(path):
        return False
    try:
        os.startfile(path)
        return True
    except OSError:
        return False
