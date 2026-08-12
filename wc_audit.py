"""Monthly WC (WorkCenter) audit.

Slices a downloaded WC export to columns D-J (Date Received → Customer),
classifies every row by the customer's Trello status into one of five
buckets, and writes the result to a five-sheet workbook at
``X:\\IE_Public\\Front Operation\\WorkCenter audits\\``.

Routing rules (see :func:`classify_row` for the exact code path):

  Pending approval lane (any board) -> "pending_approval"  (lane wins)
  Recon board                        -> "recon"
  Estimating board                   -> "estimating"  (assignee = lane name)
  EMS / Contents / Monitor / Logs    -> "active_ems"
  Anything else                      -> "not_sold"

The not_sold rule is "any row not tied to one of the other four
categories" — per user direction. Includes both rows whose customer
has no Trello card at all and rows whose card is on a board that
doesn't match any of the four classifiers above.
"""
from __future__ import annotations

import datetime as _dt
import os
import re

import openpyxl

import trello_client as tc


# Shared share where Sam picks up the file. Matches the example pattern:
# X:\IE_Public\Front Operation\EMS Admin\WorkCenter audits\WC audit 4-10-26.xlsx
# The directory is config-backed (persisted under `wc_audit_dir`) so the
# user can relocate it from the WC Audit panel without a code edit.
_DEFAULT_OUTPUT_DIR = r"X:\IE_Public\Front Operation\EMS Admin\WorkCenter audits"


def output_dir():
    """Resolve the WC-audit output directory: persisted `wc_audit_dir`
    config value first, then the default."""
    try:
        import config
        p = (config.load().get("wc_audit_dir") or "").strip()
        if p:
            return p
    except Exception:
        pass
    return _DEFAULT_OUTPUT_DIR


def set_output_dir(path):
    """Persist a new WC-audit output directory to config. Pass None /
    empty to revert to the default."""
    try:
        import config
        cfg = config.load()
        cfg["wc_audit_dir"] = (path or "").strip() or _DEFAULT_OUTPUT_DIR
        config.save(cfg)
    except Exception:
        pass


# Back-compat alias — older references read the module attribute directly.
# Prefer output_dir() so a config override is honored.
OUTPUT_DIR = _DEFAULT_OUTPUT_DIR


# Output buckets. The first five become sheets in the workbook Sam
# receives. The trailing `excluded` bucket is GUI-only — rows on a
# "Logs" Trello board (closed/archived jobs) land there so the user
# can see what was filtered out but they're NOT written to the
# workbook. Tuple order = sheet order so reviewers see the same
# layout every month.
CATEGORIES = (
    "active_ems",
    "recon",
    "estimating",
    "not_sold",
    "pending_approval",
    "needs_attention",  # ← visible in workbook + GUI; needs a human eye
    "excluded",         # ← GUI-only, skipped by write_workbook
)

# Buckets that get written to the output workbook (Sam's view).
# Anything in CATEGORIES but not in WORKBOOK_CATEGORIES is GUI-only.
# needs_attention IS in the workbook so Sam sees what didn't fit
# cleanly — these need a human to look at the row and decide whether
# they belong in active EMS, Estimating, or should be killed off.
WORKBOOK_CATEGORIES = tuple(
    c for c in CATEGORIES if c != "excluded")

CATEGORY_LABELS = {
    "active_ems":       "Active EMS & Contents",
    "recon":            "RECON",
    "estimating":       "Estimating",
    "not_sold":         "Not sold",
    "pending_approval": "Pending approvals",
    "needs_attention":  "Needs Attention",
    "excluded":         "Excluded (Logs)",
}


# Source columns D-J (1-indexed: 4..10 inclusive = 7 cols) + AP
# (col 42 = "Not Sold/Cancelled Determination") which lives outside
# the D-J slice but is needed to detect WC-side "terminal" status
# (per the 2026-06-02 classifier rules: WC terminal + Trello archived
# → drop; either alone → not_sold).
_SOURCE_START_COL    = 4    # D
_SOURCE_END_COL      = 10   # J
_NOT_SOLD_DET_COL    = 42   # AP


# Header row written at the top of every output sheet. Matches the
# headers in the example file at WC audit 4-10-26.xlsx.
COL_HEADERS = (
    "Date Received",
    "Corporate Ref #",
    "Project #",
    "Property Type",
    "Type",
    "Progress",
    "Customer",
)


def _canon_name(s):
    """Lowercase + drop non-alnum, so 'KATHLEEN ACOSTA' matches
    'Kathleen Acosta' and 'Smith, John' matches 'Smith,John'."""
    if not s:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(s).lower())


_TRAILING_TAG_RE = re.compile(r"\s*-\s*[^-]+\s*$")


def _strip_trailing_tag(name):
    """Drop a SINGLE trailing ' - <Tag>' suffix from a Trello card
    name. Mirrors `_franchise_key` in APA Monitor.

    Use `_iter_stripped_forms` instead when matching against cards
    that may carry multiple suffixes (e.g., Disaster Response cards
    are typically named like 'Verma, Vandana - Amica - WILDFIRE',
    where stripping once still leaves a carrier tag).
    """
    if not name:
        return name
    return _TRAILING_TAG_RE.sub("", str(name).strip()).strip()


def _iter_stripped_forms(name):
    """Yield every progressively-stripped form of `name` from full
    to bare. For 'Verma, Vandana - Amica - WILDFIRE' yields:
        'Verma, Vandana - Amica - WILDFIRE'  (raw)
        'Verma, Vandana - Amica'              (one strip)
        'Verma, Vandana'                      (two strips)

    Bounded at 4 strips so a long-hyphenated commercial name doesn't
    loop forever. Dedupes — names without any trailing tag yield
    just the input.
    """
    seen = set()
    cur = (name or "").strip()
    for _ in range(5):
        if not cur or cur in seen:
            return
        seen.add(cur)
        yield cur
        nxt = _TRAILING_TAG_RE.sub("", cur).strip()
        if nxt == cur:
            return
        cur = nxt


def _name_variants(name):
    """Yield canonical-key variants for a name to cover reorders:
    'John Smith' <-> 'Smith, John'. Used both for indexing Trello cards
    and for looking up a WC row's customer against that index.

    Also yields variants for the suffix-stripped form of the name so
    a Trello card 'Smith, David - Mercury' matches the WC row 'David
    Smith'. Without this strip, the carrier-tagged Trello cards never
    matched the carrier-free WC export rows (David Smith was sitting
    on a real Trello card at /c/v4fWkho4 but the auto-classifier
    couldn't see it).
    """
    base = (name or "").strip()
    if not base:
        return
    # Generate variants for EVERY progressively-stripped form of the
    # name. Disaster Response cards carry two-tag suffixes
    # ('Verma, Vandana - Amica - WILDFIRE') so a single strip leaves
    # them un-matchable against the WC export's bare 'Verma, Vandana'.
    # _iter_stripped_forms yields raw → 1-strip → 2-strip → … bounded
    # at 5 iterations.
    candidates = list(_iter_stripped_forms(base))

    seen = set()

    def _emit(s):
        k = _canon_name(s)
        if k and k not in seen:
            seen.add(k)
            return k
        return None

    for b in candidates:
        k = _emit(b)
        if k: yield k
        # Comma reorder: "Smith, John" -> "John Smith"
        if "," in b:
            last, first = (p.strip() for p in b.split(",", 1))
            if last and first:
                k = _emit(f"{first} {last}")
                if k: yield k
                k = _emit(f"{last} {first}")
                if k: yield k
        # Word-swap fallback: "John Smith" -> "Smith John"
        parts = re.split(r"\s+", b)
        if len(parts) == 2:
            k = _emit(f"{parts[1]} {parts[0]}")
            if k: yield k


def load_source(path):
    """Parse the downloaded WC export. Finds each column by HEADER NAME,
    falling back to the historic D-J slice (+ AP) when a header is
    missing. Drops rows blank across all seven mapped columns.

    Returns a list of dicts:
        {date_received, corp_ref, project_num, property_type, type,
         progress, customer, not_sold_det}

    The `not_sold_det` field is the WC "Not Sold/Cancelled
    Determination" column (col AP) — used alongside `progress` to
    detect terminal-state jobs per the 2026-06-02 classifier rules.

    ⚠ Header lookup is not a nicety. The audits on the share start at
    column A, not D, so the fixed D-J slice landed three columns to the
    right: `customer` read the empty column J on EVERY row while the real
    name silently arrived as `property_type`. A blank customer is not
    visibly wrong in the panel — the table renders `cells`, not this
    field — but it made per-row Trello pinning impossible, because
    wc_audit_web.pin_trello_card refuses a blank name ("customer +
    card_id required"). Match on the header and both layouts work.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # First non-empty header wins, so a duplicate label later in the
    # sheet can't steal a column we already mapped.
    headers = {}
    for c in range(1, (ws.max_column or 0) + 1):
        h = str(ws.cell(1, c).value or "").strip().lower()
        if h and h not in headers:
            headers[h] = c

    def _col(name, legacy_offset):
        """1-based column for `name`; legacy_offset is the 0-based index
        into the old D-J slice used when the header isn't found."""
        return headers.get(name, _SOURCE_START_COL + legacy_offset)

    cols = (_col("date received",   0), _col("corporate ref #", 1),
            _col("project #",       2), _col("property type",   3),
            _col("type",            4), _col("progress",        5),
            _col("customer",        6))
    nsd_col = headers.get("not sold/cancelled determination",
                          headers.get("not sold detail", _NOT_SOLD_DET_COL))

    rows = []
    for r in range(2, ws.max_row + 1):  # skip header row
        vals = [ws.cell(r, c).value for c in cols]
        if not any(v not in (None, "") for v in vals):
            continue
        # Pulled separately — outside the main block but needed for
        # terminal-state detection. Sheets that stop short of this
        # column simply have no determination to read.
        try:
            nsd = ws.cell(r, nsd_col).value
        except Exception:
            nsd = None
        rows.append({
            "date_received": vals[0],
            "corp_ref":      vals[1] or "",
            "project_num":   vals[2] or "",
            "property_type": vals[3] or "",
            "type":          vals[4] or "",
            "progress":      vals[5] or "",
            "customer":      (str(vals[6] or "")).strip(),
            "not_sold_det":  str(nsd or "").strip(),
        })
    return rows


def build_trello_index(*, progress_cb=None):
    """Walk every in-scope Trello board once, return a name index:
        {canon_name: {board_name, list_name, card_id, card_name, card_url}}

    Multiple cards can share a canonical name; the first one seen wins
    (board iteration order from tc.list_boards). Rare in practice — WC
    customer names are usually unique per active job.

    ``progress_cb(idx, total, board_name)`` fires once per board so the
    GUI can paint a determinate progress label during the scan.
    """
    idx = {}
    boards = tc.list_boards(exclude_quality=True) or []
    n = len(boards)
    for i, b in enumerate(boards, 1):
        if progress_cb is not None:
            try:
                progress_cb(i, n, b.get("name", ""))
            except Exception:
                pass
        try:
            lists = tc._call(
                f"/boards/{b['id']}/lists",
                params={"fields": "id,name", "filter": "open"},
            ) or []
        except Exception:
            continue
        # On Logs boards we need card description + labels too so
        # the comp-inspection / $0.00 detection can fire. Slightly
        # larger payload per card (~500-1500 bytes of desc text) but
        # the alternative was a second per-card API call. Other
        # boards keep the lean default field set.
        is_logs = _is_logs_board(b.get("name", ""))
        card_fields = ("id,name,shortUrl,desc,labels"
                       if is_logs else None)
        for lst in lists:
            try:
                if card_fields:
                    cards = tc.cards_in_list(
                        lst.get("id"), fields=card_fields) or []
                else:
                    cards = tc.cards_in_list(lst.get("id")) or []
            except Exception:
                continue
            for c in cards:
                nm = (c.get("name") or "").strip()
                if not nm:
                    continue
                for key in _name_variants(nm):
                    if not key or key in idx:
                        continue
                    entry = {
                        "board_name": b.get("name", ""),
                        "list_name":  lst.get("name", ""),
                        "card_id":    c.get("id", ""),
                        "card_name":  nm,
                        "card_url":   c.get("shortUrl", ""),
                    }
                    if is_logs:
                        entry["desc"]   = c.get("desc") or ""
                        entry["labels"] = c.get("labels") or []
                    idx[key] = entry
    return idx


def _board_classifies_as(board_name):
    """Map a Trello board name to a bucket key, or None.

    Used in the post-2026-06-02 classifier as a *secondary* signal —
    the WC `Type` column is now the authority for active_ems / recon
    / contents routing. This function only resolves the Estimating
    board (assignee = lane name) and helps `is_archived` callers
    spot Logs / Archived / AR boards.

    Substring match (case-insensitive):
      • 'estimating'/'estimate' → estimating
      • everything else → None (let WC Type drive)
    """
    if not board_name:
        return None
    bl = board_name.lower()
    if "estimating" in bl or "estimate" in bl:
        return "estimating"
    return None


# ── 2026-06-02 classifier rules ─────────────────────────────────────────
# User direction: WC source columns drive the bucket, with Trello
# board names as a routing override only for archived/estimating/
# pending-approval cases. Specifically:
#   1. WC terminal (Closed / Cancelled / Not Sold / AP populated)
#      AND Trello archived (AR / Logs / Archived board, or a
#      "Completed" lane on any board) → drop from workbook entirely.
#   2. Either signal alone → not_sold.
#   3. Pending Approval lane (any board) → pending_approval.
#   4. Estimating board → estimating (assignee = lane name).
#   5. WC `Type` column drives the remaining buckets:
#      • "Reconstruction" → recon
#      • "Contents"       → active_ems (the "Active EMS & Contents" sheet)
#      • Water / Fire / Mold / Smoke / Biohazard / Storm /
#        Trauma / Asbestos / Lead / Vandalism / General Cleaning
#        → active_ems
#   6. Empty or unknown WC Type (Other / Roofing / blank) → not_sold.
#
# Recon and Contents specifically REQUIRE the WC Type tag to confirm —
# matching a Trello card on a Contents-named board alone is not
# enough (the auto-matcher was misrouting cards that happened to live
# on Contents boards without actually being Contents work).
_AR_BOARD_RE      = re.compile(r"\bar\b", re.IGNORECASE)
_LOGS_BOARD_RE    = re.compile(r"\blogs?\b", re.IGNORECASE)
_ARCHIVED_RE      = re.compile(r"archiv", re.IGNORECASE)
_DISASTER_RESP_RE = re.compile(r"disaster", re.IGNORECASE)
# WIP / WORK IN PROGRESS board — explicitly active EMS work, even
# when other signals (Type column, name match) are ambiguous. Per
# user direction 2026-06-02: anything on the WIP board IS active,
# always. Accepts 'wip' / 'wis' / 'work in progress' variants.
_WIP_BOARD_RE     = re.compile(
    r"work\s*in\s*progress|\bwi[ps]\b", re.IGNORECASE)


def _is_wip_board(board_name):
    """True when the Trello board indicates active in-progress EMS
    work — overrides Type-mismatch and stale-lead routing."""
    return bool(_WIP_BOARD_RE.search(board_name or ""))


def _is_logs_board(board_name):
    """True for "THE LOGS - EMS" / any board with a `\bLogs?\b`
    standalone-word match."""
    return bool(_LOGS_BOARD_RE.search(board_name or ""))


_LOGS_BILLED_LANE_RE = re.compile(r"\bbilled\b", re.IGNORECASE)
_LOGS_COMP_DESC_RE   = re.compile(
    r"comp\s+inspection|\$\s*0\.00|\$\s*0\b|no\s*charge|complimentary",
    re.IGNORECASE)
_LOGS_COMP_LABEL_RE  = re.compile(
    r"comp(\b|liment)|no\s*charge", re.IGNORECASE)


def _logs_lane_is_billed(lane_name):
    """True when a Logs-board lane name encodes BILLED status (e.g.,
    'MAY 2026 - BILLED', '2025 - BILLED'). Lists like 'TO BE
    PRESERVED' and 'INVOICE QUESTIONS' do NOT match — those are
    the not-yet-billed / unresolved lanes that should route to
    not_sold per user direction."""
    return bool(_LOGS_BILLED_LANE_RE.search(lane_name or ""))


def _logs_card_is_comp(desc, labels):
    """True when a Logs-board card's description or labels indicate
    comp inspection / $0.00 / no charge. Catches the 'work wasn't
    actually performed for money' case the user calls out for the
    not_sold bucket.

    Description match: 'comp inspection', '$0.00', '$0', 'no charge',
    'complimentary'. Label match: any label whose name starts with
    'comp' (Comp, Complimentary) or is 'No Charge'.
    """
    if desc and _LOGS_COMP_DESC_RE.search(str(desc)):
        return True
    for lab in (labels or []):
        nm = lab.get("name") if isinstance(lab, dict) else lab
        if nm and _LOGS_COMP_LABEL_RE.search(str(nm)):
            return True
    return False


# Lazy cache for Logs-card biller lookups. Trello comments are a
# separate API call per card, so we only fetch them when actually
# classifying a Logs-billed row (bounded by the WC export size,
# not the 1336 open Logs cards).
_biller_cache: dict[str, str] = {}
_member_name_cache: dict[str, str] = {}


# Canonical billing message: "Billed and Uploaded Job @toniabaca"
# — the @-mention IS the biller. Capture the username after @.
# Some posters omit the message and just paste an @-mention.
_BILLED_MSG_RE = re.compile(r"billed", re.IGNORECASE)
_BILL_MENTION_RE = re.compile(r"@([A-Za-z0-9_]+)")


def _resolve_member_name(username):
    """Convert a Trello @username → full display name. Cached so a
    franchise with one biller (typical) only fetches once."""
    if not username:
        return ""
    if username in _member_name_cache:
        return _member_name_cache[username]
    try:
        mem = tc._call(
            f"/members/{username}",
            params={"fields": "fullName,username"}) or {}
        name = mem.get("fullName") or mem.get("username") or username
    except Exception:
        name = username
    _member_name_cache[username] = name
    return name


_logs_card_inspect_cache: dict[str, tuple] = {}


def _inspect_logs_card(card_id):
    """Walk a Logs card's comments once and report two facts:
        (is_comp, biller_name)

    is_comp: True when ANY comment contains comp/$0/no-charge wording
        (or the card desc/labels we already checked also flag it —
        caller adds those in). Catches the Aguilar case where the
        lane was JUNE 2026 - BILLED but a comment said
        'Comp inspection\\n$0.00'.

    biller_name: resolved per the canonical "Billed and Uploaded
        Job @username" pattern. Walks newest-first, picks the first
        billed-tagged comment's @-mention. Falls back to the most-
        recent comment author when no billed-tagged comment exists.

    Cached per card_id — one Trello call regardless of how many WC
    rows share this card.
    """
    if not card_id:
        return (False, "")
    if card_id in _logs_card_inspect_cache:
        return _logs_card_inspect_cache[card_id]
    try:
        actions = tc._call(
            f"/cards/{card_id}/actions",
            params={"filter": "commentCard", "limit": 20,
                    "fields": "data,date,memberCreator,idMemberCreator"}) or []
    except Exception:
        actions = []
    is_comp = False
    biller = ""
    fallback_author = ""
    for a in actions:
        data = a.get("data") if isinstance(a.get("data"), dict) else {}
        text = str(data.get("text") or "") if isinstance(data, dict) else ""
        # Comment-based comp/$0 signal
        if text and _LOGS_COMP_DESC_RE.search(text):
            is_comp = True
        # Capture fallback author from the most-recent comment.
        if not fallback_author:
            mc = a.get("memberCreator") if isinstance(
                a.get("memberCreator"), dict) else None
            if mc and (mc.get("fullName") or mc.get("username")):
                fallback_author = mc.get("fullName") or mc.get("username")
            elif a.get("idMemberCreator"):
                try:
                    mem = tc._call(
                        f"/members/{a['idMemberCreator']}",
                        params={"fields": "fullName,username"}) or {}
                    fallback_author = (mem.get("fullName")
                                       or mem.get("username") or "")
                except Exception:
                    pass
        # Primary billed-pattern match — newest one wins.
        if (not biller and text
                and _BILLED_MSG_RE.search(text)):
            m = _BILL_MENTION_RE.search(text)
            if m:
                biller = _resolve_member_name(m.group(1))
    result = (is_comp, biller or fallback_author)
    _logs_card_inspect_cache[card_id] = result
    return result


def _logs_card_biller(card_id):
    """Backward-compat shim — return only the biller part of
    _inspect_logs_card."""
    return _inspect_logs_card(card_id)[1]
# Lane-level terminal states. Covers Disaster Response's COMPLETED /
# CANCELED / PAID / DISPOSAL lanes and similar phrasing on other
# boards. The user direction (2026-06-02) was "all Disaster Response
# cards belong in not_sold" — handled at the board level — but the
# lane-level regex catches the same patterns on any board where they
# appear (e.g., a "COMPLETED" lane on the AR board).
_TERMINAL_LANE_RE = re.compile(
    r"complet|cancel|\bpaid\b|disposal|collection|closed", re.IGNORECASE)


def _is_archived_trello(board_name, lane_name=""):
    """True when the Trello board+lane indicate the job has left the
    EMS workflow:

      • AR Board (collections / closed)
      • Logs board (closed-out EMS work)
      • Archived board (literal Archive boards)
      • Disaster Response board (separate department — wildfire /
        catastrophic loss — not EMS work)
      • Any lane named COMPLETED / CANCELED / PAID / DISPOSAL /
        COLLECTIONS / CLOSED on any board

    Word-boundary regexes (`\bar\b`, `\blogs?\b`, `\bpaid\b`) avoid
    substring false-positives.
    """
    b = board_name or ""
    l = lane_name or ""
    if _AR_BOARD_RE.search(b):       return True
    if _LOGS_BOARD_RE.search(b):     return True
    if _ARCHIVED_RE.search(b):       return True
    if _DISASTER_RESP_RE.search(b):  return True
    if _TERMINAL_LANE_RE.search(l):  return True
    return False


_WC_TERMINAL_KEYWORDS = ("closed", "cancel", "not sold", "lost",
                          "rejected", "dead")


def _is_wc_terminal(progress, not_sold_det=""):
    """True when WorkCenter source data indicates the job is
    terminated (closed / cancelled / not sold). Either a terminal
    keyword in the Progress column OR any non-empty value in the
    'Not Sold/Cancelled Determination' column counts.
    """
    p = (progress or "").strip().lower()
    if any(kw in p for kw in _WC_TERMINAL_KEYWORDS):
        return True
    if (not_sold_det or "").strip():
        return True
    return False


_WC_TYPE_TO_BUCKET = {
    "water":              "active_ems",
    "fire":               "active_ems",
    "mold":               "active_ems",
    "smoke":              "active_ems",
    "biohazard":          "active_ems",
    "storm":              "active_ems",
    "trauma":             "active_ems",
    "asbestos":           "active_ems",
    "lead":               "active_ems",
    "vandalism":          "active_ems",
    "general cleaning":   "active_ems",
    "contents":           "active_ems",  # Active EMS & Contents sheet
    "reconstruction":     "recon",
}


def _wc_type_bucket(wc_type):
    """Map WC Type column → bucket key. Returns None for blank /
    unknown types (Other / Roofing / etc.) so the caller routes
    those to not_sold per rule #6."""
    return _WC_TYPE_TO_BUCKET.get((wc_type or "").strip().lower())


def _current_yymm():
    """Return today's YYMM as an int (e.g. 2026-06-02 → 2606).
    Used to compare a Project #'s opening month against 'now' so
    stale leads can be flagged for review."""
    d = _dt.date.today()
    return (d.year % 100) * 100 + d.month


def _project_yymm(project_num):
    """Parse a Project # prefix → YYMM int. Returns None when the
    leading 4 chars aren't digits (defensive — some legacy IDs may
    use a different shape). Project IDs in this franchise are
    formatted 'YYMM-NNNNNNLOSS' where YYMM is the year+month the job
    was opened ('2606-262058WTR' → opened June 2026)."""
    s = str(project_num or "").strip()
    if len(s) < 4 or not s[:4].isdigit():
        return None
    return int(s[:4])


def _is_stale_project(project_num):
    """True when the Project # was opened before this calendar month.
    Tied to the calendar so the rule auto-rolls each month — no
    hardcoded prefix to update.
    """
    p = _project_yymm(project_num)
    if p is None:
        return False
    return p < _current_yymm()


_PROJECT_SUFFIX_RE = re.compile(r"([A-Z]{2,4})\s*$")


def _suffix_from_project(project_num):
    """Pull the trailing loss-code suffix from a Workcenter Project #
    like '2606-267339WTR' → 'WTR'. Returns '' when no trailing letters
    are present (e.g. a numeric-only ref). Uppercased for matching."""
    if not project_num:
        return ""
    m = _PROJECT_SUFFIX_RE.search(str(project_num).strip().upper())
    return m.group(1) if m else ""


def _category_from_loss_code(code):
    """Map a Project # loss-code suffix → bucket key, or None when the
    code doesn't tell us anything useful.

    Bucket layout matches the workbook writer's sheet names:
      • EMS codes (WTR/FIR/MLD/...) → active_ems
      • CON/CTS (Contents)          → active_ems too
            The output sheet is literally named "Active EMS & Contents"
            and Sam expects active Contents jobs (e.g., Rory Nowell)
            to show up there — not in not_sold. Earlier this returned
            not_sold for CON/CTS based on the EMS-audit folder-walk
            rule (which excludes Contents/), but that's a file-ops
            constraint, not a reporting constraint.
      • REC/RCN/RST                 → recon
    """
    if not code:
        return None
    try:
        from workcenter_client import (EMS_LOSS_CODES,
                                          NON_EMS_LOSS_CODES)
    except Exception:
        return None
    if code in EMS_LOSS_CODES:
        return "active_ems"
    if code in NON_EMS_LOSS_CODES:
        if code in {"REC", "RCN", "RST"}:
            return "recon"
        if code in {"CON", "CTS"}:
            return "active_ems"  # "Active EMS & Contents" sheet
    return None


def classify_row(customer, idx, project_num="", wc_type="",
                  wc_progress="", wc_not_sold_det=""):
    """Return ``(category, assignee)``.

    2026-06-02 priority chain (per user direction):
      1. Empty customer                                    → not_sold
      2. WC terminal AND Trello archived/AR/Logs           → excluded
      3. Trello archived (AR/Logs/Archived/COMPLETED lane) → not_sold
      4. WC terminal alone (Progress=Closed/Cancelled/...) → not_sold
      5. Pending Approval lane (any board)                 → pending_approval
      6. Estimating board                                  → estimating (assignee=lane)
      7. WC `Type` column drives:
         • Reconstruction                  → recon
         • Contents                        → active_ems
         • Water/Fire/Mold/Smoke/Biohazard/
           Storm/Trauma/Asbestos/Lead/
           Vandalism/General Cleaning      → active_ems
      8. Anything else (blank Type, Other, Roofing, etc.)  → not_sold

    Note: the Project # suffix is NO LONGER consulted. The user
    confirmed the WC Type column is the authoritative tag — matching
    only on Project # suffix produced false positives (long-stale
    leads with WTR/FIR suffixes landed in active_ems even when
    no real EMS work was happening). Same goes for Contents: the
    Trello board name 'CONTENTS' alone is NOT enough — the WC Type
    column must say "Contents" to actually classify as active_ems
    via the Contents path (sister rule to the Recon-tag requirement).
    """
    if not customer:
        return ("not_sold", "")
    hit = None
    for key in _name_variants(customer):
        if key and key in idx:
            hit = idx[key]
            break

    board = (hit.get("board_name") if hit else "") or ""
    lane  = (hit.get("list_name")  if hit else "") or ""

    is_terminal = _is_wc_terminal(wc_progress, wc_not_sold_det)

    # Rule 2 (combined drop): only the "excluded" gate that requires
    # BOTH signals lives this high. Plain archived gets a refined
    # treatment below since Logs needs special billing-status
    # handling — sending it straight to not_sold here would skip the
    # billed → estimating routing.
    if is_terminal and hit and _is_archived_trello(board, lane):
        return ("excluded", "")

    # Rule 3: Logs board — branch on billed vs comp/unbilled.
    if hit and _is_logs_board(board):
        # Fetch comments ONCE per card and pull both the comp signal
        # and biller out of the same walk. Caches by card_id so
        # repeat WC shells on the same card never re-hit Trello.
        comment_is_comp, biller = _inspect_logs_card(hit.get("card_id"))
        # Comp signal can come from THREE places: card description
        # (e.g., 'comp inspection' typed into desc), labels (a 'Comp'
        # / 'No Charge' tag), or comments (the most common — Aguilar
        # case: lane was JUNE 2026 - BILLED but a comment said 'Comp
        # inspection\n$0.00'). Any of the three → not_sold.
        if (_logs_card_is_comp(hit.get("desc"), hit.get("labels"))
                or comment_is_comp):
            return ("not_sold", "")
        # Lane name encodes BILLED status. Billed → estimating with
        # the @-mentioned biller from the canonical
        # 'Billed and Uploaded Job @username' comment (falls back to
        # the most-recent comment author when no canonical message).
        if _logs_lane_is_billed(lane):
            return ("estimating", biller or "Billed (Logs)")
        # Logs but not billed and no comp signal — preserve, invoice
        # questions, etc. Goes to not_sold (no work was booked).
        return ("not_sold", "")

    # Rule 4: All other archived signals (AR Board, Disaster
    # Response, Archived boards, COMPLETED/PAID/CANCELED lanes
    # elsewhere) → not_sold.
    if hit and _is_archived_trello(board, lane):
        return ("not_sold", "")
    if is_terminal:
        return ("not_sold", "")

    # Rules 5-7: Trello lane/board overrides for live cards only.
    if hit:
        lane_low = lane.lower()
        if "pending" in lane_low and "appr" in lane_low:
            return ("pending_approval", "")
        b_bucket = _board_classifies_as(board)
        if b_bucket == "estimating":
            # Per user 2026-06-02: Estimating-board rows get a literal
            # 'Pending Review' assignee — same group label for every
            # row in the bucket — instead of the lane-name (estimator)
            # value. Sam sorts/groups on this column.
            return ("estimating", "Pending Review")
        # Rule 6.5 (per user direction 2026-06-02): WIP / WORK IN
        # PROGRESS board → active_ems, always. The WIP board IS the
        # active-work signal; it overrides downstream Type-mismatch
        # and stale-lead heuristics. Without this rule, a row whose
        # Type column happened to say "Other" or whose Project # was
        # old would land in Needs Attention even though Trello says
        # work is in progress right now.
        if _is_wip_board(board):
            return ("active_ems", "")
        # Trello-vs-WC tag mismatch flag: card sits on a Contents-
        # named board but the WC Type column doesn't agree. This was
        # the ANNA ESFAHANI case — her card is on CONTENTS · PACK OUT
        # / PODS (Contents dept owns her now) but Type=Smoke says she
        # opened as an EMS job. Send to Needs Attention so a human
        # decides whether the Trello pin is wrong or the WC tag is.
        if ("contents" in board.lower()
                and (wc_type or "").strip().lower() != "contents"):
            return ("needs_attention", "")

    # Rules 7-8: WC Type drives the remaining classification.
    type_bucket = _wc_type_bucket(wc_type)
    if type_bucket:
        # Stale-lead flag: a row that LOOKS like active EMS by Type
        # alone (Water/Fire/etc.) but has NO Trello card AND its
        # Project # was opened before this calendar month is almost
        # certainly a dead lead the franchise never closed in WC.
        # Send to Needs Attention so Sam can review instead of
        # silently bloating Active EMS with months-old non-work.
        if (type_bucket == "active_ems"
                and not hit
                and _is_stale_project(project_num)):
            return ("needs_attention", "")
        return (type_bucket, "")
    return ("not_sold", "")


def bucket_rows(rows, idx):
    """Annotate each row with ``_cat``, ``assignee``, and the matched
    Trello card metadata (``card_id``, ``card_url``, ``board_name``,
    ``list_name``, ``card_name``). Returns the same list so callers
    can re-group on demand (lets the GUI move rows between buckets
    just by updating ``_cat``).

    Project # is threaded through so classify_row can fall back to
    the loss-code suffix (WTR/FIR/etc.) when the customer isn't yet
    on a Trello card.

    Pinning the matched hit onto the row is what enables the per-row
    🔗 Trello button in the GUI — without it, the link would need a
    second lookup at render time and would silently fail for rows
    where _name_variants produced a stale key.
    """
    for r in rows:
        customer = r.get("customer", "")
        cat, assignee = classify_row(
            customer, idx,
            project_num=r.get("project_num", ""),
            wc_type=r.get("type", ""),
            wc_progress=r.get("progress", ""),
            wc_not_sold_det=r.get("not_sold_det", ""))
        r["_cat"] = cat
        r["assignee"] = assignee
        hit = None
        for key in _name_variants(customer):
            if key and key in idx:
                hit = idx[key]
                break
        if hit:
            r["card_id"]    = hit.get("card_id")    or ""
            r["card_url"]   = hit.get("card_url")   or ""
            r["card_name"]  = hit.get("card_name")  or ""
            r["board_name"] = hit.get("board_name") or ""
            r["list_name"]  = hit.get("list_name")  or ""
        else:
            # Suffix-fallback rows (e.g., David Smith WTR with no
            # Trello card yet) get empty card fields so the renderer
            # knows to skip the 🔗 button.
            r.setdefault("card_id", "")
            r.setdefault("card_url", "")
            r.setdefault("card_name", "")
            r.setdefault("board_name", "")
            r.setdefault("list_name", "")
    return rows


def rows_for_cat(rows, cat):
    return [r for r in rows if r.get("_cat") == cat]


def write_workbook(rows, output_path):
    """Emit the five-sheet output workbook. Estimating gets a trailing
    Assignee column AND is sorted by assignee so every row for the
    same estimator lands contiguously (Juan's batch together, then
    Aaron's, then Samantha's, etc.). Other sheets retain insertion
    order. Rows classified `excluded` (terminal job in an archived
    Trello board) are deliberately left OUT so Sam isn't reviewing
    completed work."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default blank sheet
    for cat in WORKBOOK_CATEGORIES:
        ws = wb.create_sheet(title=CATEGORY_LABELS[cat])
        headers = list(COL_HEADERS)
        if cat == "estimating":
            headers.append("Assignee")
        ws.append(headers)
        cat_rows = list(rows_for_cat(rows, cat))
        if cat == "estimating":
            # Stable sort: same-assignee rows preserve their original
            # order, but assignees themselves are alphabetized so
            # Sam can scan one estimator's queue at a time. Empty
            # assignee (unrouted estimating rows) sinks to the bottom.
            def _asg_sort_key(r):
                a = (r.get("assignee") or "").strip().lower()
                # Empty → sentinel '~' so it sorts after any real name
                return (a or "~", a)
            cat_rows.sort(key=_asg_sort_key)
        for r in cat_rows:
            row_vals = [
                r.get("date_received"),
                r.get("corp_ref", ""),
                r.get("project_num", ""),
                r.get("property_type", ""),
                r.get("type", ""),
                r.get("progress", ""),
                r.get("customer", ""),
            ]
            if cat == "estimating":
                row_vals.append(r.get("assignee", ""))
            ws.append(row_vals)
    wb.save(output_path)


def default_output_path(run_date=None):
    """Path under OUTPUT_DIR for a given run date:
       ``WC audit <M>-<D>-<YY>.xlsx``
    Matches the example filename pattern ``WC audit 4-10-26.xlsx``."""
    d = run_date or _dt.date.today()
    return os.path.join(
        output_dir(),
        f"WC audit {d.month}-{d.day}-{d.year % 100}.xlsx")


# ── Monthly reminder ────────────────────────────────────────────────────────


def _first_monday_of_month(year, month):
    """Return the date of the first Monday of a given month."""
    d = _dt.date(year, month, 1)
    # weekday(): Mon=0..Sun=6. Days to add to reach the first Monday.
    return d + _dt.timedelta(days=(0 - d.weekday()) % 7)


def _audit_completed_this_month(today=None):
    """True if OUTPUT_DIR already has a WC audit file dated within the
    current calendar month. Cheap glob — no shared-share contention."""
    today = today or _dt.date.today()
    _dir = output_dir()
    if not os.path.isdir(_dir):
        return False
    try:
        entries = os.listdir(_dir)
    except OSError:
        return False
    # Files look like "WC audit M-D-YY.xlsx". We compare against the
    # current month/year only.
    pat = re.compile(r"WC audit\s+(\d{1,2})-(\d{1,2})-(\d{2,4})",
                      re.IGNORECASE)
    yy_short = today.year % 100
    for name in entries:
        m = pat.search(name)
        if not m:
            continue
        mo, _da, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        # Normalize 2-digit / 4-digit year.
        if yy < 100:
            yy_full = 2000 + yy
        else:
            yy_full = yy
        if yy_full == today.year and mo == today.month:
            return True
    return False


def is_audit_due(today=None):
    """Return True when the monthly WC audit should be on the user's
    radar — fires from the first Monday of the month through the next
    Sunday, unless an audit file already exists for this month.

    Window choice: a Monday→Sunday week gives the user the work week to
    knock it out without nagging on weekends. The strict 1st-of-month
    trigger would over-fire on weekend 1sts (Sat/Sun starts) and miss
    the natural workflow."""
    today = today or _dt.date.today()
    first_mon = _first_monday_of_month(today.year, today.month)
    window_end = first_mon + _dt.timedelta(days=6)  # following Sunday
    if not (first_mon <= today <= window_end):
        return False
    return not _audit_completed_this_month(today=today)


# ── Debug helpers (run via `python wc_audit.py debug <xlsx>`) ───────────


def _nearest_index_keys(variants, idx, limit=4):
    """Return the `limit` index keys most lexically similar to the
    given variants — used to suggest "did you mean Smith vs Smiht?"
    in the debug output. Uses difflib for ordering."""
    import difflib as _dl
    keys = list(idx.keys())
    best = []
    seen = set()
    for v in variants:
        if not v:
            continue
        for k in _dl.get_close_matches(v, keys, n=limit, cutoff=0.6):
            if k not in seen:
                seen.add(k)
                best.append(k)
                if len(best) >= limit:
                    return best
    return best


def explain_not_sold(xlsx_path, idx=None, *, only_not_sold=True,
                       max_rows=200, show_nearest=True):
    """For each row in the WC export, report the classify_row outcome
    plus (when not_sold) every name variant that was tried and the
    nearest Trello index keys.

    Set ``only_not_sold=False`` to dump every row's classification.
    """
    rows = load_source(xlsx_path)
    if idx is None:
        print("Building Trello index (walks every in-scope board)…")
        idx = build_trello_index()
    print(f"Trello index size: {len(idx)} canonicalized names")
    print(f"Loaded {len(rows)} rows from {os.path.basename(xlsx_path)}")
    print("=" * 72)
    shown = 0
    for r in rows:
        customer = (r.get("customer") or "").strip()
        project_num = r.get("project_num") or ""
        cat, assignee = classify_row(
            customer, idx,
            project_num=project_num,
            wc_type=r.get("type", ""),
            wc_progress=r.get("progress", ""),
            wc_not_sold_det=r.get("not_sold_det", ""))
        if only_not_sold and cat != "not_sold":
            continue
        if shown >= max_rows:
            print(f"…stopping at {max_rows} rows. Re-run with --all to "
                  "see everything.")
            break
        shown += 1
        print(f"\n[{cat:>17}]  {customer!r}")
        if not customer:
            print("    (empty Customer cell)")
            continue
        variants = list(_name_variants(customer))
        print(f"    variants tried: {variants}")
        hits = [(v, idx[v]) for v in variants if v and v in idx]
        if hits:
            for v, hit in hits:
                print(f"    HIT on {v!r} → board={hit.get('board_name')!r} "
                      f"lane={hit.get('list_name')!r}")
            if cat == "not_sold":
                print("    (matched a card but board name didn't classify)")
        else:
            print("    no index hit on any variant")
            if show_nearest:
                near = _nearest_index_keys(variants, idx)
                if near:
                    print(f"    nearest index keys: {near}")
    print("=" * 72)
    print(f"Showed {shown} row(s).")


def verify_estimating(xlsx_path, idx=None, *,
                        fuzzy_cutoff=0.78, max_suggest=6):
    """Sanity-check: for every row NOT currently classified as
    estimating, fuzzy-match the customer name against EVERY card on
    every estimating-named board. Surfaces likely-missed matches the
    user can manually move via Move →.

    This catches the case where `_name_variants` didn't produce the
    exact canon key that `build_trello_index` indexed under — usually
    because the WC export spelled the name slightly differently than
    the Trello card (extra middle initial, missing apostrophe,
    'Mr.' prefix, comma-vs-no-comma, etc.). The fuzzy pass uses
    difflib at a cutoff ≥ `fuzzy_cutoff` (default 0.78 — empirically
    tight enough to avoid false positives but loose enough to catch
    typo-level mismatches).

    `idx` should be the full Trello workspace index from
    `build_trello_index`. Pass None to build it (slow).
    """
    import difflib as _dl
    rows = load_source(xlsx_path)
    if idx is None:
        print("Building Trello index (walks every in-scope board)…")
        idx = build_trello_index()
    # Walk the index to extract only estimating-board cards and build
    # {canon_name: full hit dict}. Same hits format build_trello_index
    # produces, so the "this is in someone's estimating lane" report
    # can show the assignee (lane name on estimating boards) directly.
    est_keys = {}
    for k, hit in idx.items():
        if _board_classifies_as(hit.get("board_name") or "") == "estimating":
            est_keys[k] = hit
    if not est_keys:
        print("No estimating boards found in the Trello workspace. "
              "Either none exist or build_trello_index couldn't reach "
              "them — verification can't run.")
        return
    print(f"Searching {len(est_keys)} estimating-board cards for "
          f"possible matches against {len(rows)} WC rows…")
    print("=" * 72)

    classified_rows = bucket_rows(rows, idx)
    suspects = []
    for r in classified_rows:
        if r.get("_cat") == "estimating":
            continue  # already correctly classified
        customer = (r.get("customer") or "").strip()
        if not customer:
            continue
        variants = list(_name_variants(customer))
        # difflib match against every estimating-board canon key. Take
        # the best score across all variants × all keys.
        best_score = 0.0
        best_key = None
        for v in variants:
            if not v:
                continue
            for k in est_keys:
                ratio = _dl.SequenceMatcher(None, v, k).ratio()
                if ratio > best_score:
                    best_score = ratio
                    best_key = k
        if best_score >= fuzzy_cutoff and best_key:
            hit = est_keys[best_key]
            suspects.append({
                "customer":   customer,
                "current":    r.get("_cat"),
                "score":      best_score,
                "match_key":  best_key,
                "card_name":  hit.get("card_name") or "",
                "assignee":   hit.get("list_name") or "",
                "card_url":   hit.get("card_url") or "",
            })
    suspects.sort(key=lambda s: -s["score"])
    if not suspects:
        print(f"No likely-missed estimating matches "
              f"(cutoff={fuzzy_cutoff}). All rows in the report appear "
              f"correctly NOT in someone's estimating column.")
        return
    print(f"Found {len(suspects)} possible estimating matches the "
          f"auto-classifier missed:\n")
    for s in suspects[:max_suggest * 20]:
        score_pct = int(round(s["score"] * 100))
        print(f"  [{score_pct:3d}%]  WC row {s['customer']!r}")
        print(f"          currently → {s['current']}")
        print(f"          looks like → {s['card_name']!r} "
              f"(estimator: {s['assignee']!r})")
        if s["card_url"]:
            print(f"          {s['card_url']}")
        print()
    print(f"Tip: in the WC Audit panel, right-click → Move to → "
          f"Estimating to override any of these manually.")


def _cli(argv):
    """Tiny CLI so the user can spot-check classification without
    opening the GUI. Subcommands:

      python wc_audit.py debug              <xlsx>  — explain not_sold
      python wc_audit.py debug-all          <xlsx>  — explain every row
      python wc_audit.py whymatch           <name>  — variants + nearest
      python wc_audit.py verify-estimating  <xlsx>  — find likely-missed
                                                      estimating matches
    """
    if not argv or argv[0] in ("-h", "--help"):
        print("Usage:")
        print("  python wc_audit.py debug              <path-to-leads.xlsx>")
        print("  python wc_audit.py debug-all          <path-to-leads.xlsx>")
        print("  python wc_audit.py verify-estimating  <path-to-leads.xlsx>")
        print("  python wc_audit.py whymatch           <Customer name>")
        return 0
    cmd = argv[0]
    if cmd in ("debug", "debug-all"):
        if len(argv) < 2:
            print("error: need an xlsx path")
            return 2
        path = argv[1]
        if not os.path.isfile(path):
            print(f"error: file not found: {path}")
            return 2
        explain_not_sold(path, only_not_sold=(cmd == "debug"))
        return 0
    if cmd == "verify-estimating":
        if len(argv) < 2:
            print("error: need an xlsx path")
            return 2
        path = argv[1]
        if not os.path.isfile(path):
            print(f"error: file not found: {path}")
            return 2
        verify_estimating(path)
        return 0
    if cmd == "whymatch":
        if len(argv) < 2:
            print("error: need a Customer name (quote it if multi-word)")
            return 2
        name = " ".join(argv[1:])
        print(f"Building Trello index…")
        idx = build_trello_index()
        print(f"Index size: {len(idx)} canonicalized names\n")
        variants = list(_name_variants(name))
        print(f"Customer:        {name!r}")
        print(f"Variants tried:  {variants}")
        hits = [(v, idx[v]) for v in variants if v and v in idx]
        if hits:
            for v, hit in hits:
                print(f"\nHIT on {v!r}:")
                print(f"  board: {hit.get('board_name')!r}")
                print(f"  lane:  {hit.get('list_name')!r}")
                print(f"  card:  {hit.get('card_name')!r}")
            cat, assignee = classify_row(name, idx)
            print(f"\nClassified as: {cat}"
                  + (f" (assignee={assignee!r})" if assignee else ""))
        else:
            print("\nNo hit on any variant.")
            near = _nearest_index_keys(variants, idx)
            if near:
                print(f"Nearest index keys: {near}")
            print("\nClassified as: not_sold")
        return 0
    print(f"error: unknown command {cmd!r}. Use --help.")
    return 2


if __name__ == "__main__":
    import sys
    sys.exit(_cli(sys.argv[1:]))
