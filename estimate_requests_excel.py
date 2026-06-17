"""Estimate-request workbook — companion source-of-truth for the 48h
SLA tracker. Mirrors the snapshots_excel.py pattern: per-year workbook
with three sheets (Pending / Overdue / Completed), idempotent upsert by
request_id, best-effort writes so a locked file never blocks the
persistence layer.

Public entry points:
    sync_request(record)        — upsert; auto-routes between sheets
    mark_completed_in_excel(id) — convenience (also handled by sync_request
                                  when the record's status is 'completed')
    workbook_path(year)         — for the Open Workbook button
    reconcile()                 — full re-sync from persistence (rare;
                                  used by the recovery button)
    pending_count(year=None)    — sidebar/badge count
"""
from __future__ import annotations

import datetime as _dt
import os
import time

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import persistence as per


# ── Paths ───────────────────────────────────────────────────────────────

_DEFAULT_ROOT = r"X:\IE_Public\Estimate Requests"
_root = _DEFAULT_ROOT


def set_root(path):
    """Override the workbook root. Tests use this to point at a temp
    directory. Pass None to revert to the default share location."""
    global _root
    _root = (path or _DEFAULT_ROOT)


def get_root():
    return _root


def workbook_path(year):
    """Absolute path to the per-year workbook. Directory is NOT created
    here — _ensure_dir handles that at write time so callers can ask
    for the path even when the share is unreachable."""
    return os.path.join(_root, f"Estimate Requests {year}.xlsx")


def _ensure_dir(path):
    try:
        os.makedirs(path, exist_ok=True)
    except OSError:
        pass


# ── Sheet layout ────────────────────────────────────────────────────────

COLUMNS = [
    "Request ID",          # stable hash (lets reconcile() find rows)
    "Received",
    "Deadline",
    "Source",              # XA / Email
    "Insured",
    "Carrier",
    "Claim#",
    "Adjuster",
    "Estimator",
    "Trello card",
    "Status",
    "Acked at",
    "Completed at",
    "Extensions",          # joined: "+24h: late photos; +12h: holiday"
    "Source link",
    "Notes",
]
_COL_INDEX = {n: i + 1 for i, n in enumerate(COLUMNS)}

_SHEET_PENDING   = "Pending"
_SHEET_OVERDUE   = "Overdue"
_SHEET_COMPLETED = "Completed"
_ALL_SHEETS = (_SHEET_PENDING, _SHEET_OVERDUE, _SHEET_COMPLETED)


def _sheet_name(base, year):
    """Two-digit year suffix matches the snapshots_excel convention so a
    user with both workbooks open sees consistent tab naming."""
    return f"{base} {year % 100:02d}"


# ── Workbook plumbing ───────────────────────────────────────────────────

def _ensure_workbook(year):
    """Open the year's workbook, creating it (with the three tabs and
    header rows) when missing. Returns (workbook, path). Caller saves."""
    path = workbook_path(year)
    _ensure_dir(os.path.dirname(path))
    if os.path.isfile(path):
        try:
            return openpyxl.load_workbook(path), path
        except Exception:
            # Corrupted (truncated mid-flush) — back it up, start fresh.
            backup = path + f".corrupt-{int(time.time())}"
            try:
                os.replace(path, backup)
            except OSError:
                pass
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    for base in _ALL_SHEETS:
        ws = wb.create_sheet(_sheet_name(base, year))
        _write_header(ws)
    return wb, path


def _write_header(ws):
    fill = PatternFill("solid", fgColor="DDEBF7")
    font = Font(bold=True)
    for name, idx in _COL_INDEX.items():
        c = ws.cell(1, idx, name)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
    widths = {
        "Request ID":   13, "Received":   17, "Deadline":     17,
        "Source":        9, "Insured":    24, "Carrier":      14,
        "Claim#":       16, "Adjuster":   20, "Estimator":    12,
        "Trello card":  30, "Status":     12, "Acked at":     17,
        "Completed at": 17, "Extensions": 28, "Source link":  18,
        "Notes":        30,
    }
    for name, w in widths.items():
        idx = _COL_INDEX.get(name)
        if idx:
            ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"


def _find_row_by_id(ws, request_id):
    """Look up by Request ID (column A). Returns row number or None."""
    if not request_id:
        return None
    col = _COL_INDEX["Request ID"]
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is None:
            continue
        if str(v).strip() == request_id:
            return r
    return None


def _delete_from_other_sheets(wb, year, request_id, target_base):
    """Remove the row from any sheet other than target_base. Without
    this, moving Pending → Completed would leave a stale Pending row
    saying the inquiry is still open."""
    for base in _ALL_SHEETS:
        if base == target_base:
            continue
        title = _sheet_name(base, year)
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        r = _find_row_by_id(ws, request_id)
        if r is not None:
            ws.delete_rows(r, 1)


# ── Record → cells mapping ──────────────────────────────────────────────

def _format_dt(iso: str | None) -> str:
    """Pretty-print an ISO timestamp for display. Falls back to the raw
    string if it doesn't parse — we never want a parse error to wedge
    the sync."""
    if not iso:
        return ""
    try:
        dt = _dt.datetime.fromisoformat(iso.split(".")[0].rstrip("Z"))
        return dt.strftime("%Y-%m-%d %H:%M")
    except (ValueError, AttributeError):
        return str(iso)


def _format_extensions(exts: list[dict] | None) -> str:
    if not exts:
        return ""
    parts = []
    for e in exts:
        if not isinstance(e, dict):
            continue
        hrs = e.get("hours") or 0
        reason = (e.get("reason") or "").strip() or "(no reason)"
        parts.append(f"+{hrs}h: {reason}")
    return "; ".join(parts)


def _record_to_cells(rec: dict) -> dict[str, str]:
    """Build the {column → value} dict for one request. Status is taken
    from the stored record; in-memory promotion to 'overdue' for routing
    happens in _route_for, not here, so the cell still shows the stored
    value (acked) — Excel users see "overdue" rows by sheet membership."""
    src_label = "XA" if rec.get("source") == "xa" else "Email"
    if not rec.get("uses_xa", True) and rec.get("source") == "adjuster_email":
        # Render non-XA adjuster contact specifically — these need
        # direct Outlook reply rather than XA paste.
        src_label = "Email (no-XA)"
    card_label = (rec.get("card_name") or "").strip()
    if card_label and rec.get("card_url"):
        card_label = f"{card_label}  ({rec['card_url']})"
    elif rec.get("card_url"):
        card_label = rec["card_url"]
    return {
        "Request ID":   rec.get("request_id", ""),
        "Received":     _format_dt(rec.get("received_at")),
        "Deadline":     _format_dt(rec.get("deadline")),
        "Source":       src_label,
        "Insured":      rec.get("insured", "") or rec.get("email_subject", ""),
        "Carrier":      rec.get("carrier", ""),
        "Claim#":       rec.get("claim", ""),
        "Adjuster":     rec.get("adjuster", ""),
        "Estimator":    rec.get("estimator", ""),
        "Trello card":  card_label,
        "Status":       rec.get("status", ""),
        "Acked at":     _format_dt(rec.get("acked_at")),
        "Completed at": _format_dt(rec.get("completed_at")),
        "Extensions":   _format_extensions(rec.get("extensions")),
        "Source link":  rec.get("source_link", ""),
        "Notes":        rec.get("completed_note", "") or rec.get("dismiss_reason", ""),
    }


def _route_for(rec: dict) -> str:
    """Decide which sheet a record belongs in. Dismissed rows are NOT
    written to Excel at all — the workbook is a working tracker, not an
    audit log. Completed go to Completed. Past-deadline non-completed go
    to Overdue. Everything else goes to Pending (acked or pending_ack)."""
    status = rec.get("status", "")
    if status == "dismissed":
        return ""
    if status == "completed":
        return _SHEET_COMPLETED
    # Overdue determination uses the stored deadline, not the in-memory
    # promoted status — that way the Excel sheet stays decisive even if
    # the GUI hasn't been opened to hydrate the records yet.
    deadline = rec.get("deadline") or ""
    try:
        d = _dt.datetime.fromisoformat(deadline.split(".")[0].rstrip("Z"))
        if d < _dt.datetime.now(_dt.timezone.utc).replace(tzinfo=None):
            return _SHEET_OVERDUE
    except (ValueError, AttributeError):
        pass
    return _SHEET_PENDING


# ── Public sync ─────────────────────────────────────────────────────────

def _year_for(rec: dict) -> int:
    """Year the request received_at falls in. Falls back to current year
    when received_at is missing or unparseable."""
    iso = rec.get("received_at") or ""
    try:
        return _dt.datetime.fromisoformat(
            iso.split(".")[0].rstrip("Z")).year
    except (ValueError, AttributeError):
        return _dt.datetime.now().year


def sync_request(record: dict) -> bool:
    """Upsert one request into the year's workbook. Returns True on
    write, False when the target sheet was '' (dismissed) or the
    workbook was locked / unreachable.

    The caller (estimate_requests._excel_sync_safe) wraps this in a
    bare try/except so partial Excel failures never break persistence —
    but most expected failures (locked file, missing share) are also
    caught here so the logs stay quiet for the common cases.
    """
    if not isinstance(record, dict):
        return False
    target_base = _route_for(record)
    if not target_base:
        # Dismissed — make sure no stale row lingers, but don't add one.
        try:
            year = _year_for(record)
            wb, path = _ensure_workbook(year)
            _delete_from_other_sheets(wb, year, record["request_id"], "<none>")
            wb.save(path)
        except (OSError, PermissionError):
            return False
        return True
    year = _year_for(record)
    try:
        wb, path = _ensure_workbook(year)
    except (OSError, PermissionError):
        return False
    target_title = _sheet_name(target_base, year)
    ws = wb[target_title]
    cells = _record_to_cells(record)
    r = _find_row_by_id(ws, record["request_id"])
    if r is None:
        r = (ws.max_row or 1) + 1
    for name, idx in _COL_INDEX.items():
        ws.cell(r, idx, cells.get(name, ""))
    _delete_from_other_sheets(wb, year, record["request_id"], target_base)
    try:
        wb.save(path)
    except (OSError, PermissionError):
        # Most common reason: Excel has the file open. Persistence has
        # the truth; the next sync_request that gets the lock will
        # catch up.
        return False
    return True


def mark_completed_in_excel(request_id: str) -> bool:
    """Convenience wrapper — same effect as sync_request when the
    record's status flips to 'completed' (which the estimate_requests
    module does before calling _excel_sync_safe). Kept as its own entry
    point so callers that bypass the persistence layer (rare) still have
    a one-call path."""
    rec = per.get_estimate_request(request_id)
    if rec is None:
        return False
    rec["status"] = "completed"
    return sync_request(rec)


def reconcile() -> dict:
    """Full re-sync from persistence. Reads every stored request,
    writes (or removes) the corresponding Excel row. Returns
    {"synced": N, "skipped": M, "errors": [...]} for the UI.

    Cheap to run — bounded by the size of estimate_requests in
    persistence. Useful when state.json and the workbook drift (rare,
    but: a fresh checkout, a manual edit, or a crashed sync mid-write).
    """
    synced = 0
    skipped = 0
    errors: list[str] = []
    for rid, rec in per.iter_estimate_requests():
        if not isinstance(rec, dict):
            skipped += 1
            continue
        try:
            if sync_request(rec):
                synced += 1
            else:
                skipped += 1
        except Exception as ex:
            errors.append(f"{rid}: {ex}")
    return {"synced": synced, "skipped": skipped, "errors": errors}


def pending_count(year=None) -> int:
    """Sidebar badge — count of rows on the Pending + Overdue sheets
    combined. Best-effort: returns 0 if the workbook can't be opened."""
    yr = year or _dt.datetime.now().year
    try:
        wb, _ = _ensure_workbook(yr)
    except (OSError, PermissionError):
        return 0
    total = 0
    for base in (_SHEET_PENDING, _SHEET_OVERDUE):
        title = _sheet_name(base, yr)
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        # max_row counts trailing empties; use the Request-ID column to
        # find the real last populated row.
        col = _COL_INDEX["Request ID"]
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, col).value:
                total += 1
    return total


# ── CLI ─────────────────────────────────────────────────────────────────

def _cli(argv):
    if not argv:
        print("Usage:")
        print("  python estimate_requests_excel.py reconcile")
        print("  python estimate_requests_excel.py path [--year=YYYY]")
        print("  python estimate_requests_excel.py count [--year=YYYY]")
        return 1
    cmd = argv[0]
    year = _dt.datetime.now().year
    for a in argv[1:]:
        if a.startswith("--year="):
            try: year = int(a.split("=", 1)[1])
            except ValueError: pass
    if cmd == "reconcile":
        result = reconcile()
        print(f"Synced: {result['synced']}  Skipped: {result['skipped']}"
              f"  Errors: {len(result['errors'])}")
        for e in result["errors"][:10]:
            print(f"  {e}")
        return 0
    if cmd == "path":
        print(workbook_path(year))
        return 0
    if cmd == "count":
        print(pending_count(year))
        return 0
    print(f"Unknown command: {cmd}")
    return 2


if __name__ == "__main__":
    import sys
    sys.exit(_cli(sys.argv[1:]))
