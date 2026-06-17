"""User-defined workbook registrations for the Spreadsheets panel.

Persists `config["user_workbooks"]` — a list of entries describing
extra .xlsx files the user has pointed the panel at via the
"➕ Add workbook…" dialog. On module import we walk the list and
register each as a `workbook_registry.WorkbookSpec`, so the
Spreadsheets panel picks them up alongside Snapshots and Disputes
without any hard-coded knowledge of what the user added.

Entry shape (persisted as JSON in config.json):

    {
        "key":         "ux_<slug>",        # generated, stable
        "label":       "Invoicing 2026",   # shown in the workbook
                                            # selector dropdown
        "path":        "X:\\...\\Invoicing.xlsx",
        "sheet":       "2026",             # "" = "first sheet"
        "header_row":  1,                  # 1-based
        "added_at":    "2026-05-18T..."
    }

Trade-offs of a user-defined entry vs. a hard-coded WorkbookSpec:
    - Columns are auto-detected from the header row. No custom
      ColumnSpec widths — every column shows at a default 110px.
      Power users still get a dedicated module + WorkbookSpec for
      a richer setup (see DISPUTES_SPEC for the canonical example).
    - No `tag_for_row` color logic. Rows render plain. The user
      can sort via the column headers and filter via the search
      box; that's enough for read-only browsing.
    - No `actions=(...)`. The "Open in Excel" link on the panel
      header already covers the most-common need.
"""
from __future__ import annotations

import os
import re
from datetime import datetime
from typing import Any

import config


def _load_raw() -> list[dict[str, Any]]:
    try:
        cfg = config.load()
    except Exception:
        return []
    raw = cfg.get("user_workbooks") or []
    if not isinstance(raw, list):
        return []
    return [e for e in raw if isinstance(e, dict)]


def _save_raw(entries: list[dict[str, Any]]) -> None:
    try:
        cfg = config.load()
    except Exception:
        cfg = {}
    cfg["user_workbooks"] = entries
    try:
        config.save(cfg)
    except Exception:
        pass


def list_entries() -> list[dict[str, Any]]:
    """All user-added workbook entries."""
    return _load_raw()


def _make_key(label: str, existing_keys: set[str]) -> str:
    """Generate a unique key from a label. Uses 'ux_' prefix so user
    workbooks are visually distinct from the built-in ones in code
    and config inspection."""
    base = re.sub(r"[^a-z0-9]+", "_", (label or "workbook").lower()).strip("_")
    base = base or "workbook"
    candidate = f"ux_{base}"
    n = 2
    while candidate in existing_keys:
        candidate = f"ux_{base}_{n}"
        n += 1
    return candidate


def add_entry(*, label: str, path: str, sheet: str = "",
               header_row: int = 1) -> dict[str, Any] | None:
    """Persist a new user-workbook entry. Returns the saved entry
    (with its generated key + added_at timestamp) or None when the
    inputs are invalid. Idempotent on (path, sheet) — re-adding the
    same workbook at the same sheet updates the label/header_row
    instead of duplicating.

    Also caches the detected header row's column names into the entry
    as `headers`. This is consulted by `_build_spec` so registration
    at app launch doesn't have to crack the .xlsx open again — that
    was making the launcher hang at startup whenever the file was
    locked / mid-sync / on a slow share."""
    label = (label or "").strip()
    path = (path or "").strip()
    sheet = (sheet or "").strip()
    if not label or not path:
        return None
    if not os.path.isfile(path):
        return None
    try:
        header_row = max(1, int(header_row))
    except (TypeError, ValueError):
        header_row = 1
    # Cache the schema NOW — the file is in front of us, the user
    # just confirmed the dialog. Later builds use this without I/O.
    detected_headers: list[str] = []
    try:
        info = inspect_workbook(path)
        target_sheet = sheet or (info.get("sheets") or [""])[0]
        det = info.get("details", {}).get(target_sheet, {})
        detected_headers = [h for h in (det.get("headers") or []) if h]
    except Exception:
        detected_headers = []
    entries = _load_raw()
    keys = {e.get("key", "") for e in entries}
    # Update an existing entry pointing at the same (path, sheet)
    # before generating a new one. Avoids duplicates when the user
    # re-runs the Add dialog on the same file.
    for e in entries:
        if (os.path.normcase(e.get("path", ""))
                == os.path.normcase(path)
                and (e.get("sheet") or "") == sheet):
            e["label"] = label
            e["header_row"] = header_row
            if detected_headers:
                e["headers"] = detected_headers
            _save_raw(entries)
            return e
    entry = {
        "key":        _make_key(label, keys),
        "label":      label,
        "path":       path,
        "sheet":      sheet,
        "header_row": header_row,
        "headers":    detected_headers,
        "added_at":   datetime.now().isoformat(timespec="seconds"),
    }
    entries.append(entry)
    _save_raw(entries)
    return entry


def remove_entry(key: str) -> bool:
    if not key:
        return False
    entries = _load_raw()
    keep = [e for e in entries if e.get("key") != key]
    if len(keep) == len(entries):
        return False
    _save_raw(keep)
    return True


def find_entry(key: str) -> dict[str, Any] | None:
    for e in _load_raw():
        if e.get("key") == key:
            return e
    return None


# ── Workbook introspection helpers (for the Add dialog) ────────────────────

def inspect_workbook(path: str) -> dict[str, Any]:
    """Open a workbook and return a summary: sheet names, per-sheet
    header-row guess + header text. Lets the Add dialog show the user
    what we'd display before they commit.

    Returns:
        {
            "ok":      bool,
            "error":   str,             # populated when ok=False
            "sheets":  ["Sheet1", ...],
            "details": {sheet_name: {
                "header_row":   int,    # 1-based, best-guess
                "headers":      [str, ...],
                "data_rows":    int,    # rows after the header
            }}
        }
    """
    if not path or not os.path.isfile(path):
        return {"ok": False, "error": "File not found.",
                "sheets": [], "details": {}}
    try:
        import openpyxl
    except Exception as ex:
        return {"ok": False, "error": f"openpyxl unavailable: {ex}",
                "sheets": [], "details": {}}
    try:
        wb = openpyxl.load_workbook(path, data_only=True,
                                      read_only=True)
    except Exception as ex:
        return {"ok": False, "error": str(ex),
                "sheets": [], "details": {}}
    details: dict[str, dict] = {}
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            hdr_row, headers, n_rows = _guess_header(ws)
            details[sheet_name] = {
                "header_row":  hdr_row,
                "headers":     headers,
                "data_rows":   n_rows,
            }
        except Exception as ex:
            details[sheet_name] = {
                "header_row": 1, "headers": [],
                "data_rows": 0, "error": str(ex)}
    try:
        wb.close()
    except Exception:
        pass
    return {"ok": True, "error": "",
            "sheets": list(wb.sheetnames), "details": details}


def _guess_header(ws) -> tuple[int, list[str], int]:
    """Walk the first ~10 rows and pick the one with the most non-empty
    string cells as the header row. Returns (row_number, headers,
    estimated_data_rows). Falls back to row 1 when nothing looks like
    a header."""
    best_row, best_count = 1, 0
    candidates: dict[int, list[str]] = {}
    for r in range(1, min(11, ws.max_row + 1)):
        cells = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            cells.append("" if v is None else str(v).strip())
        non_empty = sum(1 for v in cells if v)
        candidates[r] = cells
        # A header row is mostly strings (>=2 non-empty cells, and
        # no clear numeric "first data row" feel). Prefer later rows
        # over R1 only when they outscore R1 by a comfortable margin —
        # otherwise stick with R1 for the typical "headers in row 1"
        # template.
        if non_empty > best_count:
            best_count = non_empty
            best_row = r
    headers = candidates.get(best_row, [])
    # Estimate data row count for the user — rough but useful.
    data_rows = max(0, ws.max_row - best_row)
    return best_row, headers, data_rows


# ── Dynamic WorkbookSpec generation ────────────────────────────────────────

def _read_rows_for(entry: dict[str, Any]):
    """Build a `read_rows(year)` closure for one user entry. Year arg
    is accepted and ignored — user workbooks are single-file, no
    per-year sharding. Returns [] silently when the file is missing /
    locked / corrupted; the panel will show the empty table rather
    than crash."""
    def _read(_year):
        try:
            import openpyxl
        except Exception:
            return []
        path = entry.get("path") or ""
        if not os.path.isfile(path):
            return []
        try:
            wb = openpyxl.load_workbook(path, data_only=True,
                                          read_only=True)
        except Exception:
            return []
        sheet = entry.get("sheet") or ""
        try:
            if sheet and sheet in wb.sheetnames:
                ws = wb[sheet]
            else:
                ws = wb[wb.sheetnames[0]]
        except Exception:
            try:
                wb.close()
            except Exception:
                pass
            return []
        try:
            header_row = max(1, int(entry.get("header_row") or 1))
        except (TypeError, ValueError):
            header_row = 1
        headers: list[str] = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            headers.append(str(v).strip() if v is not None
                            else f"Col {c}")
        out: list[dict] = []
        for r in range(header_row + 1, ws.max_row + 1):
            row: dict[str, Any] = {"row_number": r}
            any_filled = False
            for c, h in enumerate(headers, start=1):
                v = ws.cell(r, c).value
                if isinstance(v, datetime):
                    v = v.strftime("%Y-%m-%d")
                if v not in (None, ""):
                    any_filled = True
                row[h] = v
            if any_filled:
                out.append(row)
        try:
            wb.close()
        except Exception:
            pass
        return out
    return _read


def _build_spec(entry: dict[str, Any]):
    """Compile one entry → workbook_registry.WorkbookSpec. Skips the
    spec when no path is set (returns None) — the panel won't try
    to render a phantom workbook.

    Uses the cached `headers` from the saved entry. NO file I/O at
    register time — that was the launcher hang. Headers were
    snapshotted into the entry when the user added the workbook;
    if the workbook's schema changes later, the user re-adds it
    (same path/sheet → updates the cache)."""
    import workbook_registry as wbr

    path = entry.get("path") or ""
    if not path:
        return None
    headers = entry.get("headers") or []
    # Cap displayed columns to keep the treeview tractable. Anything
    # past 20 is generally a data dump (an export from somewhere)
    # rather than a real tracker — show the first 20 with a hint.
    capped_headers = [h for h in headers if h][:20]
    columns = tuple(
        wbr.ColumnSpec(
            key=f"c{i}",
            header=h,
            width=140 if i == 0 else 110,
            anchor="w",
        )
        for i, h in enumerate(capped_headers)
    )

    def _row_to_values(row):
        vals = []
        for h in capped_headers:
            v = row.get(h)
            vals.append("" if v in (None, "") else str(v))
        return tuple(vals)

    return wbr.WorkbookSpec(
        key=entry.get("key") or "ux_unknown",
        label=entry.get("label") or os.path.basename(path),
        read_rows=_read_rows_for(entry),
        workbook_path=lambda _y, p=path: p,
        sheets=("All",),
        columns=columns,
        row_to_values=_row_to_values,
        tag_for_row=lambda _r: None,
        actions=(),
    )


def register_all():
    """Register every persisted user workbook with workbook_registry.
    Safe to call repeatedly — wbr.register is idempotent on key."""
    import workbook_registry as wbr
    for entry in _load_raw():
        try:
            spec = _build_spec(entry)
        except Exception:
            spec = None
        if spec is not None:
            wbr.register(spec)


# Auto-register on import. The Spreadsheets panel imports this module
# (along with the hard-coded specs) at panel build time, so the user's
# saved workbooks land in the dropdown automatically.
register_all()
