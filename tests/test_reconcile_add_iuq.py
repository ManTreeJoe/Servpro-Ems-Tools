"""reconcile_with_trello add-missing step (user-requested 2026-06-11):
every job in the Initial Upload Queue that isn't already in the Snapshots
workbook gets inserted into NEW LOSS, so every job we do the initial
upload for is tracked from the start. Dedup is by canonical name.
"""
import openpyxl

import snapshots_excel as sx
from snapshots_excel import (_SHEET_NEW, _SHEET_ATTENTION, _sheet_name,
                            _write_header, _COL_INDEX)


def _seed_workbook(tmp_path, yr, new_loss_names):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(_sheet_name(_SHEET_NEW, yr))
    _write_header(ws)
    for i, nm in enumerate(new_loss_names, start=2):
        ws.cell(i, _COL_INDEX["Name"], nm)
    wb.save(sx.workbook_path(yr))


def test_reconcile_inserts_missing_iuq_jobs(tmp_path, monkeypatch):
    monkeypatch.setattr(sx, "_root", str(tmp_path))
    yr = 2026
    _seed_workbook(tmp_path, yr, ["Existing Job"])

    import initial_upload_queue as iuq
    monkeypatch.setattr(iuq, "_fetch_queue_from_trello", lambda: [
        {"client": "Existing Job"},       # already present → deduped
        {"client": "New IUQ Job"},        # missing → inserted to NEW LOSS
        {"name": "Cardless Manual Row"},  # uses `name` fallback → inserted
        {"client": ""},                   # blank → skipped
    ])
    # No card resolves → every pre-existing row is unresolvable and the
    # move-closed loop sweeps it to Needs Attention. Newly-added IUQ jobs
    # are inserted after the row snapshot, so the loop never touches them.
    monkeypatch.setattr(sx, "_resolve_card_for_client",
                        lambda name: (None, None))

    summary = sx.reconcile_with_trello(yr)

    assert summary["added_new_loss"] == 2
    assert set(summary["added_names"]) == {"New IUQ Job", "Cardless Manual Row"}

    jobs = sx.read_jobs(yr)
    new_loss_names = {(j.get("Name") or "")
                      for j in jobs if j.get("_sheet") == _SHEET_NEW}
    # Freshly-added IUQ jobs land in NEW LOSS (not reprocessed).
    assert {"New IUQ Job", "Cardless Manual Row"} <= new_loss_names
    # The pre-existing row had no matching Trello card → Needs Attention.
    attn_names = {(j.get("Name") or "")
                  for j in jobs if j.get("_sheet") == _SHEET_ATTENTION}
    assert "Existing Job" in attn_names
    assert summary["by_target"]["needs_attention"] >= 1
    # No duplicate insert of the already-present job.
    assert sum(1 for j in jobs if (j.get("Name") or "") == "Existing Job") == 1


def test_reconcile_dedup_is_canonical(tmp_path, monkeypatch):
    # "Smith, John" in the workbook must dedup against "John Smith" in IUQ.
    monkeypatch.setattr(sx, "_root", str(tmp_path))
    yr = 2026
    _seed_workbook(tmp_path, yr, ["Smith, John"])

    import initial_upload_queue as iuq
    monkeypatch.setattr(iuq, "_fetch_queue_from_trello",
                        lambda: [{"client": "John Smith"}])
    monkeypatch.setattr(sx, "_resolve_card_for_client",
                        lambda name: (None, None))

    summary = sx.reconcile_with_trello(yr)
    assert summary["added_new_loss"] == 0


def test_reconcile_iuq_fetch_failure_is_nonfatal(tmp_path, monkeypatch):
    monkeypatch.setattr(sx, "_root", str(tmp_path))
    yr = 2026
    _seed_workbook(tmp_path, yr, ["Existing Job"])

    import initial_upload_queue as iuq
    def _boom():
        raise RuntimeError("trello down")
    monkeypatch.setattr(iuq, "_fetch_queue_from_trello", _boom)
    monkeypatch.setattr(sx, "_resolve_card_for_client",
                        lambda name: (None, None))

    summary = sx.reconcile_with_trello(yr)
    assert summary["added_new_loss"] == 0
    assert any("IUQ fetch failed" in e for e in summary["errors"])
