"""Routing + helper tests for `snapshots_excel.py`.

Pins the 5-rule `_route_for` decision tree, the pinned-then-fuzzy
`_resolve_card_for_client` lookup, and `_build_name_index` /
`_refresh_index_for_sheet` so future reconciler changes can't silently
shift sheet assignments. Routing bugs are tedious to spot from the
spreadsheet itself — every cell looks fine; only the tab is wrong.
"""
import openpyxl
import pytest

import os

import snapshots_excel as sx
from snapshots_excel import (
    _SHEET_NEW, _SHEET_COMPLETED, _SHEET_INCOMPLETE,
    _route_for, _resolve_card_for_client,
    _build_name_index, _refresh_index_for_sheet,
    _sheet_name, _write_header, COLUMNS, _COL_INDEX,
    _flagged_missing, _evidence_cell, _photo_evidence, _row_to_cells,
)


# ── _route_for ────────────────────────────────────────────────────────────

class TestRouteFor:
    """5-rule priority order — locked 2026-05-06.

    1. _route_override                             → that sheet
    2. _existing_comment has cancel/3rd-party/etc → Incomplete
    3. flagged=True                                → NEW LOSS
    4. new_loss=True                               → NEW LOSS
    5. has_claim=False (after flagged/new_loss)    → Incomplete
    6. else                                        → Completed
    """

    def test_override_completed_wins(self):
        # An override hint trumps every other signal — even flagged.
        r = {"_route_override": "completed", "flagged": True}
        assert _route_for(r) == _SHEET_COMPLETED

    def test_override_new_loss_wins(self):
        r = {"_route_override": "new_loss", "_claim": "X-123"}
        assert _route_for(r) == _SHEET_NEW

    def test_override_incomplete_wins(self):
        r = {"_route_override": "incomplete", "_claim": "X-123"}
        assert _route_for(r) == _SHEET_INCOMPLETE

    def test_override_unknown_falls_through(self):
        # An unrecognized override doesn't short-circuit the rest of the
        # decision tree — defends against typos.
        r = {"_route_override": "garbage", "_claim": "X-123"}
        assert _route_for(r) == _SHEET_COMPLETED

    def test_cancel_comment_routes_to_incomplete(self):
        r = {"_existing_comment": "Cancelled - homeowner declined",
              "_claim": "X-123"}
        assert _route_for(r) == _SHEET_INCOMPLETE

    def test_third_party_comment_routes_to_incomplete(self):
        r = {"_existing_comment": "3rd party adjusters handling",
              "_claim": "X-123"}
        assert _route_for(r) == _SHEET_INCOMPLETE

    def test_archived_comment_routes_to_incomplete(self):
        r = {"_existing_comment": "Archived - old loss",
              "_claim": "X-123"}
        assert _route_for(r) == _SHEET_INCOMPLETE

    def test_cancel_comment_beats_flagged(self):
        # Once the user types Cancelled, it sticks — even if the audit
        # surfaces new issues that would otherwise re-route to NEW LOSS.
        r = {"_existing_comment": "Cancelled", "flagged": True}
        assert _route_for(r) == _SHEET_INCOMPLETE

    def test_flagged_routes_to_new_loss(self):
        r = {"flagged": True, "_claim": "X-123"}
        assert _route_for(r) == _SHEET_NEW

    def test_new_loss_flag_routes_to_new_loss(self):
        r = {"new_loss": True, "_claim": "X-123"}
        assert _route_for(r) == _SHEET_NEW

    def test_flagged_with_no_claim_stays_in_new_loss(self):
        # NEW LOSS is the "intake" sheet — missing claim# is expected
        # there. The Incomplete demotion rule (5) doesn't apply to
        # flagged/new_loss rows.
        r = {"flagged": True}
        assert _route_for(r) == _SHEET_NEW

    def test_new_loss_with_no_claim_stays_in_new_loss(self):
        r = {"new_loss": True}
        assert _route_for(r) == _SHEET_NEW

    def test_clean_row_with_claim_routes_to_completed(self):
        r = {"_claim": "X-123"}
        assert _route_for(r) == _SHEET_COMPLETED

    def test_clean_row_with_existing_claim_routes_to_completed(self):
        # Existing workbook claim# also satisfies the has_claim check —
        # so a re-sync from audit doesn't demote a row whose claim the
        # user typed by hand.
        r = {"_existing_claim": "Y-456"}
        assert _route_for(r) == _SHEET_COMPLETED

    def test_clean_row_no_claim_demotes_to_incomplete(self):
        # The 2026-05-12 rule: would-be Completed with no claim# at all
        # gets surfaced in Incomplete as a follow-up queue.
        r = {}
        assert _route_for(r) == _SHEET_INCOMPLETE

    def test_whitespace_only_claim_counts_as_no_claim(self):
        r = {"_claim": "   "}
        assert _route_for(r) == _SHEET_INCOMPLETE


# ── Evidence-based yes/no columns ─────────────────────────────────────────

class TestFlaggedMissing:
    """`_flagged_missing` must match the EXACT labels audit_logic emits —
    "Auth to Perform" (not "ATP"), "Demo pics" (not "demo photo"). The old
    matchers missed all of these, defaulting every column to "yes"."""

    def test_matches_real_form_labels(self):
        # check_forms returns display names, not abbreviations.
        fi = ["Auth to Perform", "Customer Info Form",
              "Customer Equip Resp", "Cert of Satisfaction", "Scope"]
        assert _flagged_missing("ATP", [], fi)
        assert _flagged_missing("CIF", [], fi)
        assert _flagged_missing("CER", [], fi)
        assert _flagged_missing("COS", [], fi)
        assert _flagged_missing("Scope", [], fi)

    def test_matches_real_photo_labels(self):
        pi = ["Initial pics", "Demo pics"]
        assert _flagged_missing("Initial Photos", pi, [])
        assert _flagged_missing("Demo Photos", pi, [])

    def test_cer_not_falsely_flagged_by_cos_label(self):
        # "Cert of Satisfaction" contains the substring "cer" — CER must
        # NOT be flagged just because COS is missing.
        assert not _flagged_missing("CER", [], ["Cert of Satisfaction"])

    def test_no_issues_means_not_flagged(self):
        assert not _flagged_missing("ATP", [], [])
        assert not _flagged_missing("Demo Photos", [], [])


class TestEvidenceCell:
    def test_form_audited_present_is_yes(self):
        # Audit ran (form_issues key present) and didn't flag ATP → present.
        cell = _evidence_cell("ATP", {"form_issues": []}, {},
                              present=set(), audited=True)
        assert cell == "yes"

    def test_form_flagged_missing_is_no(self):
        cell = _evidence_cell("ATP", {"form_issues": ["Auth to Perform"]}, {},
                              present=set(), audited=True)
        assert cell == "no"

    def test_form_unaudited_preserves_existing(self):
        # No audit data: don't fabricate. Keep the user's cell.
        cell = _evidence_cell("ATP", {}, {"ATP": "yes"},
                              present=set(), audited=False)
        assert cell == "yes"

    def test_form_unaudited_no_existing_is_dash(self):
        cell = _evidence_cell("ATP", {}, {}, present=set(), audited=False)
        assert cell == "-"

    def test_photo_present_on_disk_is_yes(self):
        cell = _evidence_cell("Demo Photos", {}, {},
                              present={"Demo Photos"}, audited=True)
        assert cell == "yes"

    def test_photo_not_present_not_flagged_is_dash(self):
        # Demo never reached + no folder: honest "-", NOT a fabricated yes.
        cell = _evidence_cell("Demo Photos", {"photo_issues": []}, {},
                              present=set(), audited=True)
        assert cell == "-"

    def test_photo_flagged_missing_is_no(self):
        cell = _evidence_cell("Demo Photos",
                              {"photo_issues": ["Demo pics"]}, {},
                              present=set(), audited=True)
        assert cell == "no"

    def test_final_never_auto_yes_without_files(self):
        # The audit never emits a "Final" signal; Final Photos must depend
        # solely on disk evidence — audited alone is not enough.
        cell = _evidence_cell("Final Photos", {"photo_issues": []}, {},
                              present=set(), audited=True)
        assert cell == "-"


class TestPhotoEvidence:
    def _make_job(self, tmp_path, stage_folders):
        pics = tmp_path / "EMS" / "PICS"
        pics.mkdir(parents=True)
        for name, with_file in stage_folders.items():
            d = pics / name
            d.mkdir()
            if with_file:
                (d / "IMG_0001.jpg").write_bytes(b"x")
        return str(tmp_path)

    def test_detects_nonempty_stage_folders(self, tmp_path):
        path = self._make_job(tmp_path, {"Demo": True, "Final": True})
        present = _photo_evidence({"path": path})
        assert present == {"Demo Photos", "Final Photos"}

    def test_empty_folder_is_not_evidence(self, tmp_path):
        path = self._make_job(tmp_path, {"Demo": False})
        assert _photo_evidence({"path": path}) == set()

    def test_initial_or_inspection_folder_counts(self, tmp_path):
        path = self._make_job(tmp_path, {"Initial Inspection": True})
        assert "Initial Photos" in _photo_evidence({"path": path})

    def test_no_path_returns_empty(self):
        assert _photo_evidence({"client": "Nobody"}) == set()


class TestRowToCellsEvidence:
    """End-to-end: the cells that actually get written to the workbook."""

    def test_mark_completed_row_does_not_fabricate_yes(self):
        # A snapshot-generate / mark_completed row: only client + override,
        # no audit data, no resolvable folder → every yes/no cell honest.
        cells = _row_to_cells({"client": "Smith, John",
                               "_route_override": "completed"})
        for col in ("Initial Photos", "Demo Photos", "Scope", "Final Photos",
                    "Sketch", "ATP", "CIF", "CER", "COS"):
            assert cells[col] == "-", f"{col} fabricated {cells[col]!r}"

    def test_audited_row_marks_missing_no_present_yes(self, tmp_path):
        # Audit flagged ATP + Scope missing; everything else verified.
        pics = tmp_path / "EMS" / "PICS"
        pics.mkdir(parents=True)
        (pics / "Demo").mkdir()
        (pics / "Demo" / "a.jpg").write_bytes(b"x")
        r = {
            "client": "Doe, Jane",
            "path": str(tmp_path),
            "form_issues": ["Auth to Perform", "Scope"],
            "photo_issues": [],
        }
        cells = _row_to_cells(r)
        assert cells["ATP"] == "no"
        assert cells["Scope"] == "no"
        assert cells["CIF"] == "yes"
        assert cells["CER"] == "yes"
        assert cells["COS"] == "yes"
        assert cells["Demo Photos"] == "yes"      # folder has files
        assert cells["Final Photos"] == "-"        # no folder, not flagged

    def test_existing_manual_value_preserved_when_unaudited(self):
        cells = _row_to_cells(
            {"client": "Pat Lee", "_route_override": "completed"},
            existing={"Demo Photos": "yes", "ATP": "no"})
        assert cells["Demo Photos"] == "yes"
        assert cells["ATP"] == "no"


# ── _resolve_card_for_client ──────────────────────────────────────────────

class _FakeTrelloClient:
    """Records calls; returns canned hits per term."""
    def __init__(self, hits_by_term=None, raises=None):
        self.hits_by_term = hits_by_term or {}
        self.raises = raises or {}
        self.calls = []

    def find_cards_by_name(self, term, max_results=1):
        self.calls.append(term)
        if term in self.raises:
            raise self.raises[term]
        return self.hits_by_term.get(term, [])


class _FakePersistence:
    def __init__(self, pinned_by_client=None, aliases_by_client=None):
        self.pinned_by_client = pinned_by_client or {}
        self.aliases_by_client = aliases_by_client or {}

    def get_trello_card_ids(self, name):
        return self.pinned_by_client.get(name, [])

    def client_search_terms(self, name):
        # Mirror the real helper: canonical name + aliases, deduped.
        terms = [name] + list(self.aliases_by_client.get(name, []))
        seen, out = set(), []
        for t in terms:
            key = t.lower().strip()
            if key and key not in seen:
                seen.add(key)
                out.append(t)
        return out


@pytest.fixture
def patched_trello(monkeypatch):
    """Patch the imports `_resolve_card_for_client` does inside its body."""
    fake_tc = _FakeTrelloClient()
    fake_p  = _FakePersistence()
    import sys
    monkeypatch.setitem(sys.modules, "trello_client", fake_tc)
    monkeypatch.setitem(sys.modules, "persistence", fake_p)
    return fake_tc, fake_p


class TestResolveCardForClient:
    def test_empty_name_returns_none_pair(self):
        assert _resolve_card_for_client("") == (None, None)
        assert _resolve_card_for_client("   ") == (None, None)
        assert _resolve_card_for_client(None) == (None, None)

    def test_pinned_id_wins_without_calling_search(self, patched_trello):
        tc, p = patched_trello
        p.pinned_by_client = {"Smith, John": ["card-pinned-123"]}
        cid, err = _resolve_card_for_client("Smith, John")
        assert cid == "card-pinned-123"
        assert err is None
        # Fuzzy-search must not run when a pin exists.
        assert tc.calls == []

    def test_fuzzy_falls_back_to_first_alias_hit(self, patched_trello):
        tc, p = patched_trello
        p.aliases_by_client = {"Doe, Jane": ["Doe Jane 123 Main St"]}
        tc.hits_by_term = {
            "Doe Jane 123 Main St": [{"card_id": "card-alias-1"}]
        }
        cid, err = _resolve_card_for_client("Doe, Jane")
        assert cid == "card-alias-1"
        assert err is None
        # Should have tried canonical name first, then alias.
        assert tc.calls == ["Doe, Jane", "Doe Jane 123 Main St"]

    def test_no_match_returns_none(self, patched_trello):
        cid, err = _resolve_card_for_client("Nobody, At All")
        assert cid is None
        assert err is None  # No errors, just nothing found.

    def test_search_error_surfaces_as_error_string(self, patched_trello):
        tc, _p = patched_trello
        tc.raises = {"Smith, John": RuntimeError("connection refused")}
        cid, err = _resolve_card_for_client("Smith, John")
        assert cid is None
        assert err is not None
        assert "Smith, John" in err
        assert "connection refused" in err

    def test_canonical_hit_short_circuits_alias_search(self, patched_trello):
        tc, p = patched_trello
        p.aliases_by_client = {"Smith, John": ["Smith Property"]}
        tc.hits_by_term = {
            "Smith, John": [{"card_id": "card-canonical"}],
            "Smith Property": [{"card_id": "card-alias"}],
        }
        cid, _err = _resolve_card_for_client("Smith, John")
        assert cid == "card-canonical"
        # Alias should not have been queried.
        assert tc.calls == ["Smith, John"]


# ── _build_name_index / _refresh_index_for_sheet ──────────────────────────

def _make_workbook_with_rows(year, rows_by_sheet):
    """Build an in-memory workbook with the three sheets populated."""
    wb = openpyxl.Workbook()
    # openpyxl always creates a default "Sheet" — repurpose then move.
    default = wb.active
    wb.remove(default)
    for base, names in rows_by_sheet.items():
        title = _sheet_name(base, year)
        ws = wb.create_sheet(title)
        _write_header(ws)
        for i, name in enumerate(names, start=2):
            ws.cell(i, _COL_INDEX["Name"], name)
    return wb


class TestNameIndex:
    YEAR = 2026

    def test_index_finds_rows_across_all_three_sheets(self):
        # _canon_name_key swaps "Last, First" → "first last" so the
        # comma form and "first last" form both hash to the same key.
        wb = _make_workbook_with_rows(self.YEAR, {
            "NEW LOSS":   ["Smith, John"],
            "Completed":  ["Doe, Jane"],
            "Incomplete": ["Lee, Pat"],
        })
        idx = _build_name_index(wb, self.YEAR)
        assert idx["john smith"][0] == _sheet_name("NEW LOSS",   self.YEAR)
        assert idx["jane doe"][0]   == _sheet_name("Completed",  self.YEAR)
        assert idx["pat lee"][0]    == _sheet_name("Incomplete", self.YEAR)

    def test_index_is_case_insensitive_and_trimmed(self):
        wb = _make_workbook_with_rows(self.YEAR, {
            "NEW LOSS": ["  Smith, John  "],
        })
        idx = _build_name_index(wb, self.YEAR)
        # Padding stripped, lowercased, comma-swapped
        assert "john smith" in idx
        assert "  smith, john  " not in idx
        assert "smith, john" not in idx

    def test_index_comma_form_matches_first_last(self):
        # The whole point of the canon: both forms hash to the same key
        # so "Sanchez, Jacqueline" and "Jacqueline Sanchez" don't create
        # a duplicate row.
        wb = _make_workbook_with_rows(self.YEAR, {
            "NEW LOSS":  ["Sanchez, Jacqueline"],
            "Completed": ["Jacqueline Sanchez"],
        })
        idx = _build_name_index(wb, self.YEAR)
        # Both rows hash to "jacqueline sanchez"; second write clobbers
        # the first (last-one-wins on duplicate canon keys).
        assert "jacqueline sanchez" in idx

    def test_index_skips_blank_name_cells(self):
        wb = _make_workbook_with_rows(self.YEAR, {
            "NEW LOSS": ["Smith, John"],
        })
        # Inject a blank row
        title = _sheet_name("NEW LOSS", self.YEAR)
        wb[title].cell(3, _COL_INDEX["Name"], "")
        idx = _build_name_index(wb, self.YEAR)
        assert list(idx.keys()) == ["john smith"]

    def test_refresh_picks_up_added_row(self):
        wb = _make_workbook_with_rows(self.YEAR, {
            "Completed": ["Doe, Jane"],
        })
        idx = _build_name_index(wb, self.YEAR)
        assert "john smith" not in idx
        # Add a row to Completed.
        title = _sheet_name("Completed", self.YEAR)
        wb[title].cell(3, _COL_INDEX["Name"], "Smith, John")
        _refresh_index_for_sheet(wb, self.YEAR, "Completed", idx)
        assert "john smith" in idx
        assert idx["john smith"][1] == 3

    def test_refresh_drops_deleted_row(self):
        wb = _make_workbook_with_rows(self.YEAR, {
            "NEW LOSS": ["Smith, John", "Doe, Jane"],
        })
        idx = _build_name_index(wb, self.YEAR)
        assert "john smith" in idx and "jane doe" in idx
        # Delete the first row.
        title = _sheet_name("NEW LOSS", self.YEAR)
        wb[title].delete_rows(2, 1)
        _refresh_index_for_sheet(wb, self.YEAR, "NEW LOSS", idx)
        # Smith is gone; Jane moved up to row 2.
        assert "john smith" not in idx
        assert idx["jane doe"][1] == 2

    def test_refresh_one_sheet_does_not_affect_others(self):
        wb = _make_workbook_with_rows(self.YEAR, {
            "NEW LOSS":  ["Smith, John"],
            "Completed": ["Doe, Jane"],
        })
        idx = _build_name_index(wb, self.YEAR)
        # Add to Completed; refresh only Completed.
        title = _sheet_name("Completed", self.YEAR)
        wb[title].cell(3, _COL_INDEX["Name"], "Lee, Pat")
        _refresh_index_for_sheet(wb, self.YEAR, "Completed", idx)
        assert "john smith" in idx  # untouched
        assert "pat lee" in idx
