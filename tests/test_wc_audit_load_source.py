"""wc_audit.load_source column mapping.

The live audits on the share start at column A, but load_source used to
read a fixed D-J slice — so `customer` read three columns too far right
and came back empty on every row. Nothing looked broken in the panel
(the table renders `cells`, not this field), but per-row Trello pinning
refuses a blank customer, so 📌 Pin failed on every WC row.

These cover both layouts: header-named columns wherever they sit, and
the legacy positional slice when the headers don't match.
"""
import openpyxl
import pytest

import wc_audit


HEADERS = ["Date Received", "Corporate Ref #", "Project #",
           "Property Type", "Type", "Progress", "Customer"]
ROW_A = ["2026-07-01", "CR-1", "2607-353829WTR",
         "Residential", "WTR", "Intake & Research", "Deliah Medina"]
ROW_B = ["2026-07-02", "CR-2", "2607-352789WTR",
         "Commercial", "OTH", "Contact & Schedule", "Richard & Cathy Jenkins"]


def _book(tmp_path, name, header_row, data_rows, start_col=1):
    """Write a one-sheet xlsx with the block starting at `start_col`."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, h in enumerate(header_row):
        ws.cell(1, start_col + i, h)
    for r, row in enumerate(data_rows, start=2):
        for i, v in enumerate(row):
            ws.cell(r, start_col + i, v)
    path = tmp_path / name
    wb.save(path)
    return str(path)


def test_headers_at_column_a(tmp_path):
    """The real share layout — data starts at A, not D."""
    p = _book(tmp_path, "a.xlsx", HEADERS, [ROW_A, ROW_B])
    rows = wc_audit.load_source(p)
    assert len(rows) == 2
    assert [r["customer"] for r in rows] == ["Deliah Medina",
                                             "Richard & Cathy Jenkins"]
    # The off-by-three used to put the name here instead.
    assert rows[0]["property_type"] == "Residential"
    assert rows[0]["project_num"] == "2607-353829WTR"
    assert rows[0]["progress"] == "Intake & Research"


def test_headers_at_column_d(tmp_path):
    """Same headers, shifted to the historic D start — still found."""
    p = _book(tmp_path, "d.xlsx", HEADERS, [ROW_A], start_col=4)
    rows = wc_audit.load_source(p)
    assert len(rows) == 1
    assert rows[0]["customer"] == "Deliah Medina"
    assert rows[0]["project_num"] == "2607-353829WTR"


def test_positional_fallback_when_headers_unrecognised(tmp_path):
    """Headers renamed beyond recognition → fall back to the D-J slice,
    which is what the old code always did."""
    junk = ["a", "b", "c", "d", "e", "f", "g"]
    p = _book(tmp_path, "junk.xlsx", junk, [ROW_A], start_col=4)
    rows = wc_audit.load_source(p)
    assert len(rows) == 1
    assert rows[0]["customer"] == "Deliah Medina"


def test_blank_rows_dropped(tmp_path):
    """A row empty across every mapped column is not a row."""
    p = _book(tmp_path, "blank.xlsx", HEADERS,
              [ROW_A, [None] * 7, ["", "", "", "", "", "", ""], ROW_B])
    rows = wc_audit.load_source(p)
    assert [r["customer"] for r in rows] == ["Deliah Medina",
                                             "Richard & Cathy Jenkins"]


def test_customer_is_stripped(tmp_path):
    """Pinning keys off this name, so stray whitespace must not survive."""
    row = list(ROW_A)
    row[6] = "  Deliah Medina  "
    p = _book(tmp_path, "ws.xlsx", HEADERS, [row])
    assert wc_audit.load_source(p)[0]["customer"] == "Deliah Medina"


def test_missing_determination_column_is_empty(tmp_path):
    """Sheets that stop at Customer have no AP column — that's not an
    error, it just means no terminal-state determination."""
    p = _book(tmp_path, "nodet.xlsx", HEADERS, [ROW_A])
    assert wc_audit.load_source(p)[0]["not_sold_det"] == ""


def test_determination_found_by_header(tmp_path):
    """When the column IS present it's read by name, wherever it sits."""
    hdr = HEADERS + ["Not Sold/Cancelled Determination"]
    p = _book(tmp_path, "det.xlsx", hdr, [ROW_A + ["Cancelled by customer"]])
    assert wc_audit.load_source(p)[0]["not_sold_det"] == "Cancelled by customer"


def test_duplicate_header_keeps_the_first(tmp_path):
    """A repeated label later in the sheet must not steal the column."""
    hdr = HEADERS + ["Customer"]
    p = _book(tmp_path, "dupe.xlsx", hdr, [ROW_A + ["WRONG"]])
    assert wc_audit.load_source(p)[0]["customer"] == "Deliah Medina"
