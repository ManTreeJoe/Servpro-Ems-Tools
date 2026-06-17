"""Tracked-tab manual edits (2026-06-12): snapshots_excel.update_row_fields
(hand-edit cells) + move_existing_row to the new Needs Attention sheet.
Both back the Snapshot Tracked tab's ✎ Edit dialog.
"""
import openpyxl

import snapshots_excel as sx
from snapshots_excel import (_SHEET_COMPLETED, _SHEET_ATTENTION, _SHEET_NEW,
                            _sheet_name, _write_header, _COL_INDEX)


def _seed(tmp_path, yr, sheet_base, name):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(_sheet_name(sheet_base, yr))
    _write_header(ws)
    ws.cell(2, _COL_INDEX["Name"], name)
    wb.save(sx.workbook_path(yr))


def test_update_row_fields_writes_editable_cells(tmp_path, monkeypatch):
    monkeypatch.setattr(sx, "_root", str(tmp_path))
    yr = 2026
    _seed(tmp_path, yr, _SHEET_COMPLETED, "Smith John")

    res = sx.update_row_fields("Smith John", {
        "Carrier": "State Farm", "Claim#": "ABC-123",
        "Type of Loss": "Water", "Comment": "edited by hand",
    }, year=yr)
    assert res["ok"] is True
    assert set(res["updated"]) >= {"Carrier", "Claim#", "Type of Loss", "Comment"}

    rows = {(r.get("Name") or ""): r for r in sx.read_jobs(yr)}
    r = rows["Smith John"]
    assert r["Carrier"] == "State Farm"
    assert r["Claim#"] == "ABC-123"
    assert r["Type of Loss"] == "Water"
    assert r["Comment"] == "edited by hand"


def test_update_row_fields_ignores_unknown_columns(tmp_path, monkeypatch):
    monkeypatch.setattr(sx, "_root", str(tmp_path))
    yr = 2026
    _seed(tmp_path, yr, _SHEET_COMPLETED, "Doe Jane")
    res = sx.update_row_fields("Doe Jane", {"Bogus Col": "x"}, year=yr)
    assert res["ok"] is False  # nothing writable → not ok


def test_update_row_fields_missing_client(tmp_path, monkeypatch):
    monkeypatch.setattr(sx, "_root", str(tmp_path))
    yr = 2026
    _seed(tmp_path, yr, _SHEET_COMPLETED, "Present Person")
    res = sx.update_row_fields("Nobody Here", {"Carrier": "x"}, year=yr)
    assert res["ok"] is False
    assert "not found" in res["error"].lower()


def test_move_existing_row_to_needs_attention(tmp_path, monkeypatch):
    monkeypatch.setattr(sx, "_root", str(tmp_path))
    yr = 2026
    _seed(tmp_path, yr, _SHEET_NEW, "Mover Mike")

    assert sx.move_existing_row("Mover Mike", _SHEET_ATTENTION, year=yr) is True

    rows = sx.read_jobs(yr)
    by_name = {(r.get("Name") or ""): r.get("_sheet") for r in rows}
    assert by_name["Mover Mike"] == _SHEET_ATTENTION
    # Not left behind in NEW LOSS.
    assert sum(1 for r in rows if (r.get("Name") or "") == "Mover Mike") == 1
