"""Snapshots Spreadsheet — standalone tool panel.

Mirrors the spreadsheet viewer that lived inside the EMS Snapshot tool
(`SnapshotApp._build_spreadsheet_viewer`) so the workbook can be browsed
/ reconciled / synced without going through any automation flow. Same
widgets, same buttons, same data — just hosted on the launcher's nav
strip directly.

Public entry points:
    SpreadsheetApp(parent)   — embeddable ToolPanel
    main()                   — run standalone via launcher dispatcher
"""
import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime

import ctk_helpers as ctkh
import snapshots_excel as sx
from theme import (WHITE, BG, TEXT_DARK, TEXT_GRAY, BORDER, SURFACE_2,
                    SUCCESS_BG, WARN_BG, DANGER_BG, WARN_FG,
                    NEUTRAL_HOVER)
from tool_panel import ScrollableFrame, ToolPanel, run_standalone
from ui_buttons import (done_button, link_button, secondary_button,
                          send_button, warn_button)
import workbook_registry as wbr


class SpreadsheetApp(ToolPanel):
    """Browse + reconcile + sync the Snapshots workbook."""

    TOOL_GEOMETRY_KEY = "spreadsheet_geometry"

    def __init__(self, parent):
        super().__init__(parent)
        self.configure(bg=BG)
        self._cache = {"rows": [], "year": None}
        self._ctrl_f_handler = None
        self._build_ui()
        self.after(100, self._refresh)

    def on_close(self):
        # Symmetric unbind for the global Ctrl+F handler — without
        # this, a launcher cycle leaks the handler and Ctrl+F keeps
        # firing into a destroyed widget.
        try:
            self.unbind_all("<Control-f>")
        except tk.TclError:
            pass
        super().on_close()

    def _build_ui(self):
        # Top control band — workbook + year + sheet filter + search on
        # the left, action buttons on the right. The workbook selector
        # is new — used to pick between registered workbooks
        # (Snapshots today, more as added).
        ctl = tk.Frame(self, bg=BG, padx=14, pady=8)
        ctl.pack(fill="x")
        self._ctl_row = ctl

        # Workbook selector. Loaded from workbook_registry; persisted
        # across sessions. When only one workbook is registered we
        # still show the dropdown — gives the user a visible hook for
        # the future when more get added.
        tk.Label(ctl, text="Workbook:",
                 font=("Segoe UI Variable", 9, "bold"),
                 bg=BG, fg=TEXT_DARK).pack(side="left")
        specs = wbr.all_specs()
        wb_keys = [s.key for s in specs]
        wb_labels = [s.label for s in specs]
        try:
            import persistence as _per
            saved_key = _per.get("spreadsheet_active_workbook") or ""
        except Exception:
            saved_key = ""
        if saved_key not in wb_keys:
            saved_key = wbr.default_key()
        self._workbook_key = saved_key
        self._workbook_var = tk.StringVar(
            value=next((s.label for s in specs if s.key == saved_key),
                        wb_labels[0] if wb_labels else ""))
        wb_cb = ttk.Combobox(ctl, textvariable=self._workbook_var,
                              values=wb_labels, state="readonly",
                              width=16, font=("Segoe UI Variable", 9))
        wb_cb.pack(side="left", padx=(6, 4))
        wb_cb.bind("<<ComboboxSelected>>",
                    lambda *_: self._on_workbook_changed())
        self._wb_cb = wb_cb

        # ➕ Add workbook — opens the Add dialog which lets the user
        # point at any .xlsx file and surface it in the panel. The
        # entry persists in config["user_workbooks"] so it's still
        # here on the next launch. ✕ removes the currently-selected
        # workbook (only enabled for user-added ones — the built-in
        # specs can't be removed).
        secondary_button(
            ctl, "➕ Add",
            command=self._open_add_workbook_dialog,
            tooltip=("Point the panel at another .xlsx — auto-detects "
                     "columns from the workbook's header row. Saved "
                     "across sessions.")
        ).pack(side="left", padx=(0, 2))
        self._remove_btn = secondary_button(
            ctl, "✕",
            command=self._remove_active_user_workbook,
            tooltip="Remove this user-added workbook from the panel."
        )
        self._remove_btn.pack(side="left", padx=(0, 12))

        tk.Label(ctl, text="Year:",
                 font=("Segoe UI Variable", 9, "bold"),
                 bg=BG, fg=TEXT_DARK).pack(side="left")
        self._year_var = tk.IntVar(value=datetime.today().year)
        tk.Spinbox(ctl, from_=2020, to=2099, increment=1,
                    textvariable=self._year_var, width=6,
                    font=("Segoe UI Variable", 9)).pack(side="left", padx=(6, 12))

        tk.Label(ctl, text="Sheet:",
                 font=("Segoe UI Variable", 9, "bold"),
                 bg=BG, fg=TEXT_DARK).pack(side="left")
        self._sheet_var = tk.StringVar(value="All")
        # Sheet filter values come from the active spec — workbooks
        # with different tabs (e.g. Open/Paid vs NEW LOSS/Completed)
        # populate this list themselves.
        active = wbr.get(self._workbook_key)
        sheet_values = list(active.sheets) if active else ["All"]
        self._sheet_cb = ttk.Combobox(ctl, textvariable=self._sheet_var,
                                       values=sheet_values,
                                       state="readonly", width=14,
                                       font=("Segoe UI Variable", 9))
        self._sheet_cb.pack(side="left", padx=(6, 12))
        sheet_cb = self._sheet_cb  # local alias for the existing bind below

        tk.Label(ctl, text="🔍",
                 font=("Segoe UI Variable", 9, "bold"),
                 bg=BG, fg=TEXT_DARK).pack(side="left")
        self._filter_var = tk.StringVar()
        filt = tk.Entry(ctl, textvariable=self._filter_var,
                        font=("Segoe UI Variable", 10), width=22,
                        bg=WHITE, relief="solid", bd=1)
        filt.pack(side="left", padx=(6, 8), fill="x", expand=True)
        self._search_entry = filt
        # Esc clears, Ctrl+F focuses (gated on winfo_ismapped so the
        # shortcut doesn't leak across panels). Mirrors the Hygiene
        # search affordance so the keyboard story is identical
        # everywhere we have a row-filter box.
        def _clear_filter(_e=None):
            try: self._filter_var.set("")
            except tk.TclError: pass
            return "break"
        filt.bind("<Escape>", _clear_filter)
        def _focus_filter(_e=None):
            try:
                if not self.winfo_ismapped():
                    return None
                self._search_entry.focus_set()
                self._search_entry.select_range(0, "end")
            except tk.TclError:
                return None
            return "break"
        self.bind_all("<Control-f>", _focus_filter, add="+")
        self._ctrl_f_handler = _focus_filter

        link_button(ctl, "📂 Open in Excel", padx=10, pady=3,
                     command=lambda: sx.open_in_excel(self._year_var.get()),
                     tooltip="Open the workbook in Excel for direct editing"
                     ).pack(side="right")
        link_button(ctl, "📋 Today's digest", padx=10, pady=3,
                     command=self._open_daily_digest,
                     tooltip=("Write today's automatic activity log "
                              "(intakes, closeouts, audit resolutions, "
                              "Trello comments) and open it for review")
                     ).pack(side="right", padx=(0, 6))
        secondary_button(ctl, "🔄 Refresh", padx=12, pady=3,
                          command=self._refresh,
                          tooltip="Re-read the workbook from disk"
                          ).pack(side="right", padx=(0, 6))

        self._path_lbl = tk.Label(self, text="",
                                    font=("Segoe UI Variable", 8, "italic"),
                                    bg=BG, fg=TEXT_GRAY,
                                    padx=14, anchor="w")
        self._path_lbl.pack(fill="x")
        self._status_lbl = tk.Label(self, text="",
                                      font=("Segoe UI Variable", 8, "italic"),
                                      bg=BG, fg=TEXT_GRAY,
                                      padx=14, anchor="w")
        self._status_lbl.pack(fill="x", pady=(0, 4))

        # Treeview — compact tabular view. Click a row to select; the
        # actual edits happen in Excel proper, so we don't need
        # in-place editing. Column set comes from the active spec so a
        # different workbook can have a different shape (e.g. an
        # invoicing book with Amount/DueDate columns).
        self._tree_wrap = tk.Frame(self, bg=BG, padx=14)
        self._tree_wrap.pack(fill="both", expand=True, pady=(0, 8))
        self._tree: ttk.Treeview | None = None
        self._tree_vsb: ttk.Scrollbar | None = None
        self._build_tree_for_active_spec()

        # Bottom action row — populated per-spec.
        self._bot = tk.Frame(self, bg=BG, padx=14, pady=10)
        self._bot.pack(fill="x", side="bottom")
        self._build_actions_for_active_spec()

        # Hook controls.
        sheet_cb.bind("<<ComboboxSelected>>", lambda *_: self._redraw())
        # Debounce the text filter — full Treeview rebuild on every
        # keystroke is the dominant cause of typing lag on this panel.
        # 150ms collapses a multi-char paste / fast type into one redraw.
        self._filter_debouncer = ctkh.Debouncer(self, 150)
        self._filter_var.trace_add(
            "write", lambda *_: self._filter_debouncer.fire(self._redraw))
        self._year_var.trace_add("write", lambda *_: self._refresh())
        # Initial enabled-state of the ✕ Remove button now that the
        # spec is known and the button exists.
        self._refresh_remove_button()

    # ── Active-spec helpers ─────────────────────────────────────────────────
    def _active_spec(self):
        """Return the currently-selected WorkbookSpec, or None when
        the registry is empty (shouldn't happen in practice — Snapshots
        is registered at import)."""
        return wbr.get(self._workbook_key)

    def _build_tree_for_active_spec(self):
        """Build (or rebuild) the Treeview for the currently-active
        spec's column set. Destroys any prior tree first."""
        if self._tree is not None:
            try: self._tree.destroy()
            except Exception: pass
        if self._tree_vsb is not None:
            try: self._tree_vsb.destroy()
            except Exception: pass
        spec = self._active_spec()
        cols_specs = spec.columns if spec else ()
        col_ids = tuple(c.key for c in cols_specs)
        self._tree = ttk.Treeview(self._tree_wrap, columns=col_ids,
                                   show="headings", height=20)
        for c in cols_specs:
            self._tree.heading(c.key, text=c.header)
            self._tree.column(c.key, width=c.width, anchor=c.anchor)
        self._tree_vsb = ttk.Scrollbar(self._tree_wrap, orient="vertical",
                                         command=self._tree.yview)
        self._tree.configure(yscrollcommand=self._tree_vsb.set)
        self._tree.pack(side="left", fill="both", expand=True)
        self._tree_vsb.pack(side="right", fill="y")
        # Pre-configure tag colors so per-row tagging is a key lookup.
        for tag, color in (spec.tag_colors if spec else {}).items():
            try:
                self._tree.tag_configure(tag, background=color)
            except Exception:
                pass

    def _build_actions_for_active_spec(self):
        """Rebuild the bottom action row from the active spec.

        Action factories receive `self` so they can call panel methods
        (e.g. SnapshotsSpec returns lambdas that call self._sync_backlog)."""
        for w in self._bot.winfo_children():
            try: w.destroy()
            except Exception: pass
        spec = self._active_spec()
        if not spec:
            return
        # Map kind → button factory. Kept here, not in the spec, so the
        # spec module doesn't have to import ui_buttons.
        factories = {
            "warn":      warn_button,
            "send":      send_button,
            "secondary": secondary_button,
            "link":      link_button,
        }
        for act in spec.actions:
            factory = factories.get(act.kind) or secondary_button
            cmd = act.command_factory(self)
            btn = factory(self._bot, act.label, padx=14, pady=4,
                           command=cmd, tooltip=act.tooltip or None)
            btn.pack(side="left", padx=(0, 8))
            if act.attr_name:
                setattr(self, act.attr_name, btn)

    def _on_workbook_changed(self):
        """User picked a different workbook. Persist the choice and
        rebuild the tree + actions for the new spec."""
        label = self._workbook_var.get()
        for s in wbr.all_specs():
            if s.label == label:
                self._workbook_key = s.key
                break
        try:
            import persistence as _per
            _per.set_value("spreadsheet_active_workbook", self._workbook_key)
        except Exception:
            pass
        # Sheet filter values change per-workbook.
        spec = self._active_spec()
        if spec:
            self._sheet_cb.configure(values=list(spec.sheets))
            if self._sheet_var.get() not in spec.sheets:
                self._sheet_var.set(spec.sheets[0] if spec.sheets else "All")
        self._build_tree_for_active_spec()
        self._build_actions_for_active_spec()
        self._refresh_remove_button()
        self._refresh()

    def _refresh_remove_button(self):
        """Enable ✕ only when the active workbook is user-added.
        Built-in specs (Snapshots, Disputes) can't be removed via the
        UI — that requires a code change."""
        btn = getattr(self, "_remove_btn", None)
        if btn is None:
            return
        try:
            if (self._workbook_key or "").startswith("ux_"):
                btn.configure(state="normal")
            else:
                btn.configure(state="disabled")
        except tk.TclError:
            pass

    # ── ➕ Add workbook flow ────────────────────────────────────────────────

    def _open_add_workbook_dialog(self):
        """Two-step inline dialog: pick a file, then confirm the sheet
        + label + header row. Auto-detects what we can from the
        workbook so a typical "just add this xlsx" flow is two clicks."""
        path = filedialog.askopenfilename(
            parent=self.winfo_toplevel(),
            title="Pick a workbook to add",
            filetypes=[("Excel workbooks", "*.xlsx *.xlsm"),
                        ("All files", "*.*")],
        )
        if not path:
            return
        info = uwb.inspect_workbook(path)
        if not info.get("ok"):
            messagebox.showerror(
                "Couldn't read workbook",
                info.get("error") or "Unknown error.", parent=self)
            return
        self._open_confirm_workbook_dialog(path, info)

    def _open_confirm_workbook_dialog(self, path, info):
        """Show the auto-detected schema + let the user override the
        label / sheet / header row before saving. Renders a preview of
        the first ~6 detected column headers so they can verify the
        right row was picked."""
        sheets = info.get("sheets") or []
        if not sheets:
            messagebox.showerror(
                "Empty workbook",
                f"No sheets found in:\n{path}", parent=self)
            return

        dlg = tk.Toplevel(self)
        dlg.title("Add workbook")
        dlg.configure(bg=BG)
        dlg.transient(self.winfo_toplevel())
        dlg.grab_set()
        # Cap height to the visible screen so the dialog never
        # extends past the taskbar on a small laptop.
        try:
            screen_h = dlg.winfo_screenheight()
            dh = max(360, min(520, screen_h - 120))
            dlg.geometry(f"560x{dh}")
        except tk.TclError:
            pass

        hdr = tk.Frame(dlg, bg=BG, padx=18, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="➕  Add a workbook",
                 font=("Fraunces", 14, "bold"),
                 bg=BG, fg=TEXT_DARK).pack(anchor="w")
        tk.Label(hdr, text=path,
                 font=("Segoe UI Variable", 9),
                 bg=BG, fg=TEXT_GRAY,
                 wraplength=520, justify="left"
                 ).pack(anchor="w", pady=(2, 0))

        # Form body
        body = tk.Frame(dlg, bg=WHITE, padx=14, pady=12,
                         highlightthickness=1,
                         highlightbackground=BORDER)
        body.pack(fill="both", expand=True, padx=14, pady=(0, 8))

        # Label entry — defaults to the file's basename without extension.
        default_label = os.path.splitext(os.path.basename(path))[0]
        tk.Label(body, text="Label (shown in dropdown)",
                 font=("Segoe UI Variable", 9, "bold"),
                 bg=WHITE, fg=TEXT_DARK
                 ).pack(anchor="w")
        label_var = tk.StringVar(value=default_label)
        ttk.Entry(body, textvariable=label_var,
                  font=("Segoe UI Variable", 10)
                  ).pack(fill="x", pady=(2, 8))

        # Sheet picker — only meaningful when there's more than one
        # sheet. Single-sheet workbooks auto-pick.
        tk.Label(body, text="Sheet",
                 font=("Segoe UI Variable", 9, "bold"),
                 bg=WHITE, fg=TEXT_DARK
                 ).pack(anchor="w")
        sheet_var = tk.StringVar(value=sheets[0])
        sheet_cb = ttk.Combobox(
            body, textvariable=sheet_var, values=sheets,
            state="readonly", font=("Segoe UI Variable", 10))
        sheet_cb.pack(fill="x", pady=(2, 8))

        # Header row override — most sheets are row 1, but trackers
        # sometimes have a title row that pushes the headers to row 3
        # or 4 (the Dispute Tracker template uses row 4).
        det0 = info["details"].get(sheets[0], {})
        hdr_row_var = tk.IntVar(value=int(det0.get("header_row") or 1))
        tk.Label(body, text="Header row (1-based)",
                 font=("Segoe UI Variable", 9, "bold"),
                 bg=WHITE, fg=TEXT_DARK
                 ).pack(anchor="w")
        tk.Spinbox(body, from_=1, to=20, increment=1,
                    textvariable=hdr_row_var, width=6,
                    font=("Segoe UI Variable", 10)
                    ).pack(anchor="w", pady=(2, 8))

        # Preview area — auto-updates when sheet picker changes.
        preview_lbl = tk.Label(
            body, text="", font=("Segoe UI Variable", 9, "italic"),
            bg=WHITE, fg=TEXT_GRAY, wraplength=480,
            justify="left", anchor="w")
        preview_lbl.pack(fill="x", pady=(4, 0))

        def _refresh_preview(*_):
            sh = sheet_var.get()
            det = info["details"].get(sh, {})
            try:
                hr = int(hdr_row_var.get())
            except (TypeError, ValueError, tk.TclError):
                hr = 1
            # When the user changes header_row away from the auto-
            # detected one, we re-inspect inline so the preview
            # matches what they'll actually see in the panel.
            headers = det.get("headers") if hr == det.get(
                "header_row") else None
            if headers is None:
                # Cheap re-read of the requested row only.
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(
                        path, data_only=True, read_only=True)
                    ws = wb[sh] if sh in wb.sheetnames else (
                        wb[wb.sheetnames[0]])
                    headers = [str(ws.cell(hr, c).value or "").strip()
                                for c in range(1, ws.max_column + 1)]
                    wb.close()
                except Exception:
                    headers = []
            non_empty = [h for h in headers if h]
            data_rows = det.get("data_rows", 0)
            shown = ", ".join(non_empty[:6]) or "(none detected)"
            if len(non_empty) > 6:
                shown += f", … (+{len(non_empty) - 6} more)"
            preview_lbl.configure(
                text=(f"Detected {len(non_empty)} column(s): {shown}\n"
                      f"~{data_rows} data row(s) below the header."))

        sheet_cb.bind("<<ComboboxSelected>>",
                       lambda *_: (hdr_row_var.set(
                           info["details"].get(sheet_var.get(), {})
                                .get("header_row", 1)),
                           _refresh_preview()))
        hdr_row_var.trace_add("write", _refresh_preview)
        _refresh_preview()

        # Footer buttons
        bot = tk.Frame(dlg, bg=BG, padx=18, pady=10)
        bot.pack(fill="x", side="bottom")

        def _save():
            label = (label_var.get() or "").strip() or default_label
            sh = sheet_var.get() or sheets[0]
            try:
                hr = int(hdr_row_var.get())
            except (TypeError, ValueError, tk.TclError):
                hr = 1
            entry = uwb.add_entry(label=label, path=path,
                                    sheet=sh, header_row=hr)
            if not entry:
                messagebox.showerror(
                    "Couldn't save",
                    "Workbook entry didn't save — check the path "
                    "and label.", parent=dlg)
                return
            # Re-register everything (idempotent on key) so the new
            # entry's spec lands in the registry, then jump the panel
            # to it.
            try:
                uwb.register_all()
            except Exception:
                pass
            self._workbook_key = entry["key"]
            self._reload_workbook_list(select_key=entry["key"])
            dlg.destroy()

        secondary_button(
            bot, "Cancel", command=dlg.destroy
        ).pack(side="right", padx=(6, 0))
        done_button(
            bot, "💾 Save", command=_save
        ).pack(side="right")

        # Center on parent — clamp to screen.
        try:
            dlg.update_idletasks()
            sw = dlg.winfo_screenwidth()
            sh_px = dlg.winfo_screenheight()
            px = self.winfo_rootx() + (self.winfo_width() // 2)
            py = self.winfo_rooty() + (self.winfo_height() // 2)
            dw = dlg.winfo_width()
            dh = dlg.winfo_height()
            x = max(0, min(px - dw // 2, sw - dw))
            y = max(0, min(py - dh // 2, sh_px - dh - 40))
            dlg.geometry(f"+{x}+{y}")
        except tk.TclError:
            pass

    def _reload_workbook_list(self, *, select_key=""):
        """Refresh the workbook combobox values from the registry and
        optionally select a specific key. Called after the Add
        dialog saves so the new workbook lands in the dropdown
        immediately, no panel reload needed."""
        specs = wbr.all_specs()
        labels = [s.label for s in specs]
        try:
            self._wb_cb.configure(values=labels)
        except (tk.TclError, AttributeError):
            pass
        if select_key:
            for s in specs:
                if s.key == select_key:
                    self._workbook_key = s.key
                    try:
                        self._workbook_var.set(s.label)
                    except tk.TclError:
                        pass
                    break
            self._on_workbook_changed()

    def _remove_active_user_workbook(self):
        """Drop the currently-selected user workbook from persistence
        and from the registry's in-memory order list. Built-in specs
        (Snapshots, Disputes) skip silently — the ✕ button is
        disabled for them but a stray keyboard shortcut shouldn't
        nuke the built-in spec either."""
        if not (self._workbook_key or "").startswith("ux_"):
            return
        entry = uwb.find_entry(self._workbook_key)
        label = entry.get("label") if entry else self._workbook_key
        if not messagebox.askyesno(
                "Remove workbook?",
                f"Remove '{label}' from the Spreadsheets panel?\n\n"
                "This only removes the panel entry — the .xlsx file "
                "itself isn't touched.", parent=self):
            return
        removed_key = self._workbook_key
        uwb.remove_entry(removed_key)
        # Drop from the in-memory registry too so the dropdown updates
        # without an app restart. workbook_registry stores order in a
        # private list; this is the simplest poke that works.
        try:
            wbr._REGISTRY.pop(removed_key, None)
            if removed_key in wbr._ORDER:
                wbr._ORDER.remove(removed_key)
        except Exception:
            pass
        # Reload the dropdown + jump to whatever's now the default.
        self._reload_workbook_list(select_key=wbr.default_key())

    # ── Data ────────────────────────────────────────────────────────────────
    @staticmethod
    def _fmt_date(v):
        if isinstance(v, datetime):
            return v.strftime("%m/%d/%y")
        if isinstance(v, str):
            return v[:10]
        return str(v) if v not in (None, "") else ""

    def _refresh(self):
        spec = self._active_spec()
        if not spec:
            return
        try:
            yr = self._year_var.get()
        except tk.TclError:
            return
        try:
            rows = spec.read_rows(yr)
        except Exception as ex:
            rows = []
            self._status_lbl.config(text=f"Read error: {ex}")
        self._cache["rows"] = rows
        self._cache["year"] = yr
        try:
            self._path_lbl.config(text=f"📁  {spec.workbook_path(yr)}")
        except Exception:
            self._path_lbl.config(text="")
        self._redraw()

    def _redraw(self):
        spec = self._active_spec()
        if not spec or self._tree is None:
            return
        try:
            self._tree.delete(*self._tree.get_children())
        except tk.TclError:
            return
        sheet_filter = self._sheet_var.get()
        text_filter = (self._filter_var.get() or "").strip().lower()
        shown = 0
        for r in self._cache["rows"]:
            # Sheet filter — "All" passes everything, otherwise match
            # the row's `_sheet` field. Workbooks that don't carry a
            # _sheet field always pass.
            if (sheet_filter and sheet_filter != "All"
                    and r.get("_sheet") and r["_sheet"] != sheet_filter):
                continue
            # Text filter — substring match against a workbook-defined
            # search blob. Specs that don't supply one fall back to a
            # generic "stringify all values" approach.
            if text_filter:
                getter = getattr(spec, "search_blob", None)
                if callable(getter):
                    blob = getter(r).lower()
                else:
                    blob = " ".join(str(v) for v in r.values()
                                     if v not in (None, "")).lower()
                if text_filter not in blob:
                    continue
            values = spec.row_to_values(r)
            tag = spec.tag_for_row(r) or ""
            self._tree.insert("", "end", values=values,
                               tags=(tag,) if tag else ())
            shown += 1
        # Optional pending-count badge — only some specs publish one.
        tail = ""
        if spec and getattr(spec, "pending_count", None):
            try:
                pending = spec.pending_count(self._cache["year"])
                if pending:
                    tail = f"  ·  {pending} queued (file locked)"
            except Exception:
                pass
        self._status_lbl.config(
            text=f"{shown} of {len(self._cache['rows'])} rows{tail}")

    # ── Actions ─────────────────────────────────────────────────────────────
    def _sync_backlog(self):
        if not messagebox.askyesno(
                "Sync backlog?",
                "This re-runs the audit across every job in this "
                "year and writes the results into the workbook. "
                "It can take a few minutes. Continue?",
                parent=self):
            return
        self._backlog_btn.config(state="disabled", text="Syncing…")
        self._status_lbl.config(text="Auditing every job in the year — "
                                      "this can take a few minutes…")

        def _finish(text):
            """Called on the UI thread once both the audit AND the save
            have actually finished. Re-enables the button and refreshes
            the table from the now-current workbook."""
            self._backlog_btn.config(
                state="normal",
                text="↻ Sync backlog → spreadsheet")
            self._status_lbl.config(text=text)
            self._refresh()

        def _bg():
            try:
                # snapshot_gui.run_audit is the year-sweep helper.
                from snapshot_gui import run_audit as _run_audit
                results, e = _run_audit()
            except Exception as ex:
                self.after(0, lambda m=str(ex):
                           _finish(f"Backlog sync error: {m}"))
                return

            if e:
                self.after(0, lambda m=str(e):
                           _finish(f"Backlog sync error: {m}"))
                return
            if not results:
                self.after(0, lambda:
                           _finish("Backlog sync — no jobs to process."))
                return

            # `sync_jobs` is fire-and-forget — it spawns its own worker
            # thread. The previous code scheduled `_refresh` 1500 ms
            # later as a guess, which raced the actual `wb.save`. Pass
            # `on_done` so the refresh fires after the save genuinely
            # finishes (or after enqueue, if the file is locked).
            def _on_sync_done(applied, deferred):
                if deferred:
                    text = (f"Backlog sync: {applied} applied · "
                            f"{deferred} queued (file locked)")
                else:
                    text = f"Backlog sync: {applied} applied."
                self.after(0, lambda t=text: _finish(t))

            sx.sync_jobs(results, year=self._year_var.get(),
                          on_done=_on_sync_done)

        threading.Thread(target=_bg, daemon=True).start()

    def _reconcile(self):
        if not messagebox.askyesno(
                "Reconcile with Trello?",
                "This walks every row in the workbook, looks up the "
                "linked Trello card, and moves the row to the right "
                "sheet (NEW LOSS / Completed / Incomplete). Type of "
                "Loss + Carrier + Claim# + dates + Demo Start + "
                "Docusketch + checklist-confirmed photos/scope fill "
                "from the card when the cells are empty. Excel must be "
                "CLOSED. Continue?",
                parent=self):
            return
        self._reconcile_btn.config(state="disabled", text="Reconciling…")
        self._status_lbl.config(text="Reconciling rows against Trello…")

        def _bg():
            summary = err = None
            try:
                def _prog(i, total, name):
                    try:
                        self.after(0, lambda: self._status_lbl.config(
                            text=f"Reconciling {i}/{total}: {name}"))
                    except Exception:
                        pass
                summary = sx.reconcile_with_trello(
                    self._year_var.get(), progress_cb=_prog)
            except Exception as ex:
                err = str(ex)

            def _done():
                self._reconcile_btn.config(
                    state="normal", text="🔄 Reconcile with Trello")
                if err:
                    self._status_lbl.config(text=f"Reconcile error: {err}")
                elif summary:
                    msg = (f"Moved {summary['moved']}, "
                           f"set Type-of-Loss on {summary['loss_type_set']}, "
                           f"skipped {summary['skipped_no_match']} (no Trello match)")
                    pf = summary.get("parse_failures") or []
                    if pf:
                        msg += f"  ·  {len(pf)} cards unparseable (see log)"
                    if summary["errors"]:
                        msg += f"  ·  {len(summary['errors'])} errors"
                    self._status_lbl.config(text=msg)
                self.after(1200, self._refresh)
            self.after(0, _done)

        threading.Thread(target=_bg, daemon=True).start()

    def _cross_check(self):
        """Background scan for spreadsheet-vs-Trello sheet disagreements;
        on completion, render the list in a per-row action dialog."""
        try:
            import state_reconciler
        except Exception as ex:
            messagebox.showerror("Cross-check unavailable",
                                  str(ex), parent=self)
            return
        yr = self._year_var.get()
        self._crosscheck_btn.config(state="disabled",
                                      text="Cross-checking…")
        self._status_lbl.config(text="Comparing every row against Trello…")

        def _bg():
            disagreements = None
            err = None
            try:
                def _prog(i, total, name):
                    try:
                        self.after(0,
                                   lambda i=i, t=total, n=name:
                                   self._status_lbl.config(
                                       text=f"Cross-check {i}/{t}: {n}"))
                    except Exception:
                        pass
                disagreements = state_reconciler.find_disagreements(
                    yr, progress_cb=_prog)
            except Exception as ex:
                err = str(ex)

            def _done():
                self._crosscheck_btn.config(
                    state="normal", text="🔍 Cross-check Trello")
                if err:
                    self._status_lbl.config(
                        text=f"Cross-check error: {err}")
                    return
                if not disagreements:
                    self._status_lbl.config(
                        text="Cross-check — every row matches Trello.")
                    return
                self._status_lbl.config(
                    text=f"Cross-check — {len(disagreements)} "
                         f"disagreement"
                         f"{'s' if len(disagreements) != 1 else ''}.")
                self._open_disagreement_dialog(disagreements, yr)
            self.after(0, _done)

        threading.Thread(target=_bg, daemon=True).start()

    def _open_disagreement_dialog(self, disagreements, year):
        """Modal list of cross-check disagreements. Each row gets a
        "Move to <suggested>" button, an "Open card" link, and a
        "Snooze" — the move action calls into snapshots_excel's
        existing _move_row_between_sheets, so all the cell-preservation
        rules already proven by the reconciler apply unchanged."""
        try:
            import state_reconciler
            import snapshots_excel as sx
            import webbrowser
        except Exception:
            return
        dlg = tk.Toplevel(self)
        dlg.title(f"Cross-check disagreements — {year}")
        dlg.transient(self.winfo_toplevel())
        dlg.geometry("780x520")
        wrap = tk.Frame(dlg, bg=BG, padx=14, pady=12)
        wrap.pack(fill="both", expand=True)
        tk.Label(wrap,
                 text=(f"{len(disagreements)} row"
                       f"{'s' if len(disagreements) != 1 else ''} disagree "
                       f"with Trello state."),
                 font=("Segoe UI Variable", 10, "bold"),
                 bg=BG, fg=TEXT_DARK,
                 anchor="w").pack(fill="x")
        tk.Label(wrap,
                 text=("Per-row: \"Move to …\" relocates the workbook "
                       "row; \"Open card\" jumps to the Trello card so "
                       "you can verify."),
                 font=("Segoe UI Variable", 8, "italic"),
                 bg=BG, fg=TEXT_GRAY,
                 anchor="w").pack(fill="x", pady=(0, 8))

        from tool_panel import ScrollableFrame
        scroll = ScrollableFrame(wrap, bg=WHITE)
        scroll.pack(fill="both", expand=True)
        inner = scroll.inner

        def _open_card(url):
            if url:
                try:
                    webbrowser.open(url)
                except Exception:
                    pass

        def _move_row(d, target_route, row_frame):
            target_sheet = state_reconciler.route_to_sheet(target_route)
            if not target_sheet:
                return
            try:
                ok = sx.move_existing_row(d["name"], target_sheet,
                                            year=year)
            except Exception as ex:
                messagebox.showerror("Move failed", str(ex), parent=dlg)
                return
            if ok:
                try:
                    row_frame.destroy()
                except tk.TclError:
                    pass
                self.after(800, self._refresh)
            else:
                messagebox.showinfo(
                    "No-op",
                    f"Could not locate {d['name']} in any sheet "
                    f"(it may have been moved already).",
                    parent=dlg)

        for d in disagreements:
            row = tk.Frame(inner, bg=WHITE, padx=10, pady=6,
                            highlightbackground=BORDER if False else "#DDDDDD",
                            highlightthickness=1)
            row.pack(fill="x", padx=4, pady=3)
            top = tk.Frame(row, bg=WHITE)
            top.pack(fill="x")
            tk.Label(top, text=d["name"],
                     font=("Segoe UI Variable", 10, "bold"),
                     bg=WHITE, fg=TEXT_DARK, anchor="w"
                     ).pack(side="left")
            ctx = " · ".join(x for x in (d["carrier"], d["claim"]) if x)
            if ctx:
                tk.Label(top, text=f"  ({ctx})",
                         font=("Segoe UI Variable", 8), bg=WHITE, fg=TEXT_GRAY,
                         anchor="w").pack(side="left")
            tk.Label(row,
                     text=(f"Sheet says: {d['sheet']}    →    "
                           f"Trello suggests: "
                           f"{state_reconciler.route_to_sheet(d['trello_route']) or d['trello_route']}"),
                     font=("Segoe UI Variable", 9),
                     bg=WHITE, fg=WARN_FG,
                     anchor="w").pack(fill="x", pady=(2, 4))
            btn_row = tk.Frame(row, bg=WHITE)
            btn_row.pack(fill="x")
            target_name = state_reconciler.route_to_sheet(d["trello_route"])
            link_button(btn_row,
                         f"➡ Move to {target_name or d['trello_route']}",
                         padx=10, pady=2,
                         command=lambda d=d, t=d["trello_route"], rf=row:
                                 _move_row(d, t, rf),
                         tooltip="Move this row to the Trello-suggested sheet."
                         ).pack(side="left")
            if d.get("url"):
                link_button(btn_row, "🔗 Open card",
                             padx=10, pady=2,
                             command=lambda u=d["url"]: _open_card(u),
                             tooltip="Open this card in Trello."
                             ).pack(side="left", padx=(6, 0))
            secondary_button(btn_row, "✓ Snooze",
                              padx=10, pady=2,
                              command=lambda rf=row: rf.destroy(),
                              tooltip=("Dismiss this disagreement for the "
                                       "current session (re-run cross-check "
                                       "to surface it again).")
                              ).pack(side="left", padx=(6, 0))

        bot = tk.Frame(wrap, bg=BG)
        bot.pack(fill="x", pady=(10, 0))
        secondary_button(bot, "Close", padx=12, pady=3,
                          command=dlg.destroy).pack(side="right")
        dlg.update_idletasks()
        try:
            px = (self.winfo_rootx()
                  + max(0, (self.winfo_width() - 780) // 2))
            py = (self.winfo_rooty()
                  + max(0, (self.winfo_height() - 520) // 2))
            dlg.geometry(f"+{px}+{py}")
        except Exception:
            pass

    def _generate_notes(self):
        """Scan every row in the current view and build a draft Comment
        for each one that's empty. Backgrounds the work (Trello fetches
        are slow) then opens a confirm-each dialog so the user reviews
        + edits + approves before anything writes to the workbook."""
        try:
            import notes_generator
        except Exception as ex:
            messagebox.showerror("Notes generator unavailable",
                                  str(ex), parent=self)
            return
        yr = self._year_var.get()
        # Subject pool = currently-visible rows whose Comment cell is
        # empty. Keeps the dialog small + respects whatever sheet /
        # search filter the user has applied.
        rows = self._cache.get("rows") or []
        sheet_filter = self._sheet_var.get()
        text_filter = (self._filter_var.get() or "").strip().lower()
        subjects = []
        for r in rows:
            if sheet_filter != "All" and r.get("_sheet") != sheet_filter:
                continue
            name = str(r.get("Name") or "").strip()
            if not name:
                continue
            comment = str(r.get("Comment") or "").strip()
            if comment:
                continue
            if text_filter:
                blob = (name + " " +
                         str(r.get("Carrier") or "") + " " +
                         str(r.get("Claim#") or "")).lower()
                if text_filter not in blob:
                    continue
            subjects.append(r)
        if not subjects:
            messagebox.showinfo(
                "Nothing to generate",
                "Every visible row already has a Comment.\n\n"
                "Tip: clear a row's Comment in Excel if you want this "
                "tool to regenerate it, OR widen the sheet/search "
                "filter to include more rows.", parent=self)
            return

        self._notes_btn.config(state="disabled",
                                 text="Generating…")
        self._status_lbl.config(
            text=f"Generating drafts for {len(subjects)} rows…")

        def _bg():
            drafts = []
            for i, r in enumerate(subjects, 1):
                try:
                    self.after(0, lambda i=i, t=len(subjects), n=str(r.get("Name") or ""):
                               self._status_lbl.config(
                                   text=f"Draft {i}/{t}: {n}"))
                except Exception:
                    pass
                try:
                    text = notes_generator.generate_summary_for_row(r)
                except Exception:
                    text = ""
                drafts.append({"row": r, "text": text})

            def _done():
                self._notes_btn.config(state="normal",
                                        text="📝 Generate notes")
                if not drafts:
                    self._status_lbl.config(
                        text="No drafts produced.")
                    return
                self._status_lbl.config(
                    text=f"{len(drafts)} drafts ready — review + confirm.")
                self._open_notes_confirm_dialog(drafts, yr)
            self.after(0, _done)

        threading.Thread(target=_bg, daemon=True).start()

    def _open_notes_confirm_dialog(self, drafts, year):
        """Scrolling dialog with one editable card per draft. User can:
          • Edit the text in place
          • Click ✓ Save to write that row's Comment cell now
          • Click Skip to dismiss without writing
          • Click Save all to apply every visible draft in one shot.
        Each save happens through sx.set_comment which honors the
        workbook lock + canonical name matching."""
        try:
            import snapshots_excel as sx
        except Exception:
            return
        dlg = tk.Toplevel(self)
        dlg.title(f"Confirm generated notes — {year}")
        dlg.transient(self.winfo_toplevel())
        dlg.geometry("820x620")
        wrap = tk.Frame(dlg, bg=BG, padx=14, pady=12)
        wrap.pack(fill="both", expand=True)
        tk.Label(wrap,
                 text=(f"{len(drafts)} draft "
                       f"note{'s' if len(drafts) != 1 else ''} — "
                       "edit if needed, then save each to that row's "
                       "Comment cell. Nothing writes until you click "
                       "Save."),
                 font=("Segoe UI Variable", 10, "bold"),
                 bg=BG, fg=TEXT_DARK,
                 anchor="w").pack(fill="x")
        tk.Label(wrap,
                 text=("Save all at the bottom applies every draft "
                       "below to the workbook. Skip removes the row "
                       "from this list without writing."),
                 font=("Segoe UI Variable", 8, "italic"),
                 bg=BG, fg=TEXT_GRAY,
                 anchor="w").pack(fill="x", pady=(0, 8))

        from tool_panel import ScrollableFrame
        scroll = ScrollableFrame(wrap, bg=WHITE)
        scroll.pack(fill="both", expand=True)
        inner = scroll.inner

        # Track every row's widgets so Save all can iterate.
        editor_states = []

        def _save_one(state):
            text = state["text_widget"].get("1.0", "end").strip()
            client = (state["row"].get("Name") or "").strip()
            if not client:
                return False
            try:
                ok = sx.set_comment(client, text, year=year)
            except Exception as ex:
                messagebox.showerror("Save failed", str(ex), parent=dlg)
                return False
            if not ok:
                messagebox.showinfo(
                    "Not saved",
                    f"Couldn't find {client} in the workbook (or the "
                    "workbook is locked). Try refreshing or close "
                    "Excel.", parent=dlg)
                return False
            try:
                state["wrap"].destroy()
            except tk.TclError:
                pass
            return True

        def _save_all():
            saved = 0
            for state in list(editor_states):
                if state.get("done"):
                    continue
                if _save_one(state):
                    state["done"] = True
                    saved += 1
            if saved:
                self.after(800, self._refresh)
                self._status_lbl.config(
                    text=f"Saved {saved} note"
                         f"{'s' if saved != 1 else ''} to workbook.")
            if not any(not s.get("done") for s in editor_states):
                dlg.destroy()

        def _skip(state):
            try:
                state["wrap"].destroy()
            except tk.TclError:
                pass
            state["done"] = True
            # Close the dialog when every card has been resolved.
            if not any(not s.get("done") for s in editor_states):
                dlg.destroy()

        for d in drafts:
            row = d["row"]
            text = d["text"]
            row_wrap = tk.Frame(inner, bg=WHITE, padx=10, pady=8,
                                 highlightbackground="#DDDDDD",
                                 highlightthickness=1)
            row_wrap.pack(fill="x", padx=4, pady=3)
            top = tk.Frame(row_wrap, bg=WHITE)
            top.pack(fill="x")
            tk.Label(top, text=row.get("Name") or "(unnamed)",
                     font=("Segoe UI Variable", 11, "bold"),
                     bg=WHITE, fg=TEXT_DARK,
                     anchor="w").pack(side="left")
            ctx_bits = [b for b in (
                row.get("_sheet") or "",
                row.get("Carrier") or "",
                row.get("Claim#") or "",
            ) if b]
            if ctx_bits:
                tk.Label(top, text="   " + "  ·  ".join(ctx_bits),
                         font=("Segoe UI Variable", 8, "italic"),
                         bg=WHITE, fg=TEXT_GRAY,
                         anchor="w").pack(side="left")

            txt = tk.Text(row_wrap, wrap="word",
                           font=("Segoe UI Variable", 9), height=6,
                           bg=WHITE, relief="solid", bd=1)
            txt.insert("1.0", text or "")
            txt.pack(fill="x", pady=(6, 6))

            state = {"row": row, "text_widget": txt,
                     "wrap": row_wrap, "done": False}
            editor_states.append(state)

            btn_row = tk.Frame(row_wrap, bg=WHITE)
            btn_row.pack(fill="x")
            secondary_button(
                btn_row, "Skip", padx=12, pady=2,
                command=lambda s=state: _skip(s),
                tooltip="Dismiss without writing"
                ).pack(side="right", padx=(6, 0))
            send_button(
                btn_row, "✓ Save", padx=12, pady=2,
                command=lambda s=state: _save_one(s) and (
                    s.update({"done": True}) or None),
                tooltip="Write this draft to the row's Comment cell"
                ).pack(side="right")

        bot = tk.Frame(wrap, bg=BG)
        bot.pack(fill="x", pady=(10, 0))
        secondary_button(bot, "Close", padx=12, pady=4,
                          command=dlg.destroy
                          ).pack(side="right", padx=(8, 0))
        send_button(
            bot, "✓ Save all visible", padx=14, pady=4,
            command=_save_all,
            tooltip=("Apply every draft above to its Comment cell. "
                     "Excel must be closed.")
            ).pack(side="right")
        dlg.update_idletasks()
        try:
            px = (self.winfo_rootx()
                  + max(0, (self.winfo_width() - 820) // 2))
            py = (self.winfo_rooty()
                  + max(0, (self.winfo_height() - 620) // 2))
            dlg.geometry(f"+{px}+{py}")
        except Exception:
            pass

    def _open_daily_digest(self):
        """Generate + open today's digest .md. Falls back to a messagebox
        when the digest root (X:\\IE_Public\\Daily Digest) isn't reachable
        — usually a disconnected X: drive."""
        try:
            import daily_digest
        except Exception as ex:
            messagebox.showerror("Digest unavailable", str(ex), parent=self)
            return
        path = daily_digest.open_today_digest()
        if not path:
            messagebox.showinfo(
                "Digest",
                "Couldn't write today's digest — is X: reachable?\n"
                "Target folder: " + daily_digest.get_root(),
                parent=self)
            return
        self._status_lbl.config(text=f"Digest written: {path}")

    def _dedupe(self):
        """Two-step dedupe: dry-run preview → confirmation → apply.
        The preview matters because the merge collapses two rows into
        one and there's no version history on the server share — the
        user needs to eyeball the keeper picks before we delete."""
        yr = self._year_var.get()
        self._dedupe_btn.config(state="disabled", text="Scanning…")
        self._status_lbl.config(text="Scanning workbook for duplicate Name rows…")

        def _bg():
            preview = err = None
            try:
                preview = sx.dedupe_workbook(yr, dry_run=True)
            except Exception as ex:
                err = str(ex)

            def _after_preview():
                self._dedupe_btn.config(
                    state="normal", text="🧹 Dedupe rows")
                if err:
                    self._status_lbl.config(text=f"Dedupe preview error: {err}")
                    return
                if preview["errors"]:
                    self._status_lbl.config(
                        text="Dedupe preview: " + " · ".join(preview["errors"]))
                    return
                groups = preview["groups"]
                if not groups:
                    self._status_lbl.config(
                        text="Dedupe — no duplicates found.")
                    return

                # Build a compact summary of what will happen.
                lines = []
                for g in groups[:20]:
                    keeper = f"{g['keeper'][2]} ({g['keeper'][0]})"
                    others = ", ".join(f"{m[2]}" for m in g['merged'])
                    lines.append(f"  • Keep: {keeper}  ←  drop: {others}")
                if len(groups) > 20:
                    lines.append(f"  …and {len(groups) - 20} more groups")
                summary_text = (
                    f"Found {len(groups)} duplicate group"
                    f"{'s' if len(groups) != 1 else ''} — "
                    f"{preview['rows_removed']} row"
                    f"{'s' if preview['rows_removed'] != 1 else ''} "
                    f"will be merged into keepers:\n\n"
                    + "\n".join(lines)
                    + "\n\nApply now? (Excel must be CLOSED.)")
                if not messagebox.askyesno(
                        "Dedupe duplicate rows?",
                        summary_text, parent=self):
                    self._status_lbl.config(
                        text=f"Dedupe — preview shown, not applied "
                             f"({len(groups)} groups).")
                    return
                self._apply_dedupe(yr)

            self.after(0, _after_preview)

        threading.Thread(target=_bg, daemon=True).start()

    def _apply_dedupe(self, year):
        self._dedupe_btn.config(state="disabled", text="Deduping…")
        self._status_lbl.config(text="Merging duplicate rows…")

        def _bg():
            result = err = None
            try:
                result = sx.dedupe_workbook(year, dry_run=False)
            except Exception as ex:
                err = str(ex)

            def _done():
                self._dedupe_btn.config(
                    state="normal", text="🧹 Dedupe rows")
                if err:
                    self._status_lbl.config(text=f"Dedupe error: {err}")
                elif result and result["errors"]:
                    self._status_lbl.config(
                        text="Dedupe: " + " · ".join(result["errors"]))
                elif result:
                    self._status_lbl.config(
                        text=f"Dedupe — merged {result['rows_removed']} "
                             f"duplicate row"
                             f"{'s' if result['rows_removed'] != 1 else ''} "
                             f"into {len(result['groups'])} keeper"
                             f"{'s' if len(result['groups']) != 1 else ''}.")
                self.after(800, self._refresh)
            self.after(0, _done)

        threading.Thread(target=_bg, daemon=True).start()








def _snapshots_search_blob(r):
    return " ".join((
        str(r.get("Name") or ""),
        str(r.get("Carrier") or ""),
        str(r.get("Claim#") or ""),
        str(r.get("Comment") or ""),
    ))




# Register at import. Future workbooks: add another `wbr.register(...)`
# call from their own module (the launcher only needs to import that
# module once before the panel is built — easiest path is to import it
# from this file too).


# ── Disputes workbook spec ──────────────────────────────────────────────────
# Imports dispute_tracker lazily inside the helpers so the Spreadsheet
# panel still builds even when the dispute workbook is on a path we
# can't reach right now (server share offline, etc).











def _open_dispute_tracker_panel(app):
    """Action: open the editable Dispute Tracker panel. The Spreadsheet
    panel is read-only; routing to the dedicated panel keeps the
    treeview-vs-editor split clean.

    Uses `navigate_to` (via ToolPanel) so launcher-hosted Spreadsheets
    swaps panels in-place; only standalone Spreadsheets falls through
    to a spawn. The previous direct `spawn_tool` call produced a
    duplicate Dispute Tracker window on every click."""
    def _run():
        try:
            app.navigate_to("dispute_tracker")
        except Exception:
            try:
                from tkinter import messagebox
                messagebox.showerror(
                    "Couldn't open Dispute Tracker",
                    "Run from the launcher's Disputes button instead.",
                    parent=app)
            except Exception:
                pass
    return _run


def _refresh_disputes(app):
    """Action: re-read the workbook + redraw the table. Useful after
    making edits in the dedicated panel — the Spreadsheet panel caches
    rows in memory and won't pick up external edits otherwise."""
    return lambda: app._on_workbook_changed()





# ── User-defined workbook auto-registration ─────────────────────────────────
# Importing user_workbooks runs its `register_all()` side-effect, which
# walks `config["user_workbooks"]` and registers each entry as a
# WorkbookSpec via `wbr.register(...)`. Done after the hard-coded specs
# so the built-ins always render first in the dropdown.
# ── Workbook specs ────────────────────────────────────────────────────
# The specs themselves live in workbook_specs, which is Tk-free, so the
# WEB panel can populate the registry without importing this module.
# Importing it here registers both; we then re-register Disputes with
# the Tk-only action buttons attached. wbr.register overrides by key and
# keeps the original order, so each panel gets the spec it can use.
import dataclasses as _dc

import workbook_specs as _wbs
from workbook_specs import (  # noqa: F401 — kept importable for callers
    SNAPSHOTS_COLUMNS, SNAPSHOTS_SPEC, DISPUTES_COLUMNS, DISPUTES_SPEC,
    _fmt_date_for_row,
)

wbr.register(_dc.replace(_wbs.DISPUTES_SPEC, actions=(
    wbr.ActionSpec(
        label="✎ Open Disputes panel",
        kind="send",
        command_factory=_open_dispute_tracker_panel,
        attr_name="_open_disputes_btn",
        tooltip=("Open the dedicated Dispute Tracker panel where "
                 "rows are editable (statuses, dates, notes). "
                 "This Spreadsheets viewer is read-only.")),
    wbr.ActionSpec(
        label="↻ Refresh",
        kind="secondary",
        command_factory=_refresh_disputes,
        attr_name="_refresh_disputes_btn",
        tooltip=("Re-read the Disputes workbook from disk so edits "
                 "made in the dedicated panel show up here.")),
)))

import user_workbooks as uwb       # noqa: E402  (intentional post-spec import)


def main(argv=None):
    import ctypes
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
        "Servpro.EMS.Spreadsheet")
    # No `title=` — run_standalone has never accepted one, so this raised
    # TypeError before the window could appear. The panel titles itself,
    # and every other *_gui calls it the same way.
    run_standalone(SpreadsheetApp, geometry="1200x720")


if __name__ == "__main__":
    main()
