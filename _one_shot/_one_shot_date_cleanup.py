"""One-shot Date Received + Closing Date cleanup.

Walks every spreadsheet row, looks up the linked Trello card, and:
  - Sets Date Received to the card desc's PROPERTY DETAILS > DATE
    RECEIVED when the desc has one and it differs from the cell.
  - Clears Closing Date when the card has neither a LOGS - EMS move
    nor an archive action (i.e., the card isn't actually closed).
  - Sets Closing Date to the Trello-derived value when the card IS
    closed and the cell differs.

Run with --apply to write changes; otherwise dry-runs and prints a
diff. The Excel workbook MUST be closed (we open it for write).

Not part of the launcher — single-use cleanup for the existing
spreadsheet that has random folder-prefix dates carved into it.
"""
import sys
from datetime import datetime
from openpyxl import load_workbook

import snapshots_excel as sx
import trello_client as tc
import persistence as p


APPLY = "--apply" in sys.argv
YEAR  = datetime.today().year
PATH  = sx.workbook_path(YEAR)


def _resolve_card_id(name):
    pinned = p.get_trello_card_ids(name) or []
    if pinned:
        return pinned[0]
    for term in p.client_search_terms(name):
        try:
            hits = tc.find_cards_by_name(term, max_results=1)
        except Exception:
            hits = []
        if hits:
            return hits[0]["card_id"]
    return None


def main():
    print(f"Workbook: {PATH}")
    print(f"Mode: {'APPLY (writes)' if APPLY else 'DRY-RUN (no writes)'}")
    print()

    wb = load_workbook(PATH)
    plan = []   # (sheet, name, col, old, new)
    no_card = []
    err = []

    # Pre-count rows so we can show progress N/total. max_row is cheap.
    total_rows = 0
    for base in sx._ALL_SHEETS:
        t = sx._sheet_name(base, YEAR)
        if t in wb.sheetnames:
            total_rows += max(0, wb[t].max_row - 1)
    print(f"Scanning {total_rows} rows across {len(sx._ALL_SHEETS)} sheets…",
          flush=True)
    seen = 0

    for sheet_base in sx._ALL_SHEETS:
        title = sx._sheet_name(sheet_base, YEAR)
        if title not in wb.sheetnames:
            continue
        ws = wb[title]
        # Header is row 1; data starts row 2 (matches sx layout).
        col_name     = sx._COL_INDEX["Name"]
        col_received = sx._COL_INDEX["Date Received"]
        col_closing  = sx._COL_INDEX["Closing Date"]
        for rr in range(2, ws.max_row + 1):
            name = ws.cell(rr, col_name).value
            if not name:
                continue
            name = str(name).strip()
            seen += 1
            if seen % 10 == 0 or seen == 1:
                print(f"  [{seen}/{total_rows}] {sheet_base[:11]:11} {name[:40]}",
                      flush=True)
            cur_received = ws.cell(rr, col_received).value
            cur_closing  = ws.cell(rr, col_closing).value
            try:
                card_id = _resolve_card_id(name)
            except Exception as ex:
                err.append((sheet_base, name, f"resolve: {ex}"))
                continue
            if not card_id:
                no_card.append((sheet_base, name, cur_received, cur_closing))
                # Garbage-timestamp heuristic: real dates parsed from
                # text always have hour=minute=second=microsecond=0.
                # Anything with a microsecond came from `datetime.today()`
                # via the now-removed _date_received_for fallback. Safe
                # to wipe even without a Trello card to confirm.
                if isinstance(cur_received, datetime) and cur_received.microsecond:
                    plan.append((sheet_base, rr, name, "Date Received",
                                 cur_received, None))
                if isinstance(cur_closing, datetime) and cur_closing.microsecond:
                    plan.append((sheet_base, rr, name, "Closing Date",
                                 cur_closing, None))
                continue
            try:
                card = tc.get_card(card_id, actions_limit=10)
            except Exception as ex:
                err.append((sheet_base, name, f"get_card: {ex}"))
                continue
            # Date Received: prefer desc value; if desc is blank but
            # the cell holds the known `datetime.today()` garbage
            # (hour=14 minute=41 second=28 from a single batch run),
            # fall back to card creation date — anything is better
            # than leaving the proven-wrong stamp.
            try:
                desc_received = tc.card_received_date(card) if card else None
            except Exception:
                desc_received = None
            cell_is_today_garbage = (
                isinstance(cur_received, datetime)
                and cur_received.hour == 14
                and cur_received.minute == 41
                and cur_received.second == 28)
            if desc_received is not None and desc_received != cur_received:
                plan.append((sheet_base, rr, name, "Date Received",
                             cur_received, desc_received))
            elif desc_received is None and cell_is_today_garbage:
                try:
                    creation = tc.card_creation_date(card_id)
                except Exception:
                    creation = None
                # If we can't get a creation date, just clear the
                # garbage — blank is better than wrong.
                plan.append((sheet_base, rr, name, "Date Received",
                             cur_received, creation))
            # Closing date.
            try:
                tc_closing = tc.card_closing_date(card_id)
            except Exception:
                tc_closing = None
            if tc_closing is None and cur_closing not in (None, ""):
                plan.append((sheet_base, rr, name, "Closing Date",
                             cur_closing, None))
            elif tc_closing is not None and tc_closing != cur_closing:
                plan.append((sheet_base, rr, name, "Closing Date",
                             cur_closing, tc_closing))

    # Report.
    print(f"Rows to update: {len(plan)}")
    print(f"Rows with no Trello card matched: {len(no_card)}")
    print(f"Errors: {len(err)}")
    print()
    print("--- Planned changes ---")
    for sheet, rr, name, col, old, new in plan:
        old_s = old.strftime("%m/%d/%y") if isinstance(old, datetime) else (
            "<blank>" if old in (None, "") else str(old))
        new_s = new.strftime("%m/%d/%y") if isinstance(new, datetime) else (
            "<blank>" if new in (None, "") else str(new))
        print(f"  [{sheet:11}] r{rr:>4} {name:35} {col:14} {old_s:>10}  →  {new_s}")

    if no_card:
        print()
        print("--- No Trello match (untouched) ---")
        for sheet, name, cr, cc in no_card[:30]:
            print(f"  [{sheet:11}] {name:35}  recv={cr}  close={cc}")
        if len(no_card) > 30:
            print(f"  ... and {len(no_card)-30} more")

    if err:
        print()
        print("--- Errors ---")
        for sheet, name, msg in err[:20]:
            print(f"  [{sheet:11}] {name}: {msg}")

    if not APPLY:
        print()
        print("Dry-run only. Re-run with --apply to write changes.")
        return

    # Apply.
    # NOTE: ws.cell(row, col, value=None) treats None as "don't set" —
    # it just returns the cell. Use `.value = new` instead so clears
    # actually clear.
    for sheet, rr, name, col, _old, new in plan:
        ws = wb[sx._sheet_name(sheet, YEAR)]
        col_idx = sx._COL_INDEX[col]
        ws.cell(rr, col_idx).value = new
    try:
        wb.save(PATH)
        print(f"\nApplied {len(plan)} changes to {PATH}")
    except PermissionError:
        print(f"\nERROR: workbook is locked (close Excel and rerun).")
        sys.exit(1)


if __name__ == "__main__":
    main()
